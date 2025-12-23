"""Shared fixtures for imports tests."""

import csv
import io
import pytest
from pathlib import Path
from uuid import uuid4

import openpyxl


@pytest.fixture
def sample_products_excel(tmp_path) -> Path:
    """Create a sample products Excel file."""
    file_path = tmp_path / "sample_products.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["nombre", "precio", "stock", "categoria", "sku", "descripcion"]
    ws.append(headers)

    products = [
        ["Laptop HP", 999.99, 10, "Electrónica", "LAP-001", "Laptop 15 pulgadas"],
        ["Mouse Logitech", 29.99, 50, "Accesorios", "MOU-001", "Mouse inalámbrico"],
        ["Teclado Mecánico", 79.99, 25, "Accesorios", "TEC-001", "Teclado RGB"],
        ["Monitor Samsung", 299.99, 15, "Monitores", "MON-001", "Monitor 24 pulgadas"],
        ["Auriculares Sony", 149.99, 30, "Audio", "AUR-001", "Auriculares Bluetooth"],
    ]
    for product in products:
        ws.append(product)

    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_products_excel_bytes() -> bytes:
    """Create sample products Excel as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nombre", "precio", "stock", "categoria", "sku"])
    ws.append(["Laptop", 999.99, 10, "Electrónica", "LAP-001"])
    ws.append(["Mouse", 29.99, 50, "Accesorios", "MOU-001"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@pytest.fixture
def sample_bank_csv(tmp_path) -> Path:
    """Create a sample bank CSV file."""
    file_path = tmp_path / "sample_bank.csv"
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["fecha", "concepto", "importe", "saldo", "referencia"])
        writer.writerow(
            ["2024-01-15", "Transferencia recibida", 1500.00, 5000.00, "TRF-001"]
        )
        writer.writerow(["2024-01-16", "Pago servicios", -200.00, 4800.00, "PAG-001"])
        writer.writerow(["2024-01-17", "Depósito", 500.00, 5300.00, "DEP-001"])
        writer.writerow(
            ["2024-01-18", "Retiro cajero", -100.00, 5200.00, "RET-001"]
        )
    return file_path


@pytest.fixture
def sample_bank_csv_bytes() -> bytes:
    """Create sample bank CSV as bytes."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["fecha", "concepto", "importe", "saldo"])
    writer.writerow(["2024-01-15", "Transferencia", 1000.00, 5000.00])
    writer.writerow(["2024-01-16", "Pago", -200.00, 4800.00])
    return buffer.getvalue().encode("utf-8")


@pytest.fixture
def sample_invoice_excel(tmp_path) -> Path:
    """Create a sample invoice Excel file."""
    file_path = tmp_path / "sample_invoices.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    ws.append(["factura", "cliente", "ruc", "fecha", "subtotal", "iva", "total"])
    ws.append(["F-001", "Empresa ABC", "12345678901", "2024-01-20", 1000.00, 180.00, 1180.00])
    ws.append(["F-002", "Cliente XYZ", "98765432109", "2024-01-21", 500.00, 90.00, 590.00])

    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_expenses_excel(tmp_path) -> Path:
    """Create a sample expenses Excel file."""
    file_path = tmp_path / "sample_expenses.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    ws.append(["gasto", "categoria", "monto", "fecha", "comprobante"])
    ws.append(["Suministros oficina", "Oficina", 150.00, "2024-01-15", "REC-001"])
    ws.append(["Servicios luz", "Servicios", 200.00, "2024-01-16", "FAC-001"])

    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_ambiguous_file(tmp_path) -> Path:
    """Create an ambiguous file with unknown headers."""
    file_path = tmp_path / "ambiguous.csv"
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["columna1", "columna2", "columna3"])
        writer.writerow(["valor1", "valor2", "valor3"])
    return file_path


@pytest.fixture
def mock_tenant_id() -> str:
    """Generate a mock tenant ID."""
    return str(uuid4())


@pytest.fixture
def mock_access_claims(mock_tenant_id):
    """Generate mock access claims."""
    return {
        "tenant_id": mock_tenant_id,
        "user_id": str(uuid4()),
        "roles": ["admin"],
    }
