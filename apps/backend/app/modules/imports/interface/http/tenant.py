from __future__ import annotations

import json
import logging
import os
import re
import time
import unicodedata
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from pydantic import BaseModel

# Avoid importing heavy domain models at import time to keep router mountable in test envs
# (domain handlers perform promotion separately)
from sqlalchemy import func
from sqlalchemy import inspect as sa_inspect
from sqlalchemy import select, text
from sqlalchemy.orm import Session

from app.config.database import get_db
from app.core.access_guard import with_access_claims
from app.core.authz import require_scope
from app.db.rls import ensure_rls, tenant_id_sql_expr
from app.models.core.modelsimport import (
    ImportAttachment,
    ImportBatch,
    ImportItem,
    ImportItemCorrection,
    ImportLineage,
    ImportMapping,
    ImportOCRJob,
)
from app.models.core.products import Product
from app.models.core.ui_field_config import TenantFieldConfig
from app.models.imports import ImportColumnMapping
from app.models.inventory.stock import StockItem, StockMove
from app.models.inventory.warehouse import Warehouse
from app.models.purchases.purchase import Purchase, PurchaseLine
from app.modules.imports.application.job_runner import enqueue_job
from app.modules.imports.parsers import registry as parser_registry
from app.modules.imports.parsers.dispatcher import (
    is_parser_compatible_with_extension,
    select_fallback_parser_for_extension,
    select_parser_for_file,
)
from app.modules.imports.schemas import (
    BatchCreate,
    BatchOut,
    ImportMappingCreate,
    ImportMappingOut,
    ImportMappingUpdate,
    IngestRows,
    ItemOut,
    ItemPatch,
    OCRJobEnqueuedResponse,
    OCRJobStatusResponse,
    OkResponse,
    PromoteItems,
    UpdateClassificationRequest,
)
from app.services.excel_analyzer import analyze_excel_stream, detect_header_row, extract_headers
from app.services.inventory_costing import InventoryCostingService

logger = logging.getLogger("imports")


def _get_claims(request: Request) -> dict:
    """Helper to extract and validate access claims from request."""
    claims = getattr(request.state, "access_claims", None)
    if not isinstance(claims, dict):
        claims = with_access_claims(request)
    return claims


# Public router (no auth) for lightweight health checks
public_router = APIRouter(
    prefix="/imports",
    tags=["Imports"],
)

router = APIRouter(
    prefix="/imports",
    tags=["Imports"],
    dependencies=[
        Depends(with_access_claims),
        Depends(require_scope("tenant")),
        Depends(ensure_rls),
    ],
)

# --- simple in-memory throttling (per-tenant) --------------------------------
_INGEST_WINDOW_SEC = 60
_DEFAULT_INGEST_LIMIT = int(os.getenv("IMPORTS_MAX_INGESTS_PER_MIN", "500"))
_ingest_rate_state: dict[str, list[float]] = {}


def _throttle_ingest(tenant_id: str | int):
    key = str(tenant_id)
    now = time.time()
    window_start = now - _INGEST_WINDOW_SEC
    buf = _ingest_rate_state.get(key, [])
    buf = [t for t in buf if t >= window_start]
    # Re-read limit from env to allow test overrides
    try:
        limit = int(os.getenv("IMPORTS_MAX_INGESTS_PER_MIN", str(_DEFAULT_INGEST_LIMIT)))
    except Exception:
        limit = _DEFAULT_INGEST_LIMIT
    if len(buf) >= limit:
        raise HTTPException(status_code=429, detail="Too many ingests; try later")
    buf.append(now)
    _ingest_rate_state[key] = buf


def _file_path_from_key(file_key: str) -> str:
    """Resolver ruta local a partir de file_key almacenado en ImportBatch."""
    if file_key.startswith("imports/"):
        return os.path.join("uploads", file_key.replace("/", os.sep))
    return file_key


def _load_original_filename_from_file_key(file_key: str) -> str | None:
    """Best-effort load original filename from a sidecar meta file."""
    try:
        file_path = _file_path_from_key(file_key)
        meta_path = f"{file_path}.meta.json"
        if os.path.exists(meta_path):
            with open(meta_path, encoding="utf-8") as f:
                meta = json.load(f)
                filename = str(meta.get("filename") or "").strip()
                return filename or None
    except Exception:
        return None
    return None


def _header_similarity_score(headers: list[str] | None, mapping: dict | None) -> tuple[float, int]:
    """Compute similarity between detected headers and a saved mapping.

    Supports both mapping orientations:
    - source->target (expected by backend pipeline)
    - target->source (legacy/frontend forms)
    """
    if not headers or not isinstance(mapping, dict):
        return (0.0, 0)

    def _norm_token(value: object) -> str:
        if value is None:
            return ""
        s = str(value).strip()
        if not s:
            return ""
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = s.lower()
        s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
        return s

    header_set = {_norm_token(h) for h in headers if _norm_token(h)}
    if not header_set:
        return (0.0, 0)

    key_set = {_norm_token(k) for k in mapping.keys() if _norm_token(k)}
    val_set = {
        _norm_token(v)
        for v in mapping.values()
        if isinstance(v, str) and str(v).strip().lower() != "ignore" and _norm_token(v)
    }

    # Keep best orientation to stay backward compatible with mixed templates.
    best_score = 0.0
    best_overlap = 0
    for candidate in (key_set, val_set):
        if not candidate:
            continue
        overlap = len(header_set.intersection(candidate))
        if overlap <= 0:
            continue
        # Main signal: fraction of mapped columns present in current file.
        coverage = overlap / max(len(candidate), 1)
        # Small bonus to favor richer intersections when coverage ties.
        bonus = overlap / max(len(header_set), 1)
        score = (coverage * 0.85) + (bonus * 0.15)
        if score > best_score:
            best_score = score
            best_overlap = overlap
    return (best_score, best_overlap)


def _extract_headers_from_file_key(file_key: str) -> list[str]:
    try:
        from io import BytesIO

        file_path = _file_path_from_key(file_key)
        with open(file_path, "rb") as fh:
            analysis = analyze_excel_stream(BytesIO(fh.read()))
        headers = analysis.get("headers") or []
        normalized = [str(h) for h in headers if str(h).strip()]
        if normalized:
            return normalized
    except Exception:
        pass
    try:
        # Fallback for legacy .xls or analyzer failures.
        from app.modules.imports.parsers.generic_excel import parse_excel_generic  # type: ignore

        file_path = _file_path_from_key(file_key)
        parsed = parse_excel_generic(file_path)
        headers = parsed.get("headers") or []
        return [str(h) for h in headers if str(h).strip()]
    except Exception:
        return []


def _resolve_best_column_mapping(
    db: Session,
    *,
    tenant_id: str | UUID,
    original_filename: str | None = None,
    headers: list[str] | None = None,
    min_similarity: float = 0.45,
    min_overlap: int = 2,
) -> tuple[ImportColumnMapping | None, str | None, float]:
    """Resolve best mapping by file pattern first, then by header similarity."""
    mappings = (
        db.query(ImportColumnMapping)
        .filter(ImportColumnMapping.tenant_id == tenant_id, ImportColumnMapping.is_active)  # noqa: E712
        .order_by(
            ImportColumnMapping.last_used_at.desc().nullslast(),
            ImportColumnMapping.use_count.desc(),
            ImportColumnMapping.created_at.desc(),
        )
        .all()
    )
    if not mappings:
        return (None, None, 0.0)

    if original_filename:
        for m in mappings:
            try:
                if m.file_pattern and re.search(str(m.file_pattern), str(original_filename), re.I):
                    return (m, "file_pattern", 1.0)
            except Exception:
                continue

    if not headers:
        return (None, None, 0.0)

    best_obj: ImportColumnMapping | None = None
    best_score = 0.0
    best_overlap = 0
    for m in mappings:
        score, overlap = _header_similarity_score(headers, m.mapping or {})
        if score > best_score or (score == best_score and overlap > best_overlap):
            best_obj = m
            best_score = score
            best_overlap = overlap

    if best_obj and best_score >= min_similarity and best_overlap >= min_overlap:
        return (best_obj, "header_similarity", float(best_score))
    return (None, None, 0.0)


# ------------------------------
# Large file uploads (local chunked)
# ------------------------------


class InitChunkUploadDTO(BaseModel):
    filename: str
    content_type: str | None = None
    size: int | None = None
    part_size: int | None = None  # optional client hint in bytes


@router.post("/uploads/chunk/init")
def init_chunk_upload(dto: InitChunkUploadDTO, request: Request):
    """Initialize a local chunked upload session.

    Returns an `upload_id` and recommended `part_size` (<= max upload MB).
    Client must upload parts to PUT /imports/uploads/chunk/{upload_id}/{part_number}
    and then call POST /imports/uploads/chunk/{upload_id}/complete.
    """
    # Allow turning off on-disk storage entirely (e.g. "store only JSON results").
    if os.getenv("IMPORTS_STORE_UPLOAD_FILES", "1").lower() not in ("1", "true", "yes"):
        raise HTTPException(status_code=403, detail="file_storage_disabled")

    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    # Part size must be <= backend upload limit (default 10MB)
    max_mb = float(os.getenv("IMPORTS_MAX_UPLOAD_MB", "10"))
    # Conservative default part size (1MB) to avoid proxy limits
    default_part = 1 * 1024 * 1024
    max_bytes = int(max_mb * 1024 * 1024)
    # Honor client-provided part_size when present, clamp to [64KB, max_bytes]
    requested = 0
    try:
        requested = int(dto.part_size or 0)
    except Exception:
        requested = 0
    if requested > 0:
        part_size = max(64 * 1024, min(requested, max_bytes))
    else:
        part_size = min(default_part, max_bytes)

    upload_id = uuid4().hex
    base_tmp = os.path.join("uploads", "tmp", str(tenant_id), upload_id)
    try:
        os.makedirs(base_tmp, exist_ok=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"init_failed: {e}") from e

    # Save minimal metadata
    try:
        meta = {
            "filename": dto.filename,
            "content_type": dto.content_type or "application/octet-stream",
            "size": int(dto.size or 0),
        }
        with open(os.path.join(base_tmp, "meta.json"), "w", encoding="utf-8") as f:
            import json

            json.dump(meta, f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"meta_write_failed: {e}") from e

    return {
        "provider": "local",
        "upload_id": upload_id,
        "part_size": part_size,
        "max_part_size": max_bytes,
    }


@router.put("/uploads/chunk/{upload_id}/{part_number}")
async def upload_chunk_part(upload_id: str, part_number: int, request: Request):
    """Receive a chunk part (raw body) <= max upload size and store it on disk."""
    if os.getenv("IMPORTS_STORE_UPLOAD_FILES", "1").lower() not in ("1", "true", "yes"):
        raise HTTPException(status_code=403, detail="file_storage_disabled")

    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    max_mb = float(os.getenv("IMPORTS_MAX_UPLOAD_MB", "10"))
    max_bytes = int(max_mb * 1024 * 1024)
    body = await request.body()
    if not body:
        raise HTTPException(status_code=400, detail="empty_body")
    if len(body) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"part_too_large: {len(body)} > {max_bytes}",
        )
    base_tmp = os.path.join("uploads", "tmp", str(tenant_id), upload_id)
    if not os.path.isdir(base_tmp):
        raise HTTPException(status_code=404, detail="upload_not_found")

    try:
        part_path = os.path.join(base_tmp, f"part_{int(part_number):08d}.bin")
        with open(part_path, "wb") as f:
            f.write(body)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"write_failed: {e}") from e

    return {"ok": True, "bytes": len(body)}


class CompleteChunkUploadDTO(BaseModel):
    # Optional explicit total size and expected parts for validation
    expected_parts: int | None = None
    expected_size: int | None = None


@router.post("/uploads/chunk/{upload_id}/complete")
def complete_chunk_upload(upload_id: str, dto: CompleteChunkUploadDTO, request: Request):
    """Merge all parts into final file and return a file_key usable by imports."""
    if os.getenv("IMPORTS_STORE_UPLOAD_FILES", "1").lower() not in ("1", "true", "yes"):
        raise HTTPException(status_code=403, detail="file_storage_disabled")

    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    user_id = claims.get("user_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    import glob
    import shutil
    from uuid import uuid4 as _uuid4

    base_tmp = os.path.join("uploads", "tmp", str(tenant_id), upload_id)
    if not os.path.isdir(base_tmp):
        raise HTTPException(status_code=404, detail="upload_not_found")

    parts = sorted(glob.glob(os.path.join(base_tmp, "part_*.bin")))
    if not parts:
        raise HTTPException(status_code=400, detail="no_parts")
    if dto.expected_parts and len(parts) != int(dto.expected_parts):
        raise HTTPException(status_code=400, detail="parts_mismatch")

    # Prepare destination
    dest_dir = os.path.join("uploads", "imports", str(tenant_id))
    try:
        os.makedirs(dest_dir, exist_ok=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"dest_init_failed: {e}") from e

    # Create unique filename with original extension if available
    meta_path = os.path.join(base_tmp, "meta.json")
    original_ext = ".bin"
    meta = None
    try:
        if os.path.exists(meta_path):
            with open(meta_path, encoding="utf-8") as f:
                meta = json.load(f)
                fn = str(meta.get("filename") or "")
                _, ext = os.path.splitext(fn)
                if ext:
                    original_ext = ext
    except Exception:
        pass

    final_key = f"{_uuid4().hex}{original_ext}"
    dest_path = os.path.join(dest_dir, final_key)

    # Merge parts
    try:
        with open(dest_path, "wb") as out:
            total = 0
            for p in parts:
                with open(p, "rb") as inp:
                    shutil.copyfileobj(inp, out, length=1024 * 1024)
                    total += os.path.getsize(p)
        if dto.expected_size and int(dto.expected_size) != total:
            raise HTTPException(status_code=400, detail="size_mismatch")
        # Persist original filename metadata next to final file (best-effort).
        try:
            if meta:
                with open(f"{dest_path}.meta.json", "w", encoding="utf-8") as mf:
                    json.dump(
                        {
                            "filename": meta.get("filename"),
                            "content_type": meta.get("content_type"),
                            "size": meta.get("size"),
                        },
                        mf,
                    )
        except Exception:
            pass
    except HTTPException:
        # Re-raise validation errors
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"merge_failed: {e}") from e
    finally:
        # Best-effort cleanup of tmp parts
        try:
            shutil.rmtree(base_tmp, ignore_errors=True)
        except Exception:
            pass

    file_key = f"imports/{tenant_id}/{final_key}"
    return {
        "file_key": file_key,
        "bytes": os.path.getsize(dest_path),
        "original_filename": (locals().get("fn") or None),
    }


# ------------------------------
# Batch from uploaded file + start import
# ------------------------------


class CreateBatchFromUploadDTO(BaseModel):
    file_key: str
    source_type: str | None = "products"  # domain-specific; default products
    mapping_id: str | None = None
    original_filename: str | None = None
    parser_id: str | None = None


@router.post("/batches/from-upload", response_model=BatchOut)
def create_batch_from_upload(
    dto: CreateBatchFromUploadDTO, request: Request, db: Session = Depends(get_db)
):
    """Create an ImportBatch from a previously uploaded file_key (local)."""
    from uuid import uuid4 as _uuid4

    from app.models.core.modelsimport import ImportBatch

    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    user_id = claims.get("user_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    original_filename = dto.original_filename or _load_original_filename_from_file_key(dto.file_key)

    # Auto-pick mapping by file_pattern or header similarity if not provided.
    auto_mapping_id = None
    if not dto.mapping_id:
        try:
            headers_for_match = _extract_headers_from_file_key(dto.file_key)
            best_mapping, _, _ = _resolve_best_column_mapping(
                db,
                tenant_id=tenant_id,
                original_filename=original_filename,
                headers=headers_for_match,
            )
            if best_mapping:
                auto_mapping_id = str(best_mapping.id)
        except Exception:
            pass

    # If no mapping matches, try to auto-suggest one from headers and persist it.
    if not dto.mapping_id and not auto_mapping_id:
        try:
            from app.modules.imports.application.transform_suggest import suggest_mapping

            file_path = _file_path_from_key(dto.file_key)
            headers = _extract_headers_from_file_key(dto.file_key)
            if headers:
                mapping, transforms, defaults, _confidence = suggest_mapping(headers)
                if mapping:
                    safe_name = f"auto:{original_filename or dto.file_key}"
                    if len(safe_name) > 120:
                        safe_name = safe_name[:120]
                    pattern = None
                    if original_filename:
                        pattern = re.escape(original_filename) + "$"
                    existing = (
                        db.query(ImportColumnMapping)
                        .filter(
                            ImportColumnMapping.tenant_id == tenant_id,
                            ImportColumnMapping.name == safe_name,
                        )
                        .first()
                    )
                    if existing:
                        existing.mapping = mapping
                        existing.transforms = transforms
                        existing.defaults = defaults
                        if pattern:
                            existing.file_pattern = pattern
                        db.add(existing)
                        db.commit()
                        auto_mapping_id = str(existing.id)
                    else:
                        new_map = ImportColumnMapping(
                            id=_uuid4(),
                            tenant_id=tenant_id,
                            name=safe_name,
                            description=f"Auto mapping for {original_filename or dto.file_key}",
                            file_pattern=pattern,
                            mapping=mapping,
                            transforms=transforms,
                            defaults=defaults,
                            is_active=True,
                            created_by=user_id,
                        )
                        db.add(new_map)
                        db.commit()
                        db.refresh(new_map)
                        auto_mapping_id = str(new_map.id)
        except Exception:
            db.rollback()

    batch = ImportBatch(
        id=_uuid4(),
        tenant_id=tenant_id,
        source_type=(dto.source_type or "products"),
        origin="excel",
        file_key=dto.file_key,
        parser_id=dto.parser_id,
        original_filename=original_filename,
        mapping_id=(
            UUID(dto.mapping_id or auto_mapping_id) if (dto.mapping_id or auto_mapping_id) else None
        ),
        created_by=str(user_id or "system"),
        status="PENDING",
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)

    # Try to auto-detect parser/doc_type from the uploaded file
    try:
        file_path = _file_path_from_key(batch.file_key)
        parser_id, detected_doc_type = select_parser_for_file(
            file_path,
            hinted_doc_type=dto.source_type,
            original_filename=original_filename,
        )
        batch.parser_id = parser_id
        batch.source_type = detected_doc_type or batch.source_type
        db.add(batch)
        db.commit()
        db.refresh(batch)
    except Exception:
        db.rollback()

    return batch


@router.get("/batches", response_model=list[BatchOut])
def list_batches_private(
    request: Request,
    db: Session = Depends(get_db),
    status: str | None = None,
):
    """Lista batches del tenant autenticado (equivalente al public pero con auth/tenant_id del token)."""
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    q = db.query(ImportBatch).filter(ImportBatch.tenant_id == tenant_id)
    if status:
        q = q.filter(ImportBatch.status == status)
    return q.order_by(ImportBatch.created_at.desc()).all()


@router.post("/batches/{batch_id}/start-excel-import")
def start_excel_import(batch_id: UUID, request: Request, db: Session = Depends(get_db)):
    """Encola importación genérica de archivo usando dispatcher de parsers."""
    from app.models.core.modelsimport import ImportBatch
    from app.modules.imports.application.celery_app import celery_app
    from app.modules.imports.application.job_runner import RUNNER_MODE
    from app.modules.imports.application.tasks.task_import_excel import import_products_excel
    from app.modules.imports.application.tasks.task_import_file import (
        import_file as import_file_task,
    )

    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    batch = (
        db.query(ImportBatch)
        .filter(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id)
        .with_for_update()
        .first()
    )
    if not batch or not batch.file_key:
        raise HTTPException(status_code=404, detail="batch_not_found_or_no_file")
    if str(getattr(batch, "status", "")) == "PARSING":
        return {
            "task_id": None,
            "status": "already_enqueued",
            "parser_id": getattr(batch, "parser_id", None),
            "doc_type": getattr(batch, "source_type", None),
        }

    # Ensure mapping is auto-picked on reprocess if missing.
    if not getattr(batch, "mapping_id", None):
        try:
            original_filename = getattr(
                batch, "original_filename", None
            ) or _load_original_filename_from_file_key(batch.file_key)
            headers_for_match = _extract_headers_from_file_key(batch.file_key)
            best_mapping, _, _ = _resolve_best_column_mapping(
                db,
                tenant_id=tenant_id,
                original_filename=original_filename,
                headers=headers_for_match,
            )
            if best_mapping:
                batch.mapping_id = best_mapping.id
                db.add(batch)
                db.commit()
        except Exception:
            db.rollback()

    file_path = _file_path_from_key(batch.file_key)
    ext = Path(file_path).suffix.lower()
    selected_parser = (getattr(batch, "parser_id", None) or "").strip() or None
    parser_user_confirmed = bool(getattr(batch, "confirmed_at", None)) or bool(
        getattr(batch, "user_override", False)
    )
    selected_ext = Path(getattr(batch, "original_filename", None) or file_path).suffix.lower()
    if (
        parser_user_confirmed
        and selected_parser
        and is_parser_compatible_with_extension(selected_parser, selected_ext)
    ):
        parser_id = selected_parser
        parser_info = parser_registry.get_parser(parser_id)
        detected_doc_type = (
            parser_info["doc_type"] if parser_info else (batch.source_type or "generic")
        )
    else:
        parser_id, detected_doc_type = select_parser_for_file(
            file_path,
            hinted_doc_type=batch.source_type,
            original_filename=getattr(batch, "original_filename", None),
        )
        if not is_parser_compatible_with_extension(parser_id, selected_ext):
            parser_id, detected_doc_type = select_fallback_parser_for_extension(
                selected_ext,
            )

    # Persistimos sugerencia de parser y doc_type detectado en batch si no está seteado
    try:
        batch.parser_id = parser_id
        if detected_doc_type:
            batch.source_type = detected_doc_type
        db.add(batch)
        db.commit()
    except Exception:
        db.rollback()

    args = {
        "tenant_id": str(tenant_id),
        "batch_id": str(batch.id),
        "file_key": batch.file_key,
        "parser_id": parser_id,
    }
    use_products_task = parser_id == "products_excel" and ext in (".xlsx", ".xls", ".xlsm", ".xlsb")
    # Reserve the batch before enqueuing to avoid duplicate starts for the same batch.
    try:
        if str(getattr(batch, "status", "")) != "PARSING":
            batch.status = "PARSING"
            db.add(batch)
            db.commit()
            db.refresh(batch)
    except Exception:
        db.rollback()
    if RUNNER_MODE != "celery":
        if use_products_task:
            result = import_products_excel.run(
                tenant_id=str(tenant_id),
                batch_id=str(batch.id),
                file_key=batch.file_key,
                source_type="products",
            )
        else:
            result = import_file_task.run(**args)
        return {
            "task_id": None,
            "status": "completed",
            "parser_id": parser_id,
            "doc_type": detected_doc_type,
            "result": result,
        }
    if use_products_task:
        result = celery_app.send_task(
            "imports.import_products_excel",
            kwargs={
                "tenant_id": str(tenant_id),
                "batch_id": str(batch.id),
                "file_key": batch.file_key,
                "source_type": "products",
            },
        )
    else:
        result = celery_app.send_task("imports.import_file", kwargs=args)
    return {
        "task_id": result.id,
        "status": "enqueued",
        "parser_id": parser_id,
        "doc_type": detected_doc_type,
    }


# ------------------------------
# Mapping suggestion (synonyms + optional LLM)
# ------------------------------


class SuggestIn(BaseModel):
    headers: list[str] | None = None


@router.post("/mappings/suggest")
async def suggest_mapping_endpoint(
    request: Request,
    body: SuggestIn | None = None,
    file: UploadFile | None = File(default=None),
):
    """Suggest a column mapping + transforms for a given Excel or headers list.

    - If `file` provided: parses headers using openpyxl (stream) and suggests mapping.
    - Else uses `body.headers`.
    """
    from io import BytesIO

    from app.modules.imports.application.transform_suggest import suggest_mapping
    from app.services.excel_analyzer import analyze_excel_stream

    headers: list[str] | None = None
    if file is not None:
        data = await file.read()
        analysis = analyze_excel_stream(BytesIO(data))
        headers = analysis.get("headers") or []
    elif body and body.headers:
        headers = list(body.headers)
    else:
        raise HTTPException(status_code=400, detail="file_or_headers_required")

    mapping, transforms, defaults, confidence = suggest_mapping(headers)

    # Optional: refine with Ollama if configured (best effort)
    try:
        import json

        import httpx  # type: ignore

        ollama_url = os.getenv("OLLAMA_BASE_URL") or os.getenv("OLLAMA_URL")
        if ollama_url and ollama_url.startswith("http"):
            prompt = (
                "Sugiere un mapeo JSON de columnas de Excel a campos del sistema"
                " (sku,name,price,stock,unit,category,image_url,packs,units_per_pack,pack_price)."
                f'\nCabeceras: {headers}\nPropuesta actual: {{"mapping":{mapping},"transforms":{transforms},"defaults":{defaults}}}\n'
                "Devuelve solo JSON con keys mapping, transforms, defaults."
            )
            payload = {"model": os.getenv("OLLAMA_MODEL", "llama3.1:8b"), "prompt": prompt}
            try:
                r = httpx.post(f"{ollama_url}/api/generate", json=payload, timeout=2.0)
                if r.status_code == 200:
                    txt = r.json().get("response", "")
                    # crude JSON extract
                    import re

                    m = re.search(r"\{[\s\S]*\}$", txt)
                    if m:
                        obj = json.loads(m.group(0))
                        mapping = obj.get("mapping", mapping)
                        transforms = obj.get("transforms", transforms)
                        defaults = obj.get("defaults", defaults)
            except Exception:
                pass
    except Exception:
        pass

    return {
        "headers": headers,
        "mapping": mapping,
        "transforms": transforms,
        "defaults": defaults,
        "confidence": confidence,
    }


@router.post("/batches/{batch_id}/cancel", response_model=OkResponse)
def cancel_batch_import(batch_id: UUID, request: Request):
    """Best-effort cancel of a running Celery task for this batch.

    Busca en tareas activas/prefetch con kwargs.batch_id == batch_id y hace revoke.
    """
    from app.modules.imports.application.celery_app import celery_app

    claims = _get_claims(request)
    if not claims.get("tenant_id"):
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    try:
        insp = celery_app.control.inspect()
        candidates = []
        for attr in ("active", "reserved", "scheduled"):
            try:
                groups = getattr(insp, attr)() or {}
            except Exception:
                groups = {}
            for _, tasks in groups.items():
                for t in tasks or []:
                    try:
                        name = t.get("name") or t.get("type")
                        kw = t.get("kwargs") or {}
                        if isinstance(kw, str):
                            # when serialized as string
                            kw_str = kw
                        else:
                            kw_str = str(kw)
                        if name == "imports.import_products_excel" and str(batch_id) in kw_str:
                            candidates.append(t.get("id"))
                    except Exception:
                        continue
        for tid in candidates:
            try:
                celery_app.control.revoke(tid, terminate=True)
            except Exception:
                pass
        return OkResponse()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"cancel_failed: {e}") from e


# Primary English route
@router.post(
    "/process",
    status_code=status.HTTP_202_ACCEPTED,
    response_model=OCRJobEnqueuedResponse,
)
# DEPRECATED: legacy Spanish route - use /process instead
@router.post(
    "/procesar",
    status_code=status.HTTP_202_ACCEPTED,
    response_model=OCRJobEnqueuedResponse,
    deprecated=True,
)
# DEPRECATED: legacy Spanish route - use /process instead
@router.post(
    "/documento/procesar",
    status_code=status.HTTP_202_ACCEPTED,
    response_model=OCRJobEnqueuedResponse,
    deprecated=True,
)
async def process_document_api(
    request: Request,
    file: UploadFile = File(...),
):
    access_claims = getattr(request.state, "access_claims", None)
    tenant_id = access_claims.get("tenant_id") if access_claims else None
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Falta tenant_id en el token")

    contenido = await file.read()
    if not contenido:
        raise HTTPException(status_code=400, detail="Archivo vacio")

    max_mb = float(os.getenv("IMPORTS_MAX_UPLOAD_MB", "10"))
    max_bytes = int(max_mb * 1024 * 1024)
    if len(contenido) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"File too large: {len(contenido)} bytes > {max_bytes}",
        )

    # OCR endpoint: only PDFs and images. Excel/CSV must go through the import pipeline (batches/ingest).
    allowed = {
        "application/pdf",
        "image/jpeg",
        "image/jpg",
        "image/png",
        "image/webp",
        "image/bmp",
        "image/tiff",
    }
    content_type = file.content_type or "application/octet-stream"
    if content_type not in allowed:
        raise HTTPException(status_code=422, detail=f"Unsupported file type: {content_type}")

    filename = file.filename or "documento.pdf"
    # If you don't want to store any file bytes anywhere (DB/uploads), process inline and
    # persist only filename + extracted JSON into import_ocr_jobs.
    store_payload = os.getenv("IMPORTS_OCR_STORE_PAYLOAD", "1").lower() in ("1", "true", "yes")
    if not store_payload:
        from app.config.database import session_scope
        from app.db.rls import set_tenant_guc
        from app.models.core.modelsimport import ImportOCRJob
        from app.modules.imports import services

        def _serialize_docs(documentos: list[Any]) -> list[dict[str, Any]]:
            out: list[dict[str, Any]] = []
            for doc in documentos:
                if hasattr(doc, "model_dump"):
                    out.append(doc.model_dump())  # type: ignore[attr-defined]
                elif isinstance(doc, dict):
                    out.append(doc)
                else:
                    try:
                        out.append(dict(doc))
                    except Exception:
                        out.append({"valor": str(doc)})
            return out

        job_uuid = uuid4()
        documentos = services.procesar_documento(contenido, filename)
        result = {"archivo": filename, "documentos": _serialize_docs(documentos)}
        with session_scope() as db:
            try:
                set_tenant_guc(db, str(tenant_id), persist=False)
            except Exception:
                pass
            job = ImportOCRJob(
                id=job_uuid,
                tenant_id=tenant_id,
                filename=filename,
                content_type=content_type,
                payload=b"",  # keep schema, but store no file bytes
                status="done",
                result=result,
                error=None,
            )
            db.add(job)
            db.flush()

        return {"job_id": str(job_uuid), "status": "done"}

    job_id = enqueue_job(
        tenant_id=tenant_id,
        filename=filename,
        content_type=content_type,
        payload=contenido,
    )
    return {"job_id": str(job_id), "status": "pending"}


@router.post("/excel/parse")
async def parse_excel(
    request: Request,
    file: UploadFile = File(...),
):
    """Parsea un archivo Excel (xlsx/xls) y devuelve cabeceras y filas como strings.

    Nota: requiere tener instalado el motor de lectura (openpyxl para xlsx). Si
    falta, responde 501 con un mensaje claro.
    """
    access_claims = getattr(request.state, "access_claims", None)
    tenant_id = access_claims.get("tenant_id") if access_claims else None
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Falta tenant_id en el token")

    content_type = (file.content_type or "").lower()
    if not (
        content_type.endswith("spreadsheetml.sheet")
        or content_type.endswith("ms-excel")
        or file.filename.lower().endswith((".xlsx", ".xls"))
    ):
        raise HTTPException(status_code=415, detail=f"Unsupported file type: {file.content_type}")
    try:
        from io import BytesIO

        data = await file.read()
        if not data:
            raise HTTPException(status_code=400, detail="Archivo vac????o")
        is_legacy_xls = bool(file.filename and file.filename.lower().endswith(".xls"))

        # Prefer openpyxl for xlsx to honor header row detection and merged cells.
        try:
            import openpyxl  # type: ignore

            bio = BytesIO(data)
            wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
            ws = wb.active
            header_row = detect_header_row(ws)
            headers = extract_headers(ws, header_row)
            rows = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row_dict = {}
                has_value = False
                for col_idx, header in enumerate(headers):
                    val = ws.cell(row_idx, col_idx + 1).value
                    if val is None:
                        row_dict[header] = ""
                        continue
                    row_dict[header] = str(val)
                    if str(val).strip() != "":
                        has_value = True
                if has_value:
                    rows.append(row_dict)
            wb.close()
            return {"headers": headers, "rows": rows}
        except Exception:
            # Fallback to pandas for other formats or if openpyxl fails.
            import pandas as pd  # type: ignore

            def _detect_header_row(df_raw):
                keywords = [
                    "fecha",
                    "factura",
                    "cliente",
                    "producto",
                    "cantidad",
                    "precio",
                    "subtotal",
                    "iva",
                    "total",
                    "tipo",
                    "codigo",
                    "descripcion",
                ]
                max_scan = min(len(df_raw.index), 40)
                best_idx = 0
                best_score = -(10**9)
                for idx in range(max_scan):
                    values = [str(v).strip() for v in df_raw.iloc[idx].tolist()]
                    non_empty = [v for v in values if v and v.lower() != "nan"]
                    if len(non_empty) < 2:
                        continue
                    lowered = " ".join(v.lower() for v in non_empty)
                    kw_hits = sum(1 for kw in keywords if kw in lowered)
                    unnamed_penalty = sum(1 for v in non_empty if v.lower().startswith("unnamed"))
                    score = len(non_empty) + (kw_hits * 4) - (unnamed_penalty * 2)
                    if score > best_score:
                        best_score = score
                        best_idx = idx
                return best_idx

            def _build_headers(raw_headers):
                headers = []
                seen: dict[str, int] = {}
                for i, value in enumerate(raw_headers):
                    header = str(value).strip()
                    if (
                        not header
                        or header.lower() == "nan"
                        or header.lower().startswith("unnamed")
                    ):
                        header = f"col_{i + 1}"
                    if header in seen:
                        seen[header] += 1
                        header = f"{header}_{seen[header]}"
                    else:
                        seen[header] = 1
                    headers.append(header)
                return headers

            with BytesIO(data) as bio:
                if is_legacy_xls:
                    try:
                        df = pd.read_excel(bio, engine="xlrd", header=None)
                    except ImportError as e:
                        raise HTTPException(
                            status_code=422,
                            detail=(
                                "No se puede leer .xls antiguo: falta dependencia 'xlrd'. "
                                "Convierte el archivo a .xlsx/.csv o instala xlrd en backend."
                            ),
                        ) from e
                    except Exception as e:
                        raise HTTPException(
                            status_code=422,
                            detail=(
                                "No se pudo leer el archivo .xls. "
                                "Convierte a .xlsx/.csv e intenta nuevamente."
                            ),
                        ) from e
                else:
                    try:
                        df = pd.read_excel(bio, engine=None, header=None)
                    except Exception:
                        df = pd.read_excel(bio, engine="openpyxl", header=None)

            df = df.fillna("")
            if df.empty:
                return {"headers": [], "rows": []}

            header_row_idx = _detect_header_row(df)
            headers = _build_headers(df.iloc[header_row_idx].tolist())
            rows = []
            for row_idx in range(header_row_idx + 1, len(df.index)):
                vals = df.iloc[row_idx].tolist()
                if not any(str(v).strip() and str(v).lower() != "nan" for v in vals):
                    continue
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    val = vals[col_idx] if col_idx < len(vals) else ""
                    text = "" if str(val).lower() == "nan" else str(val)
                    row_dict[header] = text
                rows.append(row_dict)
            return {"headers": headers, "rows": rows}
    except ImportError as e:
        raise HTTPException(
            status_code=501,
            detail="Soporte XLSX no disponible en el servidor. Instala openpyxl para habilitarlo.",
        ) from e
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo procesar Excel: {e}") from e


@router.get("/jobs/{job_id}", response_model=OCRJobStatusResponse)
async def get_ocr_job(job_id: UUID, request: Request, db: Session = Depends(get_db)):
    access_claims = getattr(request.state, "access_claims", None)
    tenant_id = access_claims.get("tenant_id") if access_claims else None
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Falta tenant_id en el token")

    job = (
        db.query(ImportOCRJob)
        .filter(ImportOCRJob.id == job_id, ImportOCRJob.tenant_id == tenant_id)
        .first()
    )
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")

    return {
        "job_id": str(job.id),
        "status": job.status,
        "result": job.result,
        "error": job.error,
    }


@public_router.get("/health")
def imports_health(request: Request, db: Session = Depends(get_db)):
    """Quick schema presence check for imports module.

    Returns list of missing tables/columns useful to diagnose migration issues rapidly.
    """
    insp = sa_inspect(db.get_bind())
    required = {
        "import_batches": ["id", "tenant_id", "source_type", "origin", "status"],
        "import_items": ["id", "batch_id", "idx", "raw", "status", "idempotency_key"],
        "import_mappings": ["id", "tenant_id", "name", "source_type"],
        "import_item_corrections": ["id", "tenant_id", "item_id", "user_id", "field"],
        "import_lineage": ["id", "tenant_id", "item_id", "promoted_to"],
        "auditoria_importacion": [
            "id",
            "tenant_id",
            "documento_id",
            "fecha",
            "batch_id",
            "item_id",
        ],
    }
    missing = {"tables": [], "columns": []}
    existing_tables = set(insp.get_table_names())
    for tbl, cols in required.items():
        if tbl not in existing_tables:
            missing["tables"].append(tbl)
            continue
        cols_exist = {c["name"] for c in insp.get_columns(tbl)}
        for c in cols:
            if c not in cols_exist:
                missing["columns"].append(f"{tbl}.{c}")
    limits = {
        "max_items_per_batch": int(os.getenv("IMPORTS_MAX_ITEMS_PER_BATCH", "5000")),
        "max_upload_mb": float(os.getenv("IMPORTS_MAX_UPLOAD_MB", "10")),
        "max_ingests_per_min": int(os.getenv("IMPORTS_MAX_INGESTS_PER_MIN", "30")),
        "allowed_mimetypes": [
            "image/jpeg",
            "image/png",
            "image/heic",
            "application/pdf",
        ],
    }
    deps = {}
    try:
        import openpyxl  # noqa: F401

        deps["openpyxl"] = "ok"
    except Exception as e:  # pragma: no cover - entorno
        deps["openpyxl"] = f"missing: {e}"
    try:
        import pandas  # noqa: F401

        deps["pandas"] = "ok"
    except Exception as e:  # pragma: no cover
        deps["pandas"] = f"missing: {e}"

    return {
        "ok": not (missing["tables"] or missing["columns"]),
        "missing": missing,
        "limits": limits,
        "dependencies": deps,
    }


# ------------------------------
# Parser Registry endpoint (Sprint 2)
# ------------------------------


@router.get("/parsers/registry", tags=["Imports"])
def get_parser_registry(request: Request):
    """
    Retorna el registry de parsers disponibles.
    Usado por el frontend para mostrar opciones de override en el Wizard.

    Endpoint: GET /api/v1/imports/parsers/registry

    Respuesta:
    ```json
    {
        "parsers": {
            "csv_products": {
                "id": "csv_products",
                "doc_type": "products",
                "description": "CSV parser for product data"
            },
            ...
        }
    }
    ```
    """
    from app.modules.imports.parsers import registry

    # Construir respuesta sin los handlers (no son serializables)
    parsers_info = {}
    for parser_id, parser_data in registry.list_parsers().items():
        parsers_info[parser_id] = {
            "id": parser_data["id"],
            "doc_type": parser_data["doc_type"],
            "description": parser_data.get("description", ""),
        }

    return {
        "parsers": parsers_info,
        "count": len(parsers_info),
    }


# ------------------------------
# New batch endpoints (EPIC flow)
# ------------------------------


@router.post("/batches", response_model=BatchOut)
def create_batch_endpoint(dto: BatchCreate, request: Request, db: Session = Depends(get_db)):
    """Skeleton endpoint for batch creation. Returns placeholder until models are wired."""
    from app.modules.imports.application import use_cases

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    user_id = claims.get("user_id")
    batch = use_cases.create_batch(db, tenant_id, user_id, dto.model_dump())
    return batch


@router.patch("/batches/{batch_id}/classification", response_model=BatchOut)
def update_classification(
    batch_id: UUID,
    req: UpdateClassificationRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Actualizar campos de clasificación en un batch existente.
    Permite override manual del usuario sobre la clasificación automática.

    Endpoint: PATCH /api/v1/imports/batches/{batch_id}/classification

    Ejemplo:
    ```json
    {
        "suggested_parser": "excel_invoices",
        "classification_confidence": 0.95,
        "ai_enhanced": true,
        "ai_provider": "openai"
    }
    ```
    """
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    # Buscar batch con RLS (row-level security)
    batch = (
        db.query(ImportBatch)
        .filter(
            ImportBatch.id == batch_id,
            ImportBatch.tenant_id == tenant_id,
        )
        .first()
    )

    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    # Actualizar solo los campos que se proporcionen (no sobrescribir con None)
    if req.suggested_parser is not None:
        batch.suggested_parser = req.suggested_parser
    if req.classification_confidence is not None:
        batch.classification_confidence = req.classification_confidence
    if req.ai_enhanced is not None:
        batch.ai_enhanced = req.ai_enhanced
    if req.ai_provider is not None:
        batch.ai_provider = req.ai_provider

    db.commit()
    db.refresh(batch)
    return batch


@router.post("/batches/{batch_id}/classify-and-persist", response_model=BatchOut)
async def classify_and_persist_to_batch(
    batch_id: UUID,
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """
    Clasificar archivo Y persistir resultado en el batch.
    Integra clasificación (heurística + IA) con persistencia automática.

    Endpoint: POST /api/v1/imports/batches/{batch_id}/classify-and-persist

    Pasos:
    1. Recibe archivo para clasificar
    2. Ejecuta clasificación heurística + IA
    3. Persiste resultado en campos: suggested_parser, classification_confidence, ai_enhanced, ai_provider
    4. Retorna batch actualizado

    Retorna ClassifyResponse extendido con batch info.
    """
    import tempfile

    from app.modules.imports.services.classifier import classifier

    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    # Validar archivo
    if not file.filename:
        raise HTTPException(status_code=422, detail="Nombre de archivo requerido")

    ext = file.filename.lower().split(".")[-1]
    if ext not in ["xlsx", "xls", "csv", "xml"]:
        raise HTTPException(status_code=422, detail="Solo se aceptan archivos Excel, CSV o XML")

    # Buscar batch
    batch = (
        db.query(ImportBatch)
        .filter(
            ImportBatch.id == batch_id,
            ImportBatch.tenant_id == tenant_id,
        )
        .first()
    )

    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    tmp_path = None
    try:
        # Guardar temporalmente
        contents = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}") as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        # Clasificar con IA
        result = await classifier.classify_file_with_ai(tmp_path, file.filename)

        # Persistir en batch
        batch.suggested_parser = result.get("suggested_parser")
        batch.classification_confidence = result.get("confidence")
        batch.ai_enhanced = result.get("enhanced_by_ai", False)
        batch.ai_provider = result.get("ai_provider")

        db.commit()
        db.refresh(batch)

        return batch

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error al clasificar y persistir: {str(e)}"
        ) from e
    finally:
        # Limpiar
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/batches/{batch_id}/ingest", response_model=list)
def ingest_rows_endpoint(
    batch_id: UUID,  # <-- UUID correcto para el id del batch
    payload: IngestRows,
    request: Request,
    db: Session = Depends(get_db),
    column_mapping_id: str = None,  # Nuevo: ID del mapeo de columnas
):
    from uuid import UUID as PYUUID

    from app.models.imports import ImportColumnMapping
    from app.modules.imports.application import use_cases

    claims = _get_claims(request)
    tenant_raw = claims.get("tenant_id")
    # tenant_id is now UUID string, keep as-is
    tenant_id = tenant_raw

    # throttle per tenant
    _throttle_ingest(tenant_raw)

    # busca el batch (id UUID + tenant_id del tipo que corresponda en tu modelo)
    batch = (
        db.query(ImportBatch)
        .filter(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Batch no encontrado")

    # ðŸ†• SMART IMPORT: Aplicar mapeo de columnas si se proporciona
    if column_mapping_id:
        try:
            col_mapping = (
                db.query(ImportColumnMapping)
                .filter(
                    ImportColumnMapping.id == PYUUID(column_mapping_id),
                    ImportColumnMapping.tenant_id == tenant_id,
                )
                .first()
            )

            if col_mapping:
                # Aplicar transformaciÃ³n de columnas a cada fila
                mapped_rows = []
                for row in payload.rows:
                    mapped_row = {}
                    for excel_col, target_field in col_mapping.mapping.items():
                        if target_field != "ignore" and excel_col in row:
                            mapped_row[target_field] = row[excel_col]
                    mapped_rows.append(mapped_row)

                payload.rows = mapped_rows

                # Actualizar estadÃ­sticas de uso
                col_mapping.last_used_at = datetime.utcnow()
                col_mapping.use_count += 1
                db.add(col_mapping)
                db.commit()
        except Exception as e:
            import logging

            logging.warning(f"Error aplicando column_mapping: {e}")

    # mappings / transforms / defaults (mapping_id puede ser None; si viene, intenta parsear a UUID solo si tiene pinta de serlo)
    mappings = None
    transforms = payload.transforms
    defaults = payload.defaults

    mapping_id_val = payload.mapping_id or batch.mapping_id
    if mapping_id_val:
        # intenta convertir a UUID solo si parece un UUID; si falla, dÃƒÆ’Ã‚Â©jalo tal cual
        try:
            mapping_id_val = UUID(str(mapping_id_val))
        except ValueError:
            pass

        cm = (
            db.query(ImportColumnMapping)
            .filter(
                ImportColumnMapping.id == mapping_id_val,
                ImportColumnMapping.tenant_id == tenant_id,
                ImportColumnMapping.is_active,
            )
            .first()
        )
        if cm:
            mappings = (cm.mapping or {}) or mappings
            transforms = (cm.transforms or {}) or transforms
            defaults = (cm.defaults or {}) or defaults
        else:
            # Backward compatibility with old import_mappings table.
            mp = (
                db.query(ImportMapping)
                .filter(
                    ImportMapping.id == mapping_id_val,
                    ImportMapping.tenant_id == tenant_id,
                    ImportMapping.source_type == batch.source_type,
                )
                .first()
            )
            if mp:
                mappings = mp.mappings or mappings
                transforms = mp.transforms or transforms
                defaults = mp.defaults or defaults

    # Limit number of rows per ingest (413 on overflow)
    import logging

    logger = logging.getLogger("imports")
    rows_count = len(payload.rows) if payload.rows else 0
    print(f"DEBUG ingest_endpoint: batch_id={batch_id}, payload.rows count={rows_count}")
    logger.info(f"ingest_endpoint: batch_id={batch_id}, payload.rows count={rows_count}")

    max_items = int(os.getenv("IMPORTS_MAX_ITEMS_PER_BATCH", "5000"))
    if payload.rows and len(payload.rows) > max_items:
        raise HTTPException(
            status_code=413, detail=f"Too many rows: {len(payload.rows)} > {max_items}"
        )

    # Recipe fast-path: parse server-side file and persist recipes/ingredients
    if batch.source_type == "recipes" or batch.parser_id == "xlsx_recipes":
        from app.modules.imports.application.tasks.task_import_file import (
            _file_path_from_key,
            _persist_recipes,
        )
        from app.modules.imports.parsers.xlsx_recipes import parse_xlsx_recipes

        file_path = _file_path_from_key(batch.file_key) if batch.file_key else None
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=400, detail="Recipe file not found for batch")

        parsed = parse_xlsx_recipes(file_path)
        result = _persist_recipes(db, tenant_id, parsed)
        batch.status = "READY" if result.get("errors", 0) == 0 else "PARTIAL"
        db.add(batch)
        db.commit()
        return [
            {
                "batch_id": str(batch.id),
                "doc_type": "recipes",
                "summary": result,
            }
        ]

    items = use_cases.ingest_rows(
        db,
        tenant_id,  # <-- pasa el tenant_id tal cual (int o string, segÃƒÆ’Ã‚Âºn tu modelo)
        batch,
        payload.rows,
        mappings=mappings,
        transforms=transforms,
        defaults=defaults,
        dedupe_keys=(mp.dedupe_keys if "mp" in locals() and mp else None),
    )

    # Para el flujo MVP: dejamos el batch en estado READY tras ingesta
    # y NO encolamos el pipeline automÃ¡tico (evita publicar a productos sin revisiÃ³n).
    # El estado READY lo establece use_cases.ingest_rows; aquÃ­ no lo sobreescribimos.
    # For API ergonomics, return the created/updated items list (tests expect this)
    out = [
        {
            "id": str(getattr(it, "id", "")),
            "idx": getattr(it, "idx", None),
            "status": getattr(it, "status", None),
            "errors": getattr(it, "errors", None),
        }
        for it in (items or [])
    ]
    return out


@router.get("/batches/{batch_id}", response_model=BatchOut)
def get_batch_endpoint(batch_id: UUID, request: Request, db: Session = Depends(get_db)):
    claims = _get_claims(request)
    batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch no encontrado")
    if str(batch.tenant_id) != str(claims.get("tenant_id")):
        raise HTTPException(status_code=403, detail="No autorizado")
    return batch


@router.get("/batches/{batch_id}/status")
def get_batch_status_endpoint(batch_id: UUID, request: Request, db: Session = Depends(get_db)):
    from app.models.core.modelsimport import ImportItem
    from app.modules.imports.application.status import ImportBatchStatus, ImportItemStatus

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    batch = (
        db.query(ImportBatch)
        .filter(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Batch no encontrado")

    from sqlalchemy import func

    status_counts = dict(
        db.query(ImportItem.status, func.count(ImportItem.id))
        .filter(ImportItem.batch_id == batch_id)
        .group_by(ImportItem.status)
        .all()
    )

    total = sum(status_counts.values())
    completed = status_counts.get(ImportItemStatus.PROMOTED, 0)
    failed = status_counts.get(ImportItemStatus.ERROR_VALIDATION, 0) + status_counts.get(
        ImportItemStatus.ERROR_PROMOTION, 0
    )
    pending = status_counts.get(ImportItemStatus.PENDING, 0) + status_counts.get(
        ImportItemStatus.OK, 0
    )
    processing = max(0, total - (completed + failed + pending))
    progress_pct = round((completed / total) * 100, 1) if total > 0 else 0.0

    from datetime import datetime as _dt

    # Si no quedan pendientes/en proceso, actualizar estado del batch según resultado
    if (total == 0) or (pending == 0 and processing == 0):
        if completed == total and total > 0:
            target_status = ImportBatchStatus.PROMOTED
        elif total == 0:
            target_status = ImportBatchStatus.EMPTY
        elif failed > 0 and completed == 0:
            target_status = ImportBatchStatus.ERROR
        elif failed > 0 and completed > 0:
            target_status = ImportBatchStatus.PARTIAL
        else:
            target_status = ImportBatchStatus.READY
        if batch.status != target_status:
            batch.status = target_status
            try:
                db.add(batch)
                db.commit()
            except Exception:
                db.rollback()

    return {
        "batch_id": str(batch_id),
        "status": batch.status,
        "total_items": total,
        "completed": completed,
        "processing": processing,
        "failed": failed,
        "pending": pending,
        "progress": progress_pct,
        "status_breakdown": status_counts,
        "server_time": _dt.utcnow().isoformat() + "Z",
    }


@router.post("/batches/{batch_id}/retry")
def retry_batch_endpoint(batch_id: UUID, request: Request, db: Session = Depends(get_db)):
    from app.modules.imports.domain.pipeline import retry_failed_items

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    batch = (
        db.query(ImportBatch)
        .filter(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Batch no encontrado")

    tenant_id = getattr(batch, "tenant_id", None) or batch_id
    result = retry_failed_items(batch_id, UUID(str(tenant_id)))

    return result


@router.post("/batches/{batch_id}/reset", response_model=OkResponse)
def reset_batch_endpoint(
    batch_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    clear_items: bool = Query(default=True, description="Eliminar items del lote"),
    new_status: str = Query(default="PENDING", description="Nuevo estado del batch"),
    mapping_id: str | None = Query(default=None, description="Nuevo mapping_id (opcional)"),
):
    """Resetear un lote atascado: opcionalmente borra sus items y deja el lote en PENDING o ERROR.

    No revoca tareas Celery existentes (no guardamos task_id aÃºn), pero permite relanzar o eliminar.
    """
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    batch = (
        db.query(ImportBatch)
        .filter(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="batch_not_found")

    if clear_items:
        db.query(ImportItem).filter(ImportItem.batch_id == batch_id).delete(
            synchronize_session=False
        )

    # Sanitize status
    allowed = {"PENDING", "ERROR", "READY", "PARSING"}
    batch.status = new_status if new_status in allowed else "PENDING"
    if mapping_id:
        try:
            batch.mapping_id = UUID(str(mapping_id))
        except Exception:
            pass
    db.add(batch)
    db.commit()
    return OkResponse()


class SetMappingDTO(BaseModel):
    mapping_id: str


@router.post("/batches/{batch_id}/set-mapping", response_model=OkResponse)
def set_batch_mapping_endpoint(
    batch_id: UUID, dto: SetMappingDTO, request: Request, db: Session = Depends(get_db)
):
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    batch = (
        db.query(ImportBatch)
        .filter(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="batch_not_found")
    try:
        mapping_uuid = UUID(str(dto.mapping_id))
        batch.mapping_id = mapping_uuid
    except Exception as e:
        raise HTTPException(status_code=400, detail="invalid_mapping_id") from e
    db.add(batch)
    db.commit()

    # Learn from explicit user-approved mapping reassignment.
    try:
        mapping_obj = (
            db.query(ImportColumnMapping)
            .filter(
                ImportColumnMapping.id == mapping_uuid,
                ImportColumnMapping.tenant_id == tenant_id,
            )
            .first()
        )
        if mapping_obj and isinstance(mapping_obj.mapping, dict):
            from app.modules.imports.domain.mapping_feedback import MappingFeedback, mapping_learner

            feedback = MappingFeedback(
                tenant_id=tenant_id,
                batch_id=batch_id,
                doc_type=str(batch.source_type or "products"),
                headers=list(mapping_obj.mapping.keys()),
            )
            for source_field, canonical in mapping_obj.mapping.items():
                if canonical and canonical != "ignore":
                    feedback.mark_field_correct(
                        source_field=str(source_field),
                        canonical=str(canonical),
                        confidence=0.9,
                    )
            if feedback.field_feedbacks:
                mapping_learner.record_feedback(feedback)
    except Exception as e:
        logger.warning(f"mapping learning skipped on set-mapping: {e}")
    return OkResponse()


@public_router.get("/batches", response_model=list[BatchOut])
def list_batches_endpoint(
    db: Session = Depends(get_db),
    status: str | None = None,
    tenant_id: str = Query(..., description="Tenant ID (UUID)"),
):
    if not tenant_id:
        raise HTTPException(status_code=400, detail="tenant_id required")

    q = db.query(ImportBatch).filter(ImportBatch.tenant_id == tenant_id)
    if status:
        q = q.filter(ImportBatch.status == status)
    return q.order_by(ImportBatch.created_at.desc()).all()


@router.get("/batches/{batch_id}/items", response_model=list[ItemOut])
def list_batch_items_endpoint(
    batch_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    status: str | None = None,
    q: str | None = None,
    with_: str | None = Query(default=None, alias="with"),
):
    from app.modules.imports.infrastructure.repositories import ImportsRepository

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    repo = ImportsRepository()
    items = repo.list_items(db, tenant_id, batch_id, status=status, q=q)  # batch_id es UUID
    if with_ != "lineage":
        return items
    item_ids = [it.id for it in items]
    if not item_ids:
        return items
    lineages = (
        db.query(ImportLineage)
        .filter(ImportLineage.item_id.in_(item_ids), ImportLineage.tenant_id == tenant_id)
        .all()
    )
    corrections = (
        db.query(ImportItemCorrection)
        .filter(
            ImportItemCorrection.item_id.in_(item_ids),
            ImportItemCorrection.tenant_id == tenant_id,
        )
        .order_by(ImportItemCorrection.created_at.desc())
        .all()
    )
    # group
    from collections import defaultdict

    lineage_map = defaultdict(list)
    for lg in lineages:
        lineage_map[lg.item_id].append(
            {
                "promoted_to": lg.promoted_to,
                "promoted_ref": lg.promoted_ref,
                "created_at": lg.created_at,
            }
        )
    corr_map = {}
    for corr in corrections:
        corr_map.setdefault(corr.item_id, corr)  # keep latest

    out = []
    for it in items:
        obj = {
            "id": it.id,
            "idx": it.idx,
            "status": it.status,
            "raw": it.raw,
            "normalized": it.normalized,
            "errors": it.errors,
            "promoted_to": it.promoted_to,
            "promoted_id": it.promoted_id,
            "promoted_at": it.promoted_at,
            "lineage": lineage_map.get(it.id, []),
        }
        last = corr_map.get(it.id)
        if last is not None:
            obj["last_correction"] = {
                "field": last.field,
                "old_value": last.old_value,
                "new_value": last.new_value,
                "created_at": last.created_at,
                "user_id": str(last.user_id),
            }
        out.append(obj)
    return out


@public_router.get("/items/products")
def list_all_products_endpoint(
    request: Request,
    db: Session = Depends(get_db),
    status: str | None = None,
    q: str | None = None,
    limit: int = Query(default=5000, le=10000),
    offset: int = Query(default=0, ge=0),
    tenant_id: str | None = Query(None, description="Tenant ID (UUID) - opcional si usa RLS"),
    debug: int | None = Query(default=None),
):
    """List ALL import items formatted as products across all batches."""
    from app.modules.imports.infrastructure.repositories import ImportsRepository

    # Si no viene tenant_id, intentar obtenerlo del RLS context
    if not tenant_id:
        try:
            tenant_id = db.scalar(select(tenant_id_sql_expr()))
        except Exception:
            pass

    if not tenant_id:
        raise HTTPException(status_code=400, detail="tenant_id required (via query param or RLS)")

    repo = ImportsRepository()
    # List items from ALL batches of type 'productos'
    all_items = []
    # Aceptar tanto 'products' (nuevo) como 'productos' (legacy) para no dejar fuera lotes recientes
    batches = (
        db.query(ImportBatch)
        .filter(
            ImportBatch.tenant_id == tenant_id,
            ImportBatch.source_type.in_(("products", "productos", "product")),
        )
        .order_by(ImportBatch.created_at.desc())
        .all()
    )
    for batch in batches:
        batch_items = repo.list_items(db, tenant_id, batch.id, status=status, q=q)
        if not status:
            batch_items = [it for it in batch_items if it.status != "PROMOTED"]
        all_items.extend(batch_items)

    # Format items as products
    products = []
    for item in all_items[offset : offset + limit]:
        data = {}
        raw = item.raw or {}
        if isinstance(raw, dict) and isinstance(raw.get("datos"), dict):
            data.update(raw.get("datos") or {})
        if isinstance(raw, dict):
            data.update(raw)
        if isinstance(item.normalized, dict):
            data.update(item.normalized)
        nmap = _build_norm_map(data)

        codigo = _first(nmap, _ALIASES.get("codigo", []))
        nombre = _first(nmap, _ALIASES.get("nombre", []))
        precio = _first(nmap, _ALIASES.get("precio", []))
        precio_bulto = _first(nmap, _ALIASES.get("precio_por_bulto", []))
        costo = _first(nmap, _ALIASES.get("costo", []))
        categoria = _first(nmap, _ALIASES.get("categoria", []))
        stock_v = _first(nmap, _ALIASES.get("stock", []))
        bultos_v = _first(nmap, _ALIASES.get("bultos", []))
        unidades_por_bulto_v = _first(nmap, _ALIASES.get("cantidad_por_bulto", []))
        unidad = _first(nmap, _ALIASES.get("unidad", [])) or "unit"
        iva_v = _first(nmap, _ALIASES.get("iva", []))

        upb = _parse_num(unidades_por_bulto_v) or 0
        p_bulto = _parse_num(precio_bulto)
        p_unit = _parse_num(precio)
        if (p_unit is None or p_unit == 0) and p_bulto and upb and upb > 0:
            p_unit = round(p_bulto / upb, 6)

        stock_final = _parse_num(stock_v)
        if (stock_final is None or stock_final == 0) and upb and (_parse_num(bultos_v) or 0) > 0:
            stock_final = int((_parse_num(bultos_v) or 0) * upb)

        products.append(
            {
                "id": str(item.id),
                "idx": item.idx,
                "status": item.status,
                "errors": item.errors or [],
                "batch_id": str(item.batch_id),
                "codigo": codigo,
                "nombre": nombre,
                "precio": p_unit,
                "costo": _parse_num(costo),
                "categoria": categoria,
                "stock": int(stock_final or 0),
                "unidad": unidad,
                "iva": _parse_num(iva_v) or 0,
                "raw": item.raw,
                "normalized": item.normalized,
            }
        )

    if debug:
        try:
            sample = products[:3]
            batch_ids = {p.get("batch_id") for p in sample if p.get("batch_id")}
            batch_info = {}
            if batch_ids:
                rows = db.query(ImportBatch).filter(ImportBatch.id.in_(list(batch_ids))).all()
                for b in rows:
                    batch_info[str(b.id)] = {
                        "origin": getattr(b, "origin", None),
                        "file_key": getattr(b, "file_key", None),
                        "parser_id": getattr(b, "parser_id", None),
                        "source_type": getattr(b, "source_type", None),
                    }
            print(
                "DEBUG imports.items.products sample:",
                {"items": sample, "batches": batch_info},
            )
        except Exception:
            pass

    return {
        "items": products,
        "total": len(all_items),
        "limit": limit,
        "offset": offset,
    }


# Variante autenticada bajo /tenant (usa tenant_id del token)
@router.get("/items/products")
def list_all_products_private_endpoint(
    request: Request,
    db: Session = Depends(get_db),
    status: str | None = None,
    q: str | None = None,
    limit: int = Query(default=5000, le=10000),
    offset: int = Query(default=0, ge=0),
):
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    return list_all_products_endpoint(
        request=request,
        db=db,
        status=status,
        q=q,
        limit=limit,
        offset=offset,
        tenant_id=str(tenant_id),
    )


@router.get("/batches/{batch_id}/items/products")
def list_batch_items_products_endpoint(
    batch_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    status: str | None = None,
    # Alinear con frontend (PreviewPage usa limit=5000)
    limit: int = Query(default=5000, le=10000),
    offset: int = Query(default=0, ge=0),
    debug: int | None = Query(default=None),
):
    """List import items formatted as products with relevant fields for a specific batch."""
    from app.modules.imports.infrastructure.repositories import ImportsRepository

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    # Establecer GUC RLS usando request + db (firma correcta)
    ensure_rls(request, db)

    repo = ImportsRepository()
    items = repo.list_items(db, tenant_id, batch_id, status=status, q=None)
    if not status:
        items = [it for it in items if it.status != "PROMOTED"]

    # Format items as products
    products = []
    for item in items[offset : offset + limit]:
        # Merge raw and normalized data (normalized overrides raw) and flatten nested "datos"
        data = {}
        raw = item.raw or {}
        if isinstance(raw, dict) and isinstance(raw.get("datos"), dict):
            data.update(raw.get("datos") or {})
        data.update(raw)
        if isinstance(item.normalized, dict):
            data.update(item.normalized)
        nmap = _build_norm_map(data)

        codigo = _first(nmap, _ALIASES.get("codigo", []))
        nombre = _first(nmap, _ALIASES.get("nombre", []))
        precio = _first(nmap, _ALIASES.get("precio", []))
        precio_bulto = _first(nmap, _ALIASES.get("precio_por_bulto", []))
        costo = _first(nmap, _ALIASES.get("costo", []))
        categoria = _first(nmap, _ALIASES.get("categoria", []))
        stock_v = _first(nmap, _ALIASES.get("stock", []))
        bultos_v = _first(nmap, _ALIASES.get("bultos", []))
        unidades_por_bulto_v = _first(nmap, _ALIASES.get("cantidad_por_bulto", []))
        unidad = _first(nmap, _ALIASES.get("unidad", [])) or "unit"
        iva_v = _first(nmap, _ALIASES.get("iva", []))

        # HeurÃ­sticas: derivar precio unitario y stock si faltan
        upb = _parse_num(unidades_por_bulto_v) or 0
        p_bulto = _parse_num(precio_bulto)
        p_unit = _parse_num(precio)
        if (p_unit is None or p_unit == 0) and p_bulto and upb and upb > 0:
            p_unit = round(p_bulto / upb, 6)
        # Stock
        stock_final = _parse_num(stock_v)
        if (stock_final is None or stock_final == 0) and upb and (_parse_num(bultos_v) or 0) > 0:
            stock_final = int((_parse_num(bultos_v) or 0) * upb)

        products.append(
            {
                "id": str(item.id),
                "idx": item.idx,
                "status": item.status,
                "errors": item.errors or [],
                # Product fields (coerce numerics where reasonable)
                "codigo": codigo,
                "nombre": nombre,
                "precio": p_unit,
                "costo": _parse_num(costo),
                "categoria": categoria,
                "stock": int(stock_final or 0),
                "unidad": unidad,
                "iva": _parse_num(iva_v) or 0,
                # Metadata
                "raw": item.raw,
                "normalized": item.normalized,
            }
        )

    return {
        "items": products,
        "total": len(items),
        "limit": limit,
        "offset": offset,
    }


# ------------------------------
# Smart Import: Column analysis + saved mappings
# Endpoints expected by tenant frontend (columnMappingApi.ts)
# ------------------------------


@router.post("/analyze-file")
def analyze_file_endpoint(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Analyze uploaded Excel and return headers, sample rows and suggested mapping.

    Also returns saved column mappings for the tenant so the UI can offer reuse.
    """
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="Solo Excel (.xlsx, .xls)")

    try:
        contents = file.file.read()
    finally:
        try:
            file.file.close()
        except Exception:
            pass

    # Analyze via in-memory stream
    from io import BytesIO

    analysis = analyze_excel_stream(BytesIO(contents))

    # Load saved mappings for this tenant
    mappings = (
        db.query(ImportColumnMapping)
        .filter(ImportColumnMapping.tenant_id == tenant_id, ImportColumnMapping.is_active)  # noqa: E712
        .order_by(
            ImportColumnMapping.last_used_at.desc().nullslast(),
            ImportColumnMapping.created_at.desc(),
        )
        .all()
    )

    detected_headers = [str(h) for h in (analysis.get("headers") or []) if str(h).strip()]
    saved = []
    for m in mappings:
        similarity, overlap = _header_similarity_score(detected_headers, m.mapping or {})
        saved.append(
            {
                "id": str(m.id),
                "name": m.name,
                "description": m.description,
                "mapping": m.mapping or {},
                "transforms": m.transforms or {},
                "defaults": m.defaults or {},
                "file_pattern": m.file_pattern,
                "use_count": int(m.use_count or 0),
                "last_used_at": m.last_used_at.isoformat() if m.last_used_at else None,
                "created_at": m.created_at.isoformat() if m.created_at else None,
                "header_similarity": round(float(similarity), 4),
                "header_overlap": int(overlap),
            }
        )

    best_mapping, matched_by, best_score = _resolve_best_column_mapping(
        db,
        tenant_id=tenant_id,
        original_filename=file.filename,
        headers=detected_headers,
    )

    analysis["saved_mappings"] = saved
    analysis["best_mapping_id"] = str(best_mapping.id) if best_mapping else None
    analysis["best_mapping_name"] = best_mapping.name if best_mapping else None
    analysis["best_mapping_match"] = matched_by
    analysis["best_mapping_score"] = round(float(best_score), 4) if best_mapping else 0.0
    return analysis


@router.get("/column-mappings")
def list_column_mappings_endpoint(request: Request, db: Session = Depends(get_db)):
    """List saved column mappings for current tenant."""
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    mappings = (
        db.query(ImportColumnMapping)
        .filter(ImportColumnMapping.tenant_id == tenant_id, ImportColumnMapping.is_active)  # noqa: E712
        .order_by(ImportColumnMapping.name.asc())
        .all()
    )
    return [
        {
            "id": str(m.id),
            "name": m.name,
            "description": m.description,
            "mapping": m.mapping or {},
            "transforms": m.transforms or {},
            "defaults": m.defaults or {},
            "file_pattern": m.file_pattern,
            "use_count": int(m.use_count or 0),
            "last_used_at": m.last_used_at.isoformat() if m.last_used_at else None,
            "created_at": m.created_at.isoformat() if m.created_at else None,
        }
        for m in mappings
    ]


class CreateColumnMappingDTO(BaseModel):
    name: str
    mapping: dict
    description: str | None = None
    file_pattern: str | None = None
    transforms: dict | None = None
    defaults: dict | None = None


_IMPORT_ALIAS_DOC_TYPES = {
    "invoices",
    "bank_transactions",
    "products",
    "generic",
    "expenses",
}


def _module_for_import_aliases(doc_type: str) -> str:
    dt = str(doc_type or "").strip().lower()
    if dt not in _IMPORT_ALIAS_DOC_TYPES:
        raise HTTPException(status_code=422, detail="unsupported_doc_type")
    return f"imports_{dt}"


class ImportAliasFieldIn(BaseModel):
    field: str
    aliases: list[str] = []
    field_type: str | None = None
    required: bool = False


class ImportAliasConfigIn(BaseModel):
    doc_type: str
    fields: list[ImportAliasFieldIn]


@router.get("/field-aliases")
def get_import_field_aliases(
    request: Request,
    doc_type: str = Query(..., description="invoices|bank_transactions|products|generic|expenses"),
    db: Session = Depends(get_db),
):
    claims = _get_claims(request)
    if not bool(claims.get("is_admin_company")):
        raise HTTPException(status_code=403, detail="admin_company_required")
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    module = _module_for_import_aliases(doc_type)
    rows = (
        db.query(TenantFieldConfig)
        .filter(TenantFieldConfig.tenant_id == tenant_id, TenantFieldConfig.module == module)
        .order_by(TenantFieldConfig.ord.asc().nulls_last(), TenantFieldConfig.field.asc())
        .all()
    )
    return {
        "doc_type": doc_type,
        "module": module,
        "fields": [
            {
                "field": r.field,
                "aliases": r.aliases or [],
                "field_type": r.field_type,
                "required": bool(r.required),
            }
            for r in rows
        ],
    }


@router.put("/field-aliases")
def put_import_field_aliases(
    payload: ImportAliasConfigIn,
    request: Request,
    db: Session = Depends(get_db),
):
    claims = _get_claims(request)
    if not bool(claims.get("is_admin_company")):
        raise HTTPException(status_code=403, detail="admin_company_required")
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    module = _module_for_import_aliases(payload.doc_type)
    db.query(TenantFieldConfig).filter(
        TenantFieldConfig.tenant_id == tenant_id,
        TenantFieldConfig.module == module,
    ).delete()
    for idx, item in enumerate(payload.fields):
        field = (item.field or "").strip()
        if not field:
            continue
        aliases = [str(a).strip() for a in (item.aliases or []) if str(a).strip()]
        row = TenantFieldConfig(
            tenant_id=tenant_id,
            module=module,
            field=field,
            visible=True,
            required=bool(item.required),
            ord=idx + 1,
            aliases=aliases,
            field_type=item.field_type,
        )
        db.add(row)
    db.commit()
    return {"ok": True, "doc_type": payload.doc_type, "module": module}


@router.post("/column-mappings")
def create_column_mapping_endpoint(
    dto: CreateColumnMappingDTO,
    request: Request,
    db: Session = Depends(get_db),
):
    """Create a saved column mapping for reuse."""
    from uuid import uuid4

    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    user_id = claims.get("user_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    m = ImportColumnMapping(
        id=uuid4(),
        tenant_id=tenant_id,
        name=dto.name.strip(),
        description=(dto.description or None),
        file_pattern=(dto.file_pattern or None),
        mapping=dict(dto.mapping or {}),
        transforms=(dto.transforms or None),
        defaults=(dto.defaults or None),
        created_by=user_id,
    )
    db.add(m)
    db.commit()
    db.refresh(m)

    # Learn from manually saved mapping templates (user-curated signal).
    try:
        from app.modules.imports.domain.mapping_feedback import MappingFeedback, mapping_learner

        inferred_doc_type = "products"
        target_values = {str(v) for v in (dto.mapping or {}).values() if v and v != "ignore"}
        if any(v in target_values for v in ("expense_date", "amount", "payment_method")):
            inferred_doc_type = "expenses"
        elif any(v in target_values for v in ("invoice_number", "issue_date", "vendor_name")):
            inferred_doc_type = "invoices"
        elif any(v in target_values for v in ("value_date", "narrative", "balance")):
            inferred_doc_type = "bank_transactions"

        feedback = MappingFeedback(
            tenant_id=tenant_id,
            doc_type=inferred_doc_type,
            headers=list((dto.mapping or {}).keys()),
        )
        for source_field, canonical in (dto.mapping or {}).items():
            if canonical and canonical != "ignore":
                feedback.mark_field_correct(str(source_field), str(canonical), confidence=0.85)
        if feedback.field_feedbacks:
            mapping_learner.record_feedback(feedback)
    except Exception as e:
        logger.warning(f"mapping learning skipped on save-mapping: {e}")

    return {
        "id": str(m.id),
        "name": m.name,
        "description": m.description,
        "mapping": m.mapping or {},
        "transforms": m.transforms or {},
        "defaults": m.defaults or {},
        "file_pattern": m.file_pattern,
        "use_count": int(m.use_count or 0),
        "last_used_at": m.last_used_at.isoformat() if m.last_used_at else None,
        "created_at": m.created_at.isoformat() if m.created_at else None,
    }


@router.delete("/column-mappings/{mapping_id}")
def delete_column_mapping_endpoint(
    mapping_id: UUID, request: Request, db: Session = Depends(get_db)
):
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    m = (
        db.query(ImportColumnMapping)
        .filter(ImportColumnMapping.id == mapping_id, ImportColumnMapping.tenant_id == tenant_id)
        .first()
    )
    if not m:
        raise HTTPException(status_code=404, detail="mapping_not_found")
    db.delete(m)
    db.commit()
    return {"ok": True}


@router.patch("/batches/{batch_id}/items/{item_id}", response_model=ItemOut)
def patch_batch_item_endpoint(
    batch_id: UUID,
    item_id: UUID,
    patch: ItemPatch,
    request: Request,
    db: Session = Depends(get_db),
):
    from app.modules.imports.application import use_cases

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    user_id = claims.get("user_id")
    it = use_cases.patch_item(db, tenant_id, user_id, batch_id, item_id, patch.field, patch.value)
    if not it:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    return it


@router.delete("/batches/{batch_id}/items/{item_id}")
def delete_batch_item_endpoint(
    batch_id: UUID,
    item_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
):
    """Elimina un item del batch (soft delete)"""
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    # Verificar que el item existe y pertenece al tenant
    item = (
        db.query(ImportItem)
        .filter(
            ImportItem.id == item_id,
            ImportItem.batch_id == batch_id,
            ImportItem.tenant_id == tenant_id,
        )
        .first()
    )

    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")

    # Eliminar el item
    db.delete(item)
    db.commit()

    return {"status": "ok", "message": "Item eliminado"}


@router.post("/items/delete-multiple")
def delete_multiple_items_endpoint(
    payload: dict,
    request: Request,
    db: Session = Depends(get_db),
):
    """Elimina mÃºltiples items por sus IDs"""
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    item_ids = payload.get("item_ids", [])
    if not item_ids:
        raise HTTPException(status_code=400, detail="item_ids requerido")

    # Convertir a UUID
    try:
        uuid_ids = [UUID(str(id)) for id in item_ids]
    except Exception as e:
        raise HTTPException(status_code=400, detail="IDs invÃ¡lidos") from e

    # Eliminar items
    deleted = (
        db.query(ImportItem)
        .filter(ImportItem.id.in_(uuid_ids), ImportItem.tenant_id == tenant_id)
        .delete(synchronize_session=False)
    )

    db.commit()

    return {
        "status": "ok",
        "deleted": deleted,
        "message": f"{deleted} items eliminados",
    }


@router.post("/batches/{batch_id}/validate", response_model=list[ItemOut])
def validate_batch_endpoint(batch_id: UUID, request: Request, db: Session = Depends(get_db)):
    from app.modules.imports.application import use_cases

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    return use_cases.revalidate_batch(db, tenant_id, batch_id)


@router.delete("/batches/{batch_id}", response_model=OkResponse)
def delete_batch_endpoint(batch_id: UUID, request: Request, db: Session = Depends(get_db)):
    """Eliminar un lote completo y sus dependencias (items, attachments, lineage, corrections).

    Requiere contexto de tenant (RLS). Opera con seguridad multi-tenant.
    """
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    ensure_rls(request, db)

    batch = (
        db.query(ImportBatch)
        .filter(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="batch_not_found")

    # Obtener ids de items del lote para eliminar relaciones
    item_ids = [
        it.id for it in db.query(ImportItem.id).filter(ImportItem.batch_id == batch_id).all()
    ]

    if item_ids:
        db.query(ImportAttachment).filter(ImportAttachment.item_id.in_(item_ids)).delete(
            synchronize_session=False
        )
        # lineage
        db.query(ImportLineage).filter(ImportLineage.item_id.in_(item_ids)).delete(
            synchronize_session=False
        )
        # corrections
        db.query(ImportItemCorrection).filter(ImportItemCorrection.item_id.in_(item_ids)).delete(
            synchronize_session=False
        )
        # items
        db.query(ImportItem).filter(ImportItem.id.in_(item_ids)).delete(synchronize_session=False)

    # batch
    db.delete(batch)
    db.commit()

    return OkResponse(ok=True)


@router.post("/items/promote")
def promote_items_endpoint(
    payload: PromoteItems,
    request: Request,
    db: Session = Depends(get_db),
    auto: bool = Query(default=True, description="Modo automÃ¡tico"),
    target_warehouse: str | None = Query(default="ALM-1"),
    create_warehouse: bool = Query(default=True),
    allow_missing_price: bool = Query(default=True),
    activate: bool = Query(default=True),
    create_purchase: bool = Query(default=False),
    supplier_id: str | None = Query(default=None),
    purchase_date: str | None = Query(default=None),
    purchase_status: str | None = Query(default="received"),
):
    """Promote specific items by their IDs (works across batches)."""
    from app.modules.imports.application.status import ImportItemStatus
    from app.modules.imports.domain.handlers import ProductHandler

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    tenant_id_str = claims.get("tenant_uuid")

    # tenant_id already from claims (converted to UUID string above)
    if not tenant_id_str:
        # Use tenant_id from claims directly (already UUID)
        tenant_id_str = str(tenant_id) if tenant_id else None

    if not tenant_id_str:
        raise HTTPException(status_code=400, detail="tenant_id not found")

    _tenant_uuid = UUID(tenant_id_str)
    ensure_rls(request, db)
    # Asegurar que la sesión arranca limpia por si viene de un intento previo fallido
    try:
        db.rollback()
    except Exception:
        pass
    # Reaplicar GUC de RLS después del rollback inicial
    try:
        db.execute(text("SET LOCAL app.tenant_id = :tid"), {"tid": str(tenant_id)})
        _uid = claims.get("user_id") or claims.get("tenant_user_id")
        if _uid:
            db.execute(text("SET LOCAL app.user_id = :uid"), {"uid": str(_uid)})
    except Exception:
        pass

    # Set RLS context for product inserts
    db.execute(text("SET LOCAL app.tenant_id = :tid"), {"tid": str(tenant_id)})

    item_ids = payload.item_ids
    if not item_ids:
        raise HTTPException(status_code=400, detail="item_ids required")

    promoted_count = 0
    errors = []
    purchase_lines: list[dict] = []

    opts = {
        "allow_missing_price": allow_missing_price if auto else False,
        "activate": activate if auto else False,
        "target_warehouse": target_warehouse if auto else None,
        "create_missing_warehouses": create_warehouse if auto else False,
    }
    if create_purchase:
        opts["skip_stock_init"] = True

    for item_id in item_ids:
        try:
            # Usar savepoint por item para que un fallo no aborte toda la transacción
            with db.begin_nested():
                item = db.query(ImportItem).filter_by(id=UUID(item_id)).first()
                if not item:
                    errors.append({"item_id": item_id, "error": "not_found"})
                    continue

                # Promote individual item using ProductHandler
                # Normalize keys to lowercase for handler compatibility
                raw = item.raw or {}
                normalized = item.normalized or {}
                src = {k.lower(): v for k, v in raw.items()}
                src.update({k.lower(): v for k, v in normalized.items()})

                # Map common Spanish column names
                if "producto" in src and "name" not in src:
                    src["name"] = src["producto"]
                if "precio unitario venta" in src and "precio" not in src:
                    src["precio"] = src["precio unitario venta"]
                if "cantidad" not in src and "sobrante diario" in src:
                    src["cantidad"] = src["sobrante diario"]

                # Skip rows without name (category headers)
                if not src.get("name") or not src.get("name").strip():
                    errors.append(
                        {
                            "item_id": item_id,
                            "error": "missing_name",
                            "sku": src.get("sku") or src.get("codigo"),
                            "name": src.get("name"),
                        }
                    )
                    continue

                result = ProductHandler.promote(
                    db, tenant_id, src, promoted_id=item.promoted_id, options=opts
                )

                if result.domain_id:
                    item.status = ImportItemStatus.PROMOTED
                    item.promoted_id = result.domain_id
                    db.add(item)
                    # Detectar fallos de integridad en el acto (no esperar al commit global)
                    db.flush()
                    try:
                        audit_user = (
                            claims.get("user_id")
                            or claims.get("tenant_user_id")
                            or claims.get("user_uuid")
                        )
                        if not audit_user:
                            audit_user = str(tenant_id)
                        audit_payload = {
                            "action": "promote",
                            "product_id": str(result.domain_id),
                            "item_id": str(item.id),
                            "batch_id": str(item.batch_id),
                            "sku": src.get("sku") or src.get("codigo"),
                            "name": src.get("name"),
                        }
                        db.execute(
                            text(
                                "INSERT INTO import_audits "
                                "(document_id, batch_id, item_id, tenant_id, user_id, changes) "
                                "VALUES (:document_id, :batch_id, :item_id, :tenant_id, :user_id, "
                                "CAST(:changes AS jsonb))"
                            ),
                            {
                                "document_id": 0,
                                "batch_id": str(item.batch_id),
                                "item_id": str(item.id),
                                "tenant_id": str(tenant_id),
                                "user_id": str(audit_user),
                                "changes": json.dumps(audit_payload, ensure_ascii=False),
                            },
                        )
                    except Exception:
                        # Audit is best-effort; never block promotion
                        pass
                    promoted_count += 1
                    if create_purchase:
                        line = _build_purchase_line(src, result.domain_id)
                        if line:
                            purchase_lines.append(line)
                else:
                    errors.append(
                        {
                            "item_id": item_id,
                            "error": "promotion_failed",
                            "sku": src.get("sku") or src.get("codigo"),
                            "name": src.get("name"),
                        }
                    )
        except Exception as e:
            # El savepoint se revierte automáticamente; limpiamos el estado de sesión y continuamos
            try:
                db.rollback()
            except Exception:
                pass
            # Intenta exponer la causa real (e.orig en SQLAlchemy)
            real_error = getattr(e, "orig", None) or e
            trace = None
            try:
                import traceback as _tb

                trace = _tb.format_exc()
            except Exception:
                trace = None
            errors.append(
                {
                    "item_id": item_id,
                    "error": str(real_error),
                    "sku": locals().get("src", {}).get("sku") if "src" in locals() else None,
                    "name": locals().get("src", {}).get("name") if "src" in locals() else None,
                    "trace": trace,
                }
            )

    # Log para depurar qué falló en producción/dev sin depender del cliente
    if errors:
        logging.error(
            "Import promote items: %s errores (primeros 3: %s)",
            len(errors),
            errors[:3],
        )

    db.commit()

    purchase_result = None
    if create_purchase:
        try:
            warehouse = _resolve_purchase_warehouse(
                db,
                tenant_id,
                target_warehouse,
                create_warehouse,
            )
            purchase_result = _create_purchase_from_lines(
                db,
                tenant_id,
                user_id=claims.get("user_id")
                or claims.get("tenant_user_id")
                or claims.get("user_uuid"),
                supplier_id=supplier_id,
                purchase_date=_parse_purchase_date(purchase_date),
                status=purchase_status or "received",
                warehouse=warehouse,
                lines=purchase_lines,
            )
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            purchase_result = {"created": False, "error": str(exc)}

    response = {"promoted": promoted_count, "total": len(item_ids), "errors": errors}
    if purchase_result is not None:
        response["purchase"] = purchase_result
    return response


@router.post("/batches/{batch_id}/promote")
def promote_batch_endpoint(
    batch_id: str,
    request: Request,
    db: Session = Depends(get_db),
    auto: bool = Query(default=True),
    target_warehouse: str | None = Query(default="ALM-1"),
    create_warehouse: bool = Query(default=True),
    allow_missing_price: bool = Query(default=True),
    activate: bool = Query(default=True),
    create_purchase: bool = Query(default=False),
    supplier_id: str | None = Query(default=None),
    purchase_date: str | None = Query(default=None),
    purchase_status: str | None = Query(default="received"),
    post_accounting: bool = Query(
        default=False, description="Generar asiento contable automaticamente al promover"
    ),
    payment_status: str | None = Query(
        default=None,
        description="Estado de pago al promover (pending|paid). Si no se indica, se mantiene pending.",
        pattern="^(pending|paid)$",
    ),
    payment_method: str | None = Query(
        default=None,
        description="Metodo de pago si payment_status=paid (cash|bank|card|transfer|direct_debit|check|other)",
        pattern="^(cash|bank|card|transfer|direct_debit|check|other)$",
    ),
    paid_at: str | None = Query(
        default=None, description="Fecha de pago YYYY-MM-DD (si payment_status=paid)"
    ),
    save_as_products: str | None = Query(
        default=None, description="Para recetas: guardar ingredientes como productos (true|false)"
    ),
):
    from app.modules.imports.application import use_cases
    from app.modules.imports.infrastructure.repositories import ImportsRepository

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    # Special handling for recipes
    def to_uuid_safe(value):
        try:
            return value if isinstance(value, UUID) else UUID(str(value))
        except Exception:
            return None

    repo = ImportsRepository()
    batch_uuid = to_uuid_safe(batch_id)
    if not batch_uuid:
        raise HTTPException(status_code=400, detail="Invalid batch_id format")
    batch = repo.get_batch(db, tenant_id, batch_uuid)
    if batch and batch.source_type == "recipes":
        from app.modules.imports.application.tasks.task_import_file import (
            _file_path_from_key,
            _persist_recipes,
        )
        from app.modules.imports.parsers.xlsx_recipes import parse_xlsx_recipes

        file_path = _file_path_from_key(batch.file_key) if batch.file_key else None
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=400, detail="Recipe file not found for batch")

        parsed = parse_xlsx_recipes(file_path)
        result = _persist_recipes(db, tenant_id, parsed)

        # Optionally save ingredients as products
        if save_as_products == "true" and result.get("created", 0) > 0:
            # Already handled by _persist_recipes
            pass

        batch.status = "READY" if result.get("errors", 0) == 0 else "PARTIAL"
        db.add(batch)
        db.commit()
        return {
            "created": result.get("created", 0),
            "skipped": 0,
            "failed": result.get("errors", 0),
        }

    options = {
        "allow_missing_price": allow_missing_price if auto else False,
        "activate": activate if auto else False,
        "target_warehouse": target_warehouse if auto else None,
        "create_missing_warehouses": create_warehouse if auto else False,
        "post_accounting": bool(post_accounting) if auto else False,
        "user_id": claims.get("user_id") or claims.get("tenant_user_id") or claims.get("user_uuid"),
        "payment_status": payment_status,
        "payment_method": payment_method,
        "paid_at": paid_at,
        "save_as_products": save_as_products == "true" if save_as_products else False,
    }
    if create_purchase:
        options["skip_stock_init"] = True
        options["collect_promoted"] = True
    res = use_cases.promote_batch(db, tenant_id, batch_id, options=options)
    purchase_result = None
    if create_purchase:
        try:
            promoted_items = res.get("promoted_items") or []
            purchase_lines = []
            for item in promoted_items:
                line = _build_purchase_line(item.get("src") or {}, item.get("promoted_id"))
                if line:
                    purchase_lines.append(line)
            warehouse = _resolve_purchase_warehouse(
                db,
                tenant_id,
                target_warehouse,
                create_warehouse,
            )
            purchase_result = _create_purchase_from_lines(
                db,
                tenant_id,
                user_id=claims.get("user_id")
                or claims.get("tenant_user_id")
                or claims.get("user_uuid"),
                supplier_id=supplier_id,
                purchase_date=_parse_purchase_date(purchase_date),
                status=purchase_status or "received",
                warehouse=warehouse,
                lines=purchase_lines,
            )
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            purchase_result = {"created": False, "error": str(exc)}
    res.pop("promoted_items", None)
    if purchase_result is not None:
        res["purchase"] = purchase_result
    return res


@router.get("/batches/{batch_id}/errors.csv")
def errors_csv_endpoint(batch_id: UUID, request: Request, db: Session = Depends(get_db)):
    import csv
    from io import StringIO

    from app.modules.imports.infrastructure.repositories import ImportsRepository

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    repo = ImportsRepository()
    items = repo.list_items(db, tenant_id, batch_id, status="ERROR_VALIDATION")
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(["idx", "campo", "error", "valor"])
    for it in items:
        errs = it.errors or []
        for e in errs:
            field = e.get("field") or e.get("campo") or ""
            msg = e.get("msg") or e.get("error") or ""
            val = None
            if it.normalized and field in it.normalized:
                val = it.normalized.get(field)
            elif it.raw and field in it.raw:
                val = it.raw.get(field)
            writer.writerow([it.idx, field, msg, val])
    csv_text = buf.getvalue()
    buf.close()
    from fastapi import Response

    return Response(content=csv_text, media_type="text/csv")


# ------------------------------
# Acciones masivas y reproceso
# ------------------------------

# payloads tipados de forma dinÃƒÆ’Ã‚Â¡mica en bulk_patch_items


@router.post("/batches/{batch_id}/items/bulk-patch", response_model=OkResponse)
def bulk_patch_items(
    batch_id: UUID,
    payload: dict,
    request: Request,
    db: Session = Depends(get_db),
):
    from app.modules.imports.application import use_cases

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    user_id = claims.get("user_id")

    ids = payload.get("ids") or []
    changes = payload.get("changes") or {}
    if not isinstance(ids, list) or not isinstance(changes, dict) or not ids or not changes:
        raise HTTPException(
            status_code=400, detail="Body invalido: {ids:[], changes:{campo:valor}}"
        )

    # Aplicar cambios en cada id, campo por campo
    for item_id in ids:
        for field, value in changes.items():
            use_cases.patch_item(db, tenant_id, user_id, batch_id, item_id, field, value)

    # Revalidar lote completo
    use_cases.revalidate_batch(db, tenant_id, batch_id)
    return OkResponse()


@router.post("/batches/{batch_id}/reprocess", response_model=OkResponse)
def reprocess_batch(
    batch_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
    scope: str | None = Query(default="all", pattern="^(all|errors)$"),
    promote: bool | None = Query(default=True),
):
    from app.modules.imports.application import use_cases

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    # Revalidar: para simplificar se revalida todo el lote; en futuras mejoras se puede filtrar
    use_cases.revalidate_batch(db, tenant_id, batch_id)

    # Promover sÃƒÆ’Ã‚Â³lo vÃƒÆ’Ã‚Â¡lidos si promote=true
    if promote:
        use_cases.promote_batch(db, tenant_id, batch_id)
    return OkResponse()


# ------------------------------
# Excel Column Analyzer - Smart Import
# ------------------------------


@router.post("/_legacy/analyze-file")
async def analyze_import_file(
    file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db)
):
    """
    Analiza un archivo Excel y detecta columnas automÃ¡ticamente.
    Retorna:
    - headers: Lista de columnas detectadas
    - sample_data: Primeras 5 filas de ejemplo
    - suggested_mapping: Sugerencias automÃ¡ticas de mapeo
    - saved_mappings: Mapeos guardados por este tenant (para reutilizar)
    """
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    # Validar tipo de archivo
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")

    try:
        # Analizar Excel
        analysis = analyze_excel_stream(file.file)

        # Obtener mapeos guardados del tenant
        from app.models.imports import ImportColumnMapping

        saved_mappings = (
            db.query(ImportColumnMapping)
            .filter(
                ImportColumnMapping.tenant_id == tenant_id,
                ImportColumnMapping.is_active,
            )
            .order_by(ImportColumnMapping.last_used_at.desc().nullslast())
            .all()
        )

        return {
            "headers": analysis["headers"],
            "header_row": analysis["header_row"],
            "sample_data": analysis["sample_data"],
            "suggested_mapping": analysis["suggested_mapping"],
            "total_rows": analysis["total_rows"],
            "total_columns": analysis["total_columns"],
            "saved_mappings": [
                {
                    "id": str(m.id),
                    "name": m.name,
                    "description": m.description,
                    "mapping": m.mapping,
                    "use_count": m.use_count,
                    "last_used_at": m.last_used_at.isoformat() if m.last_used_at else None,
                }
                for m in saved_mappings
            ],
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error analizando archivo: {str(e)}") from e


# ------------------------------
# CRUD ImportMappings (plantillas)
# ------------------------------


@router.post("/_legacy/column-mappings")
def create_column_mapping(
    name: str,
    mapping: dict[str, str],
    description: str = None,
    file_pattern: str = None,
    transforms: dict | None = None,
    defaults: dict | None = None,
    request: Request = None,
    db: Session = Depends(get_db),
):
    """
    Guarda una configuraciÃ³n de mapeo de columnas para reutilizar
    Ejemplo:
    {
      "name": "Proveedor LÃ¡cteos",
      "description": "Excel mensual de stock",
      "mapping": {"FORMATO...": "ignore", "Precio $": "precio", ...}
    }
    """
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    user_id = claims.get("user_id")

    from uuid import uuid4

    from app.models.imports import ImportColumnMapping

    # Verificar que no exista con mismo nombre
    existing = (
        db.query(ImportColumnMapping)
        .filter(ImportColumnMapping.tenant_id == tenant_id, ImportColumnMapping.name == name)
        .first()
    )

    if existing:
        raise HTTPException(status_code=409, detail=f"Ya existe un mapeo con el nombre '{name}'")

    obj = ImportColumnMapping(
        id=uuid4(),
        tenant_id=tenant_id,
        name=name,
        description=description,
        file_pattern=file_pattern,
        mapping=mapping,
        transforms=transforms,
        defaults=defaults,
        created_by=user_id,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)

    return {
        "id": str(obj.id),
        "name": obj.name,
        "description": obj.description,
        "mapping": obj.mapping,
        "transforms": obj.transforms,
        "defaults": obj.defaults,
        "created_at": obj.created_at.isoformat(),
    }


@router.get("/_legacy/column-mappings")
def list_column_mappings(request: Request, db: Session = Depends(get_db)):
    """Lista todos los mapeos guardados del tenant"""
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    from app.models.imports import ImportColumnMapping

    mappings = (
        db.query(ImportColumnMapping)
        .filter(
            ImportColumnMapping.tenant_id == tenant_id,
            ImportColumnMapping.is_active,
        )
        .order_by(ImportColumnMapping.last_used_at.desc().nullslast())
        .all()
    )

    return [
        {
            "id": str(m.id),
            "name": m.name,
            "description": m.description,
            "mapping": m.mapping,
            "transforms": m.transforms or {},
            "defaults": m.defaults or {},
            "file_pattern": m.file_pattern,
            "use_count": m.use_count,
            "last_used_at": m.last_used_at.isoformat() if m.last_used_at else None,
            "created_at": m.created_at.isoformat(),
        }
        for m in mappings
    ]


@router.delete("/_legacy/column-mappings/{mapping_id}")
def delete_column_mapping(mapping_id: str, request: Request, db: Session = Depends(get_db)):
    """Elimina (soft delete) un mapeo guardado"""
    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    from uuid import UUID

    from app.models.imports import ImportColumnMapping

    obj = (
        db.query(ImportColumnMapping)
        .filter(
            ImportColumnMapping.id == UUID(mapping_id),
            ImportColumnMapping.tenant_id == tenant_id,
        )
        .first()
    )

    if not obj:
        raise HTTPException(status_code=404, detail="Mapeo no encontrado")

    obj.is_active = False
    db.commit()

    return {"status": "ok", "message": "Mapeo eliminado"}


@router.post("/mappings", response_model=ImportMappingOut)
def create_mapping(dto: ImportMappingCreate, request: Request, db: Session = Depends(get_db)):
    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    user_id = claims.get("user_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    from uuid import uuid4

    obj = ImportColumnMapping(
        id=uuid4(),
        tenant_id=tenant_id,
        name=dto.name,
        description=f"compat:/mappings source_type={dto.source_type}",
        mapping=dto.mappings or {},
        transforms=dto.transforms or {},
        defaults=dto.defaults or {},
        created_by=user_id,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return {
        "id": obj.id,
        "tenant_id": obj.tenant_id,
        "name": obj.name,
        "source_type": dto.source_type,
        "version": dto.version or 1,
        "mappings": obj.mapping or {},
        "transforms": obj.transforms or {},
        "defaults": obj.defaults or {},
        "dedupe_keys": dto.dedupe_keys,
        "created_at": obj.created_at,
    }


@router.get("/mappings", response_model=list[ImportMappingOut])
def list_mappings(request: Request, db: Session = Depends(get_db), source_type: str | None = None):
    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    items = (
        db.query(ImportColumnMapping)
        .filter(ImportColumnMapping.tenant_id == tenant_id, ImportColumnMapping.is_active)  # noqa: E712
        .order_by(ImportColumnMapping.created_at.desc())
        .all()
    )
    out = []
    for m in items:
        mapped_source_type = source_type or "generic"
        if source_type and m.description:
            marker = f"source_type={source_type}"
            if marker not in m.description:
                continue
        out.append(
            {
                "id": m.id,
                "tenant_id": m.tenant_id,
                "name": m.name,
                "source_type": mapped_source_type,
                "version": 1,
                "mappings": m.mapping or {},
                "transforms": m.transforms or {},
                "defaults": m.defaults or {},
                "dedupe_keys": None,
                "created_at": m.created_at,
            }
        )
    return out


@router.get("/mappings/{mapping_id}", response_model=ImportMappingOut)
def get_mapping(mapping_id: UUID, request: Request, db: Session = Depends(get_db)):
    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    obj = (
        db.query(ImportColumnMapping)
        .filter(
            ImportColumnMapping.id == mapping_id,
            ImportColumnMapping.tenant_id == tenant_id,
            ImportColumnMapping.is_active,
        )
        .first()
    )
    if not obj:
        raise HTTPException(status_code=404, detail="Mapping no encontrado")
    source_type = "generic"
    if obj.description and "source_type=" in obj.description:
        try:
            source_type = obj.description.split("source_type=", 1)[1].split()[0].strip()
        except Exception:
            source_type = "generic"
    return {
        "id": obj.id,
        "tenant_id": obj.tenant_id,
        "name": obj.name,
        "source_type": source_type,
        "version": 1,
        "mappings": obj.mapping or {},
        "transforms": obj.transforms or {},
        "defaults": obj.defaults or {},
        "dedupe_keys": None,
        "created_at": obj.created_at,
    }


@router.put("/mappings/{mapping_id}", response_model=ImportMappingOut)
def update_mapping(
    mapping_id: UUID,
    dto: ImportMappingUpdate,
    request: Request,
    db: Session = Depends(get_db),
):
    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    obj = (
        db.query(ImportColumnMapping)
        .filter(
            ImportColumnMapping.id == mapping_id,
            ImportColumnMapping.tenant_id == tenant_id,
            ImportColumnMapping.is_active,
        )
        .first()
    )
    if not obj:
        raise HTTPException(status_code=404, detail="Mapping no encontrado")
    if dto.name is not None:
        obj.name = dto.name
    if dto.mappings is not None:
        obj.mapping = dto.mappings
    if dto.transforms is not None:
        obj.transforms = dto.transforms
    if dto.defaults is not None:
        obj.defaults = dto.defaults
    db.add(obj)
    db.commit()
    db.refresh(obj)
    source_type = "generic"
    if obj.description and "source_type=" in obj.description:
        try:
            source_type = obj.description.split("source_type=", 1)[1].split()[0].strip()
        except Exception:
            source_type = "generic"
    return {
        "id": obj.id,
        "tenant_id": obj.tenant_id,
        "name": obj.name,
        "source_type": source_type,
        "version": dto.version or 1,
        "mappings": obj.mapping or {},
        "transforms": obj.transforms or {},
        "defaults": obj.defaults or {},
        "dedupe_keys": dto.dedupe_keys,
        "created_at": obj.created_at,
    }


@router.post("/mappings/{mapping_id}/clone", response_model=ImportMappingOut)
def clone_mapping(mapping_id: UUID, request: Request, db: Session = Depends(get_db)):
    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    from uuid import uuid4

    src = (
        db.query(ImportColumnMapping)
        .filter(
            ImportColumnMapping.id == mapping_id,
            ImportColumnMapping.tenant_id == tenant_id,
            ImportColumnMapping.is_active,
        )
        .first()
    )
    if not src:
        raise HTTPException(status_code=404, detail="Mapping no encontrado")
    clone = ImportColumnMapping(
        id=uuid4(),
        tenant_id=tenant_id,
        name=f"{src.name} (copia)",
        description=src.description,
        mapping=src.mapping,
        transforms=src.transforms,
        defaults=src.defaults,
        created_by=claims.get("user_id"),
    )
    db.add(clone)
    db.commit()
    db.refresh(clone)
    source_type = "generic"
    if clone.description and "source_type=" in clone.description:
        try:
            source_type = clone.description.split("source_type=", 1)[1].split()[0].strip()
        except Exception:
            source_type = "generic"
    return {
        "id": clone.id,
        "tenant_id": clone.tenant_id,
        "name": clone.name,
        "source_type": source_type,
        "version": 1,
        "mappings": clone.mapping or {},
        "transforms": clone.transforms or {},
        "defaults": clone.defaults or {},
        "dedupe_keys": None,
        "created_at": clone.created_at,
    }


@router.delete("/mappings/{mapping_id}", response_model=OkResponse)
def delete_mapping(mapping_id: UUID, request: Request, db: Session = Depends(get_db)):
    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    obj = (
        db.query(ImportColumnMapping)
        .filter(
            ImportColumnMapping.id == mapping_id,
            ImportColumnMapping.tenant_id == tenant_id,
            ImportColumnMapping.is_active,
        )
        .first()
    )
    if not obj:
        raise HTTPException(status_code=404, detail="Mapping no encontrado")
    obj.is_active = False
    db.add(obj)
    db.commit()
    return OkResponse()


# ------------------------------
# Fotos: subir y adjuntar con OCR
# ------------------------------


@router.post("/batches/{batch_id}/photos", response_model=ItemOut)
async def upload_photo_to_batch(
    batch_id: UUID,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    from app.modules.imports.application import use_cases

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    user_id = claims.get("user_id")

    batch = (
        db.query(ImportBatch)
        .filter(ImportBatch.id == batch_id, ImportBatch.tenant_id == tenant_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Batch no encontrado")

    # Limit file size (read & rewind) + mimetype
    max_mb = float(os.getenv("IMPORTS_MAX_UPLOAD_MB", "10"))
    max_bytes = int(max_mb * 1024 * 1024)
    content = await file.read()
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"File too large: {len(content)} bytes > {max_bytes}",
        )
    allowed = {"image/jpeg", "image/png", "image/heic", "application/pdf"}
    if (file.content_type or "") not in allowed:
        raise HTTPException(status_code=422, detail=f"Unsupported file type: {file.content_type}")
    file.file.seek(0)

    # Crea item a partir de la foto + adjunta OCR y metadata
    # Size + mimetype limits and throttling handled above
    item = use_cases.ingest_photo(db, str(tenant_id), str(user_id), batch, file)
    return item


@router.post("/batches/{batch_id}/items/{item_id}/photos", response_model=ItemOut)
async def attach_photo_to_item(
    batch_id: UUID,
    item_id: UUID,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    from app.models.core.modelsimport import ImportItem
    from app.modules.imports.application import use_cases

    claims = _get_claims(request)
    # tenant_id is now UUID, not int
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")
    user_id = claims.get("user_id")

    # Limit file size + mimetype
    max_mb = float(os.getenv("IMPORTS_MAX_UPLOAD_MB", "10"))
    max_bytes = int(max_mb * 1024 * 1024)
    content = await file.read()
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"File too large: {len(content)} bytes > {max_bytes}",
        )
    allowed = {"image/jpeg", "image/png", "image/heic", "application/pdf"}
    if (file.content_type or "") not in allowed:
        raise HTTPException(status_code=422, detail=f"Unsupported file type: {file.content_type}")
    file.file.seek(0)

    # Verificar scoping: item pertenece al batch y al tenant
    it = (
        db.query(ImportItem)
        .join(ImportBatch, ImportItem.batch_id == ImportBatch.id)
        .filter(
            ImportItem.id == item_id,
            ImportBatch.id == batch_id,
            ImportBatch.tenant_id == tenant_id,
        )
        .first()
    )
    if not it:
        raise HTTPException(status_code=404, detail="Item no encontrado")

    # Limit file size (read & rewind)
    max_mb = float(os.getenv("IMPORTS_MAX_UPLOAD_MB", "10"))
    max_bytes = int(max_mb * 1024 * 1024)
    content = await file.read()
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"File too large: {len(content)} bytes > {max_bytes}",
        )
    file.file.seek(0)

    # Size + mimetype limits and throttling handled above
    item = use_cases.attach_photo_and_reocr(db, str(tenant_id), str(user_id), it, file)
    return item


# ------------------------------
# Helpers: key normalization + fuzzy field lookup
# ------------------------------


def _norm_key(s: str) -> str:
    try:
        s = unicodedata.normalize("NFKD", s)
        # Normalizar espacios: NBSP y mÃºltiples espacios â†’ espacio simple
        s = str(s).replace("\u00a0", " ")
        s = re.sub(r"\s+", " ", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = s.strip().lower()
        s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
        return s
    except Exception:
        return str(s).strip().lower()


def _build_norm_map(data: dict) -> dict:
    nm = {}
    for k, v in (data or {}).items():
        if isinstance(k, str):
            nm[_norm_key(k)] = v
    return nm


def _first(nmap: dict, keys: list[str]):
    for k in keys:
        nk = _norm_key(k)
        if nk in nmap and nmap[nk] not in (None, ""):
            return nmap[nk]
    return None


def _parse_num(val):
    if val is None:
        return None
    if isinstance(val, int | float):
        return float(val)
    try:
        s = str(val).strip()
        if s == "":
            return None
        s = s.replace("\u00a0", " ")
        s = s.replace(" ", "")
        # Heuristic: if there is one comma and at least one dot, and comma comes after dot: treat dot as thousands
        if s.count(",") == 1 and s.count(".") >= 1 and s.find(",") > s.find("."):
            s = s.replace(".", "").replace(",", ".")
        elif s.count(",") == 1 and "." not in s:
            s = s.replace(",", ".")
        s = re.sub(r"[^0-9\.-]", "", s)
        if s in ("", "-", ".", "-."):
            return None
        return float(s)
    except Exception:
        return None


def _to_uuid_or_none(value) -> UUID | None:
    try:
        return value if isinstance(value, UUID) else UUID(str(value))
    except Exception:
        return None


def _dec(value: float | Decimal | None, q: str = "0.000001") -> Decimal:
    if value is None:
        value = 0
    return Decimal(str(value)).quantize(Decimal(q), rounding=ROUND_HALF_UP)


def _parse_purchase_date(value: str | None) -> date:
    if not value:
        return date.today()
    try:
        return date.fromisoformat(value)
    except Exception:
        return date.today()


def _resolve_purchase_warehouse(
    db: Session,
    tenant_id,
    target_code: str | None,
    create_missing: bool,
) -> Warehouse:
    wh = None
    if target_code:
        wh = (
            db.query(Warehouse)
            .filter(
                Warehouse.tenant_id == tenant_id,
                Warehouse.code == target_code,
            )
            .first()
        )
        if not wh and create_missing:
            wh = Warehouse(
                tenant_id=tenant_id,
                code=target_code,
                name=target_code,
                is_active=True,
            )
            db.add(wh)
            db.flush()
    if not wh:
        wh = (
            db.query(Warehouse)
            .filter(Warehouse.tenant_id == tenant_id)
            .order_by(Warehouse.code.asc())
            .first()
        )
    if not wh:
        wh = Warehouse(
            tenant_id=tenant_id,
            code="ALM-1",
            name="Principal",
            is_active=True,
        )
        db.add(wh)
        db.flush()
    return wh


def _build_purchase_line(src: dict, product_id: str | None) -> dict | None:
    if not product_id:
        return None
    nmap = _build_norm_map(src or {})
    qty = _parse_num(
        _first(
            nmap,
            [
                "stock",
                "cantidad",
                "quantity",
                "existencia",
                "existencias",
                "unidades",
                "sobrante_diario",
            ],
        )
    )
    if qty is None or qty <= 0:
        return None
    unit_cost = _parse_num(
        _first(
            nmap,
            [
                "costo",
                "cost",
                "cost_price",
                "unit_cost",
                "purchase_price",
                "precio_compra",
                "precio_costo",
            ],
        )
    )
    if unit_cost is None:
        unit_cost = 0
    return {
        "product_id": str(product_id),
        "qty": float(qty),
        "unit_cost": float(unit_cost),
    }


def _create_purchase_from_lines(
    db: Session,
    tenant_id,
    *,
    user_id,
    supplier_id: str | None,
    purchase_date: date,
    status: str,
    warehouse: Warehouse,
    lines: list[dict],
) -> dict:
    if not lines:
        return {"created": False, "reason": "no_lines"}
    purchase_total = sum((line["qty"] or 0) * (line["unit_cost"] or 0) for line in lines)
    purchase_number = f"IMP-{purchase_date.strftime('%Y%m%d')}-{uuid4().hex[:6].upper()}"

    purchase = Purchase(
        id=uuid4(),
        tenant_id=tenant_id,
        number=purchase_number,
        supplier_id=_to_uuid_or_none(supplier_id) if supplier_id else None,
        date=purchase_date,
        subtotal=float(purchase_total),
        taxes=0,
        total=float(purchase_total),
        status=status,
        notes="Importador",
        user_id=_to_uuid_or_none(user_id) or _to_uuid_or_none(tenant_id) or uuid4(),
    )
    db.add(purchase)
    db.flush()

    costing = InventoryCostingService(db)
    for line in lines:
        product_id = str(line["product_id"])
        qty_dec = _dec(line["qty"])
        unit_cost_dec = _dec(line["unit_cost"])
        row = (
            db.query(StockItem)
            .filter(
                StockItem.warehouse_id == str(warehouse.id),
                StockItem.product_id == product_id,
            )
            .with_for_update()
            .first()
        )
        if not row:
            row = StockItem(
                tenant_id=str(tenant_id),
                warehouse_id=str(warehouse.id),
                product_id=product_id,
                qty=0,
            )
            db.add(row)
            db.flush()

        costing.apply_inbound(
            str(tenant_id),
            str(warehouse.id),
            product_id,
            qty=qty_dec,
            unit_cost=unit_cost_dec,
            initial_qty=_dec(row.qty),
            initial_avg_cost=unit_cost_dec,
        )

        row.qty = (row.qty or 0) + float(qty_dec)
        db.add(row)

        db.add(
            StockMove(
                tenant_id=str(tenant_id),
                product_id=product_id,
                warehouse_id=str(warehouse.id),
                qty=float(qty_dec),
                kind="receipt",
                tentative=False,
                posted=True,
                ref_type="purchase",
                ref_id=str(purchase.id),
                unit_cost=float(unit_cost_dec),
                total_cost=float(unit_cost_dec * qty_dec),
            )
        )

        db.add(
            PurchaseLine(
                purchase_id=purchase.id,
                product_id=_to_uuid_or_none(product_id),
                description=None,
                quantity=float(qty_dec),
                unit_price=float(unit_cost_dec),
                tax_rate=0,
                total=float(unit_cost_dec * qty_dec),
            )
        )

        try:
            total_qty = (
                db.query(func.coalesce(func.sum(StockItem.qty), 0.0))
                .filter(StockItem.product_id == product_id, StockItem.tenant_id == str(tenant_id))
                .scalar()
            ) or 0.0
            prod = (
                db.query(Product)
                .filter(Product.id == product_id, Product.tenant_id == tenant_id)
                .first()
            )
            if prod:
                prod.stock = float(total_qty)
                db.add(prod)
        except Exception:
            pass

    purchase.status = status
    db.add(purchase)
    db.commit()
    return {
        "created": True,
        "purchase_id": str(purchase.id),
        "status": purchase.status,
        "lines": len(lines),
    }


# ------------------------------
# Aliases registry (JSON-configurable)
# ------------------------------


def _load_aliases() -> dict:
    default_aliases = {
        "codigo": [
            "codigo",
            "cÃ³digo",
            "sku",
            "code",
            "cod",
            "codigo_barras",
            "ean",
            "ean13",
            "upc",
            "producto_codigo",
        ],
        "nombre": [
            "nombre",
            "name",
            "producto",
            "descripcion",
            "descripciÃ³n",
            "detalle",
            "articulo",
            "artÃ­culo",
            "nombre_articulo",
            "nombre_producto",
            "descripcion_articulo",
            "descripcion_producto",
            "desc",
        ],
        "precio": [
            "precio",
            "price",
            "pvp",
            "venta",
            "precio_unitario_venta",
            "precio_unitario",
            "precio_venta",
            "precio_lista",
            "precio_final",
            "importe",
            "valor",
        ],
        "precio_por_bulto": [
            "precio_por_bulto",
            "precio_x_bulto",
            "precio_bulto",
            "pvp_bulto",
            "venta_bulto",
        ],
        "costo": ["costo", "cost", "precio_costo", "costo_unitario"],
        "categoria": [
            "categoria",
            "category",
            "rubro",
            "familia",
            "seccion",
            "secciÃ³n",
            "grupo",
            "linea",
            "lÃ­nea",
            "subcategoria",
            "sub_categoria",
        ],
        "stock": [
            "stock",
            "cantidad",
            "existencia",
            "existencias",
            "unidades",
            "disponible",
            "cantidad_actual",
            "sobrante_diario",
        ],
        "bultos": ["bultos", "packs", "paquetes"],
        "cantidad_por_bulto": [
            "cantidad_por_bulto",
            "cant_por_bulto",
            "unidades_por_bulto",
            "cantidad_x_bulto",
            "cant_x_bulto",
            "u_x_bulto",
        ],
        "unidad": ["unidad", "uom", "un", "unidad_medida", "medida"],
        "iva": ["iva", "tax", "impuesto"],
    }
    # Optional override via env
    cfg_path = os.getenv("IMPORTS_ALIASES_FILE")
    candidates: list[Path] = []
    if cfg_path:
        candidates.append(Path(cfg_path))
    # project file default
    candidates.append(Path(__file__).parent.parent / "config" / "field_aliases_es.json")
    for p in candidates:
        try:
            if p.exists():
                with p.open("r", encoding="utf-8") as fh:
                    data = json.load(fh)
                    if isinstance(data, dict):
                        # merge with defaults to keep missing keys
                        for k, v in data.items():
                            if isinstance(v, list):
                                default_aliases[k] = v
                        break
        except Exception:
            # Ignore and keep defaults
            pass
    return default_aliases


@router.post("/batches/cleanup/stuck", response_model=OkResponse)
def cleanup_stuck_batches(
    request: Request,
    hours: int = Query(default=2, ge=1, le=24),
    db: Session = Depends(get_db),
):
    """Limpia batches atascados en estado PENDING/PARSING desde hace más de X horas.

    Parámetros:
    - hours: cuántas horas de inactividad requieren para eliminar (default: 2)
    """
    from datetime import datetime, timedelta

    claims = _get_claims(request)
    tenant_id = claims.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=401, detail="tenant_id_missing")

    ensure_rls(request, db)

    # Batches que llevan más de X horas sin cambios y están en estado processing
    cutoff_time = datetime.utcnow() - timedelta(hours=hours)

    stuck_batches = (
        db.query(ImportBatch)
        .filter(
            ImportBatch.tenant_id == tenant_id,
            ImportBatch.status.in_(["PENDING", "PARSING"]),
            ImportBatch.created_at < cutoff_time,
        )
        .all()
    )

    deleted_count = 0
    for batch in stuck_batches:
        try:
            db.delete(batch)
            deleted_count += 1
        except Exception as e:
            logger.error(f"Error deleting batch {batch.id}: {e}")

    db.commit()

    return OkResponse(
        ok=True, detail=f"Deleted {deleted_count} stuck batches (older than {hours} hours)"
    )


_ALIASES = _load_aliases()
