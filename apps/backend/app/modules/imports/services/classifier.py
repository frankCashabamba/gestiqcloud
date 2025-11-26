"""File classification service for import module."""

import csv
import logging
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

import openpyxl


class FileClassifier:
    """Classifies uploaded files to determine appropriate parser."""

    def __init__(self):
        self.logger = logging.getLogger("app.imports.classifier")
        self._ai_provider = None
        self.parsers_info = {
            "generic_excel": {
                "doc_type": "generic",
                "description": "Generic Excel parser",
                "supported_extensions": [".xlsx", ".xls"],
            },
            "products_excel": {
                "doc_type": "products",
                "description": "Products Excel parser",
                "supported_extensions": [".xlsx", ".xls"],
            },
            "csv_invoices": {
                "doc_type": "invoices",
                "description": "CSV invoices parser",
                "supported_extensions": [".csv"],
            },
            "csv_bank": {
                "doc_type": "bank_transactions",
                "description": "CSV bank transactions parser",
                "supported_extensions": [".csv"],
            },
            "xml_invoice": {
                "doc_type": "invoices",
                "description": "XML invoice parser (UBL/CFDI)",
                "supported_extensions": [".xml"],
            },
            "xml_camt053_bank": {
                "doc_type": "bank_transactions",
                "description": "ISO 20022 CAMT.053 bank statement parser",
                "supported_extensions": [".xml"],
            },
        }

    def classify_file(self, file_path: str, filename: str) -> dict[str, Any]:
        """
        Classify a file to suggest the best parser.

        Uses heuristics based on file type and content analysis.

        Args:
            file_path: Path to the file
            filename: Original filename

        Returns:
            Dict with classification results
        """
        # Get file extension
        ext = Path(filename).suffix.lower()

        # Dispatch by extension
        if ext in [".xlsx", ".xls"]:
            return self._classify_excel(file_path, filename)
        elif ext == ".csv":
            return self._classify_csv(file_path, filename)
        elif ext == ".xml":
            return self._classify_xml(file_path, filename)
        else:
            return {
                "suggested_parser": None,
                "confidence": 0.0,
                "reason": f"Unsupported file type: {ext}",
                "available_parsers": list(self.parsers_info.keys()),
            }

    async def classify_file_with_ai(self, file_path: str, filename: str) -> dict[str, Any]:
        """
        Classify file with AI enhancement (Fase D).

        First uses heuristics, then optionally improves with IA if confidence is low.

        Args:
            file_path: Path to the file
            filename: Original filename

        Returns:
            Dict with classification results (potentially enhanced by AI)
        """
        from app.config.settings import settings

        # Step 1: Get base classification (heuristics)
        base_result = self.classify_file(file_path, filename)

        # Step 2: If confidence is low, try to improve with IA
        if base_result.get("confidence", 0) < settings.IMPORT_AI_CONFIDENCE_THRESHOLD:
            try:
                # Extract text from file
                text = self._extract_text(file_path, filename)
                if not text or len(text.strip()) < 10:
                    self.logger.debug("Not enough text to classify with AI, using base result")
                    return base_result

                # Get AI provider (singleton)
                try:
                    from app.modules.imports.ai import get_ai_provider_singleton

                    ai_provider = await get_ai_provider_singleton()
                except Exception as e:
                    self.logger.warning(f"Could not load AI provider: {e}, using base result")
                    return base_result

                # Classify with IA
                ai_result = await ai_provider.classify_document(
                    text=text,
                    available_parsers=list(self.parsers_info.keys()),
                    doc_metadata={"filename": filename},
                )

                # Compare: use AI result if more confident
                if ai_result.confidence > base_result.get("confidence", 0):
                    self.logger.info(
                        f"AI enhanced classification: {ai_result.suggested_parser} "
                        f"(base: {base_result.get('suggested_parser')}, "
                        f"confidence: {ai_result.confidence:.2f})"
                    )
                    return {
                        "suggested_parser": ai_result.suggested_parser,
                        "confidence": ai_result.confidence,
                        "probabilities": ai_result.probabilities,
                        "reasoning": ai_result.reasoning,
                        "enhanced_by_ai": True,
                        "ai_provider": ai_result.provider,
                        "available_parsers": list(self.parsers_info.keys()),
                    }
                else:
                    self.logger.debug(
                        f"Base classification more confident, keeping it "
                        f"(base: {base_result.get('confidence'):.2f}, "
                        f"ai: {ai_result.confidence:.2f})"
                    )
            except Exception as e:
                # Fallback gracefully if AI fails
                self.logger.warning(f"AI classification failed: {e}, using base result")

        return base_result

    def _extract_text(self, file_path: str, filename: str) -> str:
        """Extract text content from file for AI analysis."""
        ext = Path(filename).suffix.lower()
        text = ""

        try:
            if ext in [".xlsx", ".xls"]:
                # Extract from Excel
                text = self._extract_text_excel(file_path)
            elif ext == ".csv":
                # Read CSV as text
                with open(file_path, encoding="utf-8-sig", errors="ignore") as f:
                    text = f.read()[:2000]  # First 2000 chars
            elif ext == ".xml":
                # Read XML as text
                with open(file_path, encoding="utf-8", errors="ignore") as f:
                    text = f.read()[:2000]  # First 2000 chars
            else:
                # Try generic text read
                with open(file_path, encoding="utf-8", errors="ignore") as f:
                    text = f.read()[:2000]
        except Exception as e:
            self.logger.debug(f"Error extracting text: {e}")

        return text.strip()

    def _extract_text_excel(self, file_path: str) -> str:
        """Extract text from Excel file."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            text_parts = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 10:  # First 10 rows
                    break
                text_parts.extend([str(cell) for cell in row if cell])

            return " ".join(text_parts)[:1000]  # First 1000 chars
        except Exception:
            return ""

    def _classify_excel(self, file_path: str, filename: str) -> dict[str, Any]:
        """Classify Excel files."""
        try:
            content_analysis = self._analyze_excel_content(file_path)

            # Simple rule-based classification
            if self._looks_like_products(content_analysis):
                suggested_parser = "products_excel"
                confidence = 0.8
                reason = "Detected product-related columns (name, price, quantity)"
            else:
                suggested_parser = "generic_excel"
                confidence = 0.6
                reason = "Generic Excel structure detected"

            return {
                "suggested_parser": suggested_parser,
                "confidence": confidence,
                "reason": reason,
                "available_parsers": list(self.parsers_info.keys()),
                "content_analysis": content_analysis,
            }

        except Exception as e:
            return {
                "suggested_parser": "generic_excel",  # Fallback
                "confidence": 0.3,
                "reason": f"Error analyzing file: {str(e)}",
                "available_parsers": list(self.parsers_info.keys()),
            }

    def _classify_csv(self, file_path: str, filename: str) -> dict[str, Any]:
        """Classify CSV files."""
        try:
            content_analysis = self._analyze_csv_content(file_path)

            # Check if it looks like invoices or bank transactions
            if self._looks_like_bank_csv(content_analysis):
                suggested_parser = "csv_bank"
                confidence = 0.85
                reason = "Detected bank transaction columns (amount, direction, date)"
            elif self._looks_like_invoices_csv(content_analysis):
                suggested_parser = "csv_invoices"
                confidence = 0.8
                reason = "Detected invoice columns (invoice_number, total, vendor)"
            else:
                suggested_parser = "csv_invoices"  # Default
                confidence = 0.5
                reason = "CSV file; assuming invoices (can override)"

            return {
                "suggested_parser": suggested_parser,
                "confidence": confidence,
                "reason": reason,
                "available_parsers": [
                    p
                    for p in self.parsers_info.keys()
                    if ".csv" in str(self.parsers_info[p].get("supported_extensions", []))
                ],
                "content_analysis": content_analysis,
            }
        except Exception as e:
            return {
                "suggested_parser": "csv_invoices",
                "confidence": 0.4,
                "reason": f"Error analyzing CSV: {str(e)}",
                "available_parsers": list(self.parsers_info.keys()),
            }

    def _classify_xml(self, file_path: str, filename: str) -> dict[str, Any]:
        """Classify XML files."""
        try:
            content_analysis = self._analyze_xml_content(file_path)

            # Check if it's a CAMT.053 bank statement
            if content_analysis.get("is_camt053"):
                suggested_parser = "xml_camt053_bank"
                confidence = 0.95
                reason = "Detected ISO 20022 CAMT.053 bank statement"
            # Check if it's an invoice (UBL/CFDI)
            elif content_analysis.get("is_invoice"):
                suggested_parser = "xml_invoice"
                confidence = 0.9
                reason = "Detected UBL/CFDI invoice structure"
            else:
                suggested_parser = "xml_invoice"
                confidence = 0.5
                reason = "XML file; assuming invoice (can override)"

            return {
                "suggested_parser": suggested_parser,
                "confidence": confidence,
                "reason": reason,
                "available_parsers": [
                    p
                    for p in self.parsers_info.keys()
                    if ".xml" in str(self.parsers_info[p].get("supported_extensions", []))
                ],
                "content_analysis": content_analysis,
            }
        except Exception as e:
            return {
                "suggested_parser": "xml_invoice",
                "confidence": 0.3,
                "reason": f"Error analyzing XML: {str(e)}",
                "available_parsers": list(self.parsers_info.keys()),
            }

    def _analyze_excel_content(self, file_path: str) -> dict[str, Any]:
        """Analyze Excel file content for classification hints."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first few rows
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if any(cell for cell in row if cell):
                headers.extend([str(cell).lower() for cell in row if cell])
                break

        # Count keywords for different document types
        product_keywords = [
            "producto",
            "nombre",
            "name",
            "precio",
            "price",
            "cantidad",
            "quantity",
            "stock",
            "categoria",
            "category",
            "sku",
            "codigo",
        ]
        bank_keywords = [
            "fecha",
            "date",
            "importe",
            "amount",
            "saldo",
            "balance",
            "banco",
            "bank",
            "iban",
            "cuenta",
            "account",
        ]
        invoice_keywords = [
            "factura",
            "invoice",
            "vendor",
            "supplier",
            "cliente",
            "customer",
            "iva",
            "tax",
            "total",
            "subtotal",
        ]

        product_score = sum(1 for h in headers for kw in product_keywords if kw in h)
        bank_score = sum(1 for h in headers for kw in bank_keywords if kw in h)
        invoice_score = sum(1 for h in headers for kw in invoice_keywords if kw in h)

        return {
            "headers": headers[:20],  # First 20 headers
            "total_rows": sum(1 for _ in ws.iter_rows()),
            "scores": {
                "products": product_score,
                "bank": bank_score,
                "invoices": invoice_score,
            },
        }

    def _looks_like_products(self, analysis: dict[str, Any]) -> bool:
        """Check if content looks like product data."""
        scores = analysis.get("scores", {})
        return scores.get("products", 0) > scores.get("bank", 0) and scores.get(
            "products", 0
        ) > scores.get("invoices", 0)

    def _analyze_csv_content(self, file_path: str) -> dict[str, Any]:
        """Analyze CSV file content for classification."""
        headers = []
        rows_count = 0

        try:
            with open(file_path, encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                headers = [h.lower().strip() for h in (reader.fieldnames or [])]
                rows_count = sum(1 for _ in reader)
        except Exception:
            pass

        return {
            "headers": headers,
            "rows_count": rows_count,
            "header_keywords": self._extract_keywords_from_headers(headers),
        }

    def _looks_like_bank_csv(self, analysis: dict[str, Any]) -> bool:
        """Check if CSV looks like bank transactions."""
        keywords = analysis.get("header_keywords", {})
        bank_score = keywords.get("bank", 0)
        invoice_score = keywords.get("invoices", 0)
        return bank_score > 2 and bank_score > invoice_score

    def _looks_like_invoices_csv(self, analysis: dict[str, Any]) -> bool:
        """Check if CSV looks like invoices."""
        keywords = analysis.get("header_keywords", {})
        invoice_score = keywords.get("invoices", 0)
        bank_score = keywords.get("bank", 0)
        return invoice_score > 1 and invoice_score > bank_score

    def _analyze_xml_content(self, file_path: str) -> dict[str, Any]:
        """Analyze XML file content for classification."""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()

            # Get root tag and namespace
            tag = root.tag.lower()

            # Check for CAMT.053
            is_camt053 = "camt" in tag or "053" in tag or "statement" in tag

            # Check for UBL/CFDI invoices
            is_invoice = "invoice" in tag or "cfdi" in tag or "ubl" in tag

            return {
                "root_tag": tag,
                "is_camt053": is_camt053,
                "is_invoice": is_invoice,
                "element_count": len(list(root.iter())),
            }
        except Exception:
            return {
                "root_tag": "",
                "is_camt053": False,
                "is_invoice": False,
                "element_count": 0,
            }

    def _extract_keywords_from_headers(self, headers: list) -> dict[str, int]:
        """Extract keywords from CSV headers."""
        bank_keywords = [
            "amount",
            "importe",
            "monto",
            "date",
            "fecha",
            "direction",
            "debit",
            "credit",
            "iban",
            "account",
            "balance",
        ]
        invoice_keywords = [
            "invoice",
            "factura",
            "number",
            "vendor",
            "supplier",
            "total",
            "subtotal",
            "iva",
            "tax",
            "customer",
        ]

        bank_score = sum(1 for h in headers for kw in bank_keywords if kw in h)
        invoice_score = sum(1 for h in headers for kw in invoice_keywords if kw in h)

        return {
            "bank": bank_score,
            "invoices": invoice_score,
        }


# Global classifier instance
classifier = FileClassifier()
