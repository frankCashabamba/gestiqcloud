"""Smart Router: Unifica dispatcher + classifier + AI para análisis inteligente de archivos."""

from __future__ import annotations

import csv
import logging
import mimetypes
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from app.services.excel_analyzer import detect_header_row, extract_headers

from app.config.settings import settings
from app.modules.imports.ai.mapping_suggester import mapping_suggester
from app.modules.imports.parsers import registry
from app.modules.imports.parsers.dispatcher import select_parser_for_file
from app.modules.imports.services.classifier import FileClassifier
from app.modules.imports.services.ocr_service import DocumentLayout, ocr_service

logger = logging.getLogger("app.imports.smart_router")


@dataclass
class AnalysisResult:
    """Resultado del análisis inteligente de un archivo."""

    suggested_parser: str
    suggested_doc_type: str
    confidence: float
    headers_sample: list[str]
    mapping_suggestion: dict[str, str] | None
    mapping_transforms: dict[str, str] | None
    mapping_defaults: dict[str, Any] | None
    mapping_confidence: float | None
    mapping_provider: str | None
    explanation: str
    decision_log: list[dict[str, Any]]
    requires_confirmation: bool
    available_parsers: list[str]
    probabilities: dict[str, float] | None
    ai_enhanced: bool
    ai_provider: str | None
    ocr_text: str | None = None
    ocr_layout: str | None = None
    ocr_confidence: float | None = None


STANDARD_FIELD_MAPPINGS = {
    "products": {
        "nombre": "name",
        "name": "name",
        "producto": "name",
        "product": "name",
        "descripcion": "description",
        "description": "description",
        "precio": "price",
        "price": "price",
        "costo": "cost",
        "cost": "cost",
        "cantidad": "quantity",
        "quantity": "quantity",
        "stock": "stock",
        "sku": "sku",
        "codigo": "sku",
        "code": "sku",
        "categoria": "category",
        "category": "category",
        "unidad": "unit",
        "unit": "unit",
    },
    "bank_transactions": {
        "fecha": "date",
        "date": "date",
        "importe": "amount",
        "amount": "amount",
        "monto": "amount",
        "concepto": "description",
        "description": "description",
        "descripcion": "description",
        "saldo": "balance",
        "balance": "balance",
        "cuenta": "account",
        "account": "account",
        "iban": "iban",
        "referencia": "reference",
        "reference": "reference",
    },
    "invoices": {
        "numero": "invoice_number",
        "number": "invoice_number",
        "factura": "invoice_number",
        "invoice": "invoice_number",
        "fecha": "date",
        "date": "date",
        "proveedor": "vendor",
        "vendor": "vendor",
        "supplier": "vendor",
        "cliente": "customer",
        "customer": "customer",
        "total": "total",
        "subtotal": "subtotal",
        "iva": "tax",
        "tax": "tax",
        "impuesto": "tax",
        "ruc": "tax_id",
        "nit": "tax_id",
        "cif": "tax_id",
    },
    "expenses": {
        "fecha": "date",
        "date": "date",
        "concepto": "description",
        "description": "description",
        "gasto": "description",
        "expense": "description",
        "monto": "amount",
        "amount": "amount",
        "importe": "amount",
        "categoria": "category",
        "category": "category",
        "recibo": "receipt_number",
        "receipt": "receipt_number",
    },
}


class SmartRouter:
    """Router inteligente que unifica dispatcher, classifier y AI."""

    def __init__(self):
        self.classifier = FileClassifier()
        self.confidence_threshold = getattr(settings, "IMPORT_CONFIRMATION_THRESHOLD", 0.7)

    async def analyze_file(
        self,
        file_path: str,
        filename: str,
        content_type: str | None = None,
        tenant_id: str | None = None,
    ) -> AnalysisResult:
        """
        Analiza un archivo y retorna sugerencias de parser, doc_type y mapeo.

        Pasos:
        1. Detectar extensión/mime
        2. Extraer headers/muestra de filas
        3. Usar heurísticas del dispatcher
        4. Si confianza < threshold, intentar mejorar con IA
        5. Sugerir mapping inicial basado en headers
        6. Registrar cada paso en decision_log
        """
        decision_log: list[dict[str, Any]] = []
        ext = Path(filename).suffix.lower()

        decision_log.append(
            {
                "step": "file_detection",
                "extension": ext,
                "content_type": content_type,
                "filename": filename,
            }
        )

        ocr_text: str | None = None
        ocr_layout: str | None = None
        ocr_confidence: float | None = None

        if ext in (".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".bmp") and ocr_service.is_available():
            try:
                ocr_result = await ocr_service.extract_text(file_path)
                ocr_text = ocr_result.text[:10000] if ocr_result.text else None
                ocr_layout = ocr_result.layout.value
                ocr_confidence = ocr_result.confidence

                decision_log.append(
                    {
                        "step": "ocr_extraction",
                        "pages": ocr_result.pages,
                        "layout": ocr_layout,
                        "confidence": ocr_confidence,
                        "text_length": len(ocr_result.text) if ocr_result.text else 0,
                        "tables_found": len(ocr_result.tables),
                    }
                )
            except Exception as e:
                logger.warning(f"OCR extraction failed: {e}")
                decision_log.append(
                    {
                        "step": "ocr_extraction_failed",
                        "error": str(e),
                    }
                )

        headers_sample = self._extract_headers(file_path, ext)
        decision_log.append(
            {
                "step": "headers_extraction",
                "headers_count": len(headers_sample),
                "headers": headers_sample[:10],
            }
        )

        dispatcher_parser, dispatcher_doc_type = select_parser_for_file(
            file_path,
            content_type=content_type,
            original_filename=filename,
        )
        decision_log.append(
            {
                "step": "dispatcher_heuristics",
                "suggested_parser": dispatcher_parser,
                "suggested_doc_type": dispatcher_doc_type,
            }
        )

        classifier_result = self.classifier.classify_file(file_path, filename)
        classifier_confidence = classifier_result.get("confidence", 0.0)
        classifier_parser = classifier_result.get("suggested_parser")

        decision_log.append(
            {
                "step": "classifier_analysis",
                "suggested_parser": classifier_parser,
                "confidence": classifier_confidence,
                "reason": classifier_result.get("reason", ""),
            }
        )

        # If XLSX and recipe keywords are present, force recipes early
        if Path(filename).suffix.lower() in (".xlsx", ".xls"):
            headers_blob = " ".join(h.lower() for h in headers_sample)
            if any(
                kw in headers_blob
                for kw in (
                    "ingredientes",
                    "costo total ingredientes",
                    "formato de costeo",
                    "temperatura de servicio",
                )
            ):
                dispatcher_parser = "xlsx_recipes"
                dispatcher_doc_type = "recipes"
                decision_log.append(
                    {
                        "step": "force_recipe_by_keyword",
                        "headers_sample": headers_sample[:10],
                        "suggested_parser": dispatcher_parser,
                        "suggested_doc_type": dispatcher_doc_type,
                    }
                )

        if classifier_confidence > 0.7 and classifier_parser:
            base_parser = classifier_parser
            base_confidence = classifier_confidence
            parser_info = registry.get_parser(base_parser)
            base_doc_type = parser_info["doc_type"] if parser_info else dispatcher_doc_type
        else:
            base_parser = dispatcher_parser
            base_confidence = classifier_confidence if classifier_confidence > 0 else 0.5
            base_doc_type = dispatcher_doc_type

        decision_log.append(
            {
                "step": "base_decision",
                "chosen_parser": base_parser,
                "chosen_doc_type": base_doc_type,
                "base_confidence": base_confidence,
            }
        )

        ai_enhanced = False
        ai_provider_name: str | None = None
        probabilities: dict[str, float] | None = None
        final_parser = base_parser
        final_doc_type = base_doc_type
        final_confidence = base_confidence
        explanation = classifier_result.get("reason", "Análisis basado en heurísticas")

        if base_confidence < self.confidence_threshold:
            decision_log.append(
                {
                    "step": "ai_enhancement_attempt",
                    "reason": f"confidence {base_confidence:.2f} < threshold {self.confidence_threshold}",
                }
            )

            try:
                ai_result = await self.classifier.classify_file_with_ai(file_path, filename)

                if ai_result.get("enhanced_by_ai"):
                    ai_enhanced = True
                    ai_provider_name = ai_result.get("ai_provider")
                    ai_confidence = ai_result.get("confidence", 0.0)
                    ai_parser = ai_result.get("suggested_parser")
                    probabilities = ai_result.get("probabilities")

                    if ai_confidence > base_confidence and ai_parser:
                        final_parser = ai_parser
                        final_confidence = ai_confidence
                        parser_info = registry.get_parser(final_parser)
                        final_doc_type = parser_info["doc_type"] if parser_info else base_doc_type
                        explanation = ai_result.get("reasoning", "Clasificación mejorada por IA")

                        decision_log.append(
                            {
                                "step": "ai_enhancement_success",
                                "ai_parser": ai_parser,
                                "ai_confidence": ai_confidence,
                                "ai_provider": ai_provider_name,
                            }
                        )
                    else:
                        decision_log.append(
                            {
                                "step": "ai_enhancement_no_improvement",
                                "ai_confidence": ai_confidence,
                                "base_confidence": base_confidence,
                            }
                        )
                else:
                    decision_log.append(
                        {
                            "step": "ai_enhancement_skipped",
                            "reason": "AI did not enhance result",
                        }
                    )
            except Exception as e:
                logger.warning(f"AI enhancement failed: {e}")
                decision_log.append(
                    {
                        "step": "ai_enhancement_error",
                        "error": str(e),
                    }
                )

        sample_rows = self._extract_sample_rows(file_path, ext, headers_sample)
        ai_mapping = await mapping_suggester.suggest_mapping(
            headers=headers_sample,
            sample_rows=sample_rows,
            doc_type=final_doc_type,
            tenant_id=tenant_id,
            use_ai=(settings.IMPORT_AI_PROVIDER != "local"),
        )

        mapping_suggestion = ai_mapping.mappings if ai_mapping.mappings else None
        mapping_transforms = ai_mapping.transforms
        mapping_defaults = ai_mapping.defaults
        mapping_confidence = ai_mapping.confidence
        mapping_provider = ai_mapping.provider

        decision_log.append(
            {
                "step": "mapping_suggestion",
                "mappings_count": len(mapping_suggestion) if mapping_suggestion else 0,
                "mapping_confidence": mapping_confidence,
                "mapping_provider": mapping_provider,
                "from_cache": ai_mapping.from_cache,
            }
        )

        available_parsers = list(registry.list_parsers().keys())
        requires_confirmation = final_confidence < self.confidence_threshold

        decision_log.append(
            {
                "step": "final_decision",
                "parser": final_parser,
                "doc_type": final_doc_type,
                "confidence": final_confidence,
                "requires_confirmation": requires_confirmation,
                "ai_enhanced": ai_enhanced,
            }
        )

        logger.info(
            f"SmartRouter analysis complete: parser={final_parser}, "
            f"doc_type={final_doc_type}, confidence={final_confidence:.2f}, "
            f"ai_enhanced={ai_enhanced}"
        )

        return AnalysisResult(
            suggested_parser=final_parser,
            suggested_doc_type=final_doc_type,
            confidence=final_confidence,
            headers_sample=headers_sample,
            mapping_suggestion=mapping_suggestion,
            mapping_transforms=mapping_transforms,
            mapping_defaults=mapping_defaults,
            mapping_confidence=mapping_confidence,
            mapping_provider=mapping_provider,
            explanation=explanation,
            decision_log=decision_log,
            requires_confirmation=requires_confirmation,
            available_parsers=available_parsers,
            probabilities=probabilities,
            ai_enhanced=ai_enhanced,
            ai_provider=ai_provider_name,
            ocr_text=ocr_text,
            ocr_layout=ocr_layout,
            ocr_confidence=ocr_confidence,
        )

    def _extract_headers(self, file_path: str, ext: str) -> list[str]:
        """Extrae headers del archivo según su tipo."""
        headers: list[str] = []

        try:
            if ext in (".xlsx", ".xls", ".xlsm", ".xlsb"):
                headers = self._extract_excel_headers(file_path)
            elif ext == ".csv":
                headers = self._extract_csv_headers(file_path)
            elif ext == ".xml":
                headers = self._extract_xml_elements(file_path)
        except Exception as e:
            logger.warning(f"Error extracting headers: {e}")

        return headers

    def _extract_sample_rows(self, file_path: str, ext: str, headers: list[str]) -> list[list[Any]]:
        """Extrae filas de muestra del archivo."""
        try:
            if ext in (".xlsx", ".xls", ".xlsm", ".xlsb"):
                return self._extract_excel_sample_rows(file_path, len(headers))
            elif ext == ".csv":
                return self._extract_csv_sample_rows(file_path)
        except Exception as e:
            logger.warning(f"Error extracting sample rows: {e}")
        return []

    def _extract_excel_headers(self, file_path: str) -> list[str]:
        """Extrae headers de archivo Excel."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            header_row = detect_header_row(ws)
            headers = extract_headers(ws, header_row)
            wb.close()
            return headers
        except Exception:
            return []

    def _extract_excel_sample_rows(self, file_path: str, num_cols: int) -> list[list[Any]]:
        """Extrae filas de muestra de archivo Excel."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows: list[list[Any]] = []
            header_found = False

            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                if not header_found:
                    if any(cell for cell in row if cell):
                        header_found = True
                    continue

                row_data = list(row[:num_cols]) if num_cols else list(row)
                if any(cell for cell in row_data if cell):
                    rows.append(row_data)
                    if len(rows) >= 3:
                        break

            wb.close()
            return rows
        except Exception:
            return []

    def _extract_csv_sample_rows(self, file_path: str) -> list[list[Any]]:
        """Extrae filas de muestra de archivo CSV."""
        try:
            rows: list[list[Any]] = []
            with open(file_path, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if row:
                        rows.append(row)
                        if len(rows) >= 3:
                            break
            return rows
        except Exception:
            return []

    def _extract_csv_headers(self, file_path: str) -> list[str]:
        """Extrae headers de archivo CSV."""
        try:
            with open(file_path, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                first_row = next(reader, [])
                return [h.strip() for h in first_row if h.strip()]
        except Exception:
            return []

    def _extract_xml_elements(self, file_path: str) -> list[str]:
        """Extrae nombres de elementos principales de XML."""
        import xml.etree.ElementTree as ET

        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            elements = set()

            for child in root.iter():
                tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                elements.add(tag)
                if len(elements) >= 20:
                    break

            return list(elements)[:20]
        except Exception:
            return []

    def _suggest_mapping(self, headers: list[str], doc_type: str) -> dict[str, str] | None:
        """Sugiere mapeo de headers a campos estándar."""
        if not headers:
            return None

        field_map = STANDARD_FIELD_MAPPINGS.get(doc_type, {})
        if not field_map:
            return None

        mapping: dict[str, str] = {}

        for header in headers:
            header_lower = header.lower().strip()
            header_normalized = header_lower.replace("_", " ").replace("-", " ")

            for keyword, target_field in field_map.items():
                if keyword in header_normalized or header_normalized in keyword:
                    if target_field not in mapping.values():
                        mapping[header] = target_field
                        break

        return mapping if mapping else None


smart_router = SmartRouter()
