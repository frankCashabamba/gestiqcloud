"""
Large file handling for imports >100MB.
Chunked processing with streaming parser.
"""

from dataclasses import dataclass
from enum import Enum
from typing import AsyncIterator, Optional
import hashlib


class ChunkStatus(str, Enum):
    """Status of file chunk."""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    MERGED = "merged"


@dataclass
class FileChunk:
    """Single chunk of uploaded file."""
    upload_id: str
    chunk_number: int
    size_bytes: int
    hash_md5: str  # For integrity check
    status: ChunkStatus = ChunkStatus.PENDING
    error: Optional[str] = None


@dataclass
class ChunkedFileSession:
    """Session for chunked file upload."""
    upload_id: str
    filename: str
    total_size: int
    chunk_size: int
    expected_chunks: int
    chunks_received: dict[int, FileChunk]
    content_type: Optional[str] = None
    created_at: Optional[str] = None
    
    def add_chunk(self, chunk: FileChunk):
        """Add received chunk to session."""
        self.chunks_received[chunk.chunk_number] = chunk
    
    def is_complete(self) -> bool:
        """Check if all chunks received."""
        return len(self.chunks_received) == self.expected_chunks
    
    def get_missing_chunks(self) -> list[int]:
        """Get list of missing chunk numbers."""
        all_chunks = set(range(self.expected_chunks))
        received = set(self.chunks_received.keys())
        return sorted(list(all_chunks - received))
    
    def get_progress_percent(self) -> float:
        """Get upload progress as percentage."""
        if self.expected_chunks == 0:
            return 0.0
        return (len(self.chunks_received) / self.expected_chunks) * 100


class StreamingExcelParser:
    """
    Streaming parser for large Excel files.
    Processes rows in batches without loading entire file into memory.
    """
    
    def __init__(self, batch_size: int = 1000):
        self.batch_size = batch_size
    
    async def parse_file_streaming(
        self,
        file_path: str,
        header_row: int = 1,
    ) -> AsyncIterator[dict]:
        """
        Parse Excel file in streaming fashion.
        Yields batches of rows without loading entire file.
        
        Yields:
            Dict with batch metadata and rows
        """
        try:
            import openpyxl
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl required for streaming parser")
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Extract headers from specified row
        headers = []
        for cell in ws[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                break
        
        # Stream rows in batches
        batch = []
        row_number = header_row + 1
        
        for openpyxl_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Skip empty rows
            if not any(cell for cell in openpyxl_row):
                continue
            
            # Convert to dict
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(openpyxl_row):
                    row_dict[header] = openpyxl_row[col_idx]
            
            batch.append(row_dict)
            
            # Yield batch when it reaches batch_size
            if len(batch) >= self.batch_size:
                yield {
                    "batch_number": len(batch) // self.batch_size,
                    "batch_size": len(batch),
                    "headers": headers,
                    "rows": batch,
                }
                batch = []
            
            row_number += 1
        
        # Yield remaining rows
        if batch:
            yield {
                "batch_number": (row_number - header_row) // self.batch_size,
                "batch_size": len(batch),
                "headers": headers,
                "rows": batch,
            }
        
        wb.close()
    
    async def parse_csv_streaming(
        self,
        file_path: str,
    ) -> AsyncIterator[dict]:
        """Stream CSV in batches."""
        import csv
        
        batch = []
        headers = None
        batch_num = 0
        
        try:
            with open(file_path, encoding='utf-8-sig', newline='') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                
                for row in reader:
                    batch.append(row)
                    
                    if len(batch) >= self.batch_size:
                        yield {
                            "batch_number": batch_num,
                            "batch_size": len(batch),
                            "headers": headers,
                            "rows": batch,
                        }
                        batch = []
                        batch_num += 1
                
                if batch:
                    yield {
                        "batch_number": batch_num,
                        "batch_size": len(batch),
                        "headers": headers,
                        "rows": batch,
                    }
        except Exception as e:
            raise IOError(f"Error streaming CSV: {e}")


class LargeFileOptimizer:
    """
    Optimization strategies for large files.
    Decides between streaming, chunking, or async processing.
    """
    
    # Thresholds
    SIZE_THRESHOLD_MB = {
        "stream": 50,  # >50MB use streaming
        "async_worker": 100,  # >100MB use async worker
        "chunk_upload": 10,  # >10MB suggest chunked upload
    }
    
    def __init__(self):
        self.streaming_parser = StreamingExcelParser(batch_size=1000)
    
    def get_optimal_strategy(self, file_size_mb: float) -> dict[str, str]:
        """
        Recommend processing strategy based on file size.
        
        Returns:
            {
                "strategy": "inline" | "streaming" | "chunked_upload" | "async_worker",
                "batch_size": int,
                "reason": str,
            }
        """
        if file_size_mb < self.SIZE_THRESHOLD_MB["chunk_upload"]:
            return {
                "strategy": "inline",
                "batch_size": 10000,
                "reason": "File small enough for inline processing",
            }
        elif file_size_mb < self.SIZE_THRESHOLD_MB["stream"]:
            return {
                "strategy": "chunked_upload",
                "batch_size": 5000,
                "reason": "Recommend chunked upload for reliability",
            }
        elif file_size_mb < self.SIZE_THRESHOLD_MB["async_worker"]:
            return {
                "strategy": "streaming",
                "batch_size": 1000,
                "reason": "Use streaming parser to avoid memory issues",
            }
        else:
            return {
                "strategy": "async_worker",
                "batch_size": 500,
                "reason": "Use Celery async worker for very large files",
            }
    
    def estimate_processing_time(self, file_size_mb: float, rows: int) -> dict[str, float]:
        """
        Estimate processing time based on size and row count.
        
        Returns:
            {
                "parse_seconds": float,
                "validate_seconds": float,
                "total_seconds": float,
            }
        """
        # Empirical: 100K rows per second for parsing
        parse_rate = 100000
        parse_seconds = max(rows / parse_rate, 1)
        
        # Validation adds ~30% overhead
        validate_seconds = parse_seconds * 0.3
        
        return {
            "parse_seconds": parse_seconds,
            "validate_seconds": validate_seconds,
            "total_seconds": parse_seconds + validate_seconds,
        }
    
    def estimate_memory_usage(self, file_size_mb: float, batch_size: int = 1000) -> dict[str, float]:
        """
        Estimate memory usage for processing.
        
        Returns:
            {
                "per_batch_mb": float,
                "max_memory_mb": float,
            }
        """
        # Rough estimate: 1KB per row
        bytes_per_row = 1024
        batch_memory_mb = (batch_size * bytes_per_row) / (1024 * 1024)
        
        # Parser overhead ~50MB
        parser_overhead_mb = 50
        
        max_memory = batch_memory_mb + parser_overhead_mb
        
        return {
            "per_batch_mb": batch_memory_mb,
            "max_memory_mb": max_memory,
        }


# Global instances
streaming_parser = StreamingExcelParser()
large_file_optimizer = LargeFileOptimizer()
