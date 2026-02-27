from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell import MergedCell

from .header_norm import normalize_headers
from .schema import ExtractRule, HeaderNormalization, SheetRule


def _get_merged_value(ws, cell) -> Any:
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def _fill_merged_cells(ws, rows: list[list[Any]]) -> list[list[Any]]:
    filled = [list(row) for row in rows]
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                row_idx = r - 1
                col_idx = c - 1
                if row_idx < len(filled) and col_idx < len(filled[row_idx]):
                    filled[row_idx][col_idx] = value
    return filled


def _read_sheet_rows(ws, fill_merged: bool = False) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows():
        rows.append([_get_merged_value(ws, cell) for cell in row])

    if fill_merged:
        rows = _fill_merged_cells(ws, rows)

    return rows


def _detect_header_row(
    rows: list[list[Any]],
    map_keys: set[str],
) -> int:
    best_idx = 0
    best_hits = 0
    for idx, row in enumerate(rows[:30]):
        hits = 0
        for cell in row:
            if cell is None:
                continue
            cell_lower = str(cell).strip().lower()
            if cell_lower in map_keys:
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_idx = idx
    return best_idx


def _match_sheet_rule(sheet_name: str, rules: list[SheetRule]) -> SheetRule | None:
    for rule in rules:
        try:
            if re.search(rule.sheet_name_regex, sheet_name, re.IGNORECASE):
                return rule
        except re.error:
            continue
    return None


def read_excel_tables(
    file_path: str | Path,
    extract: ExtractRule,
    header_norm: HeaderNormalization,
    map_keys: dict[str, list[str]],
    language: str = "es",
) -> list[dict[str, Any]]:
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == ".xls":
        try:
            import xlrd  # type: ignore

            return _read_xls(file_path, extract, header_norm, map_keys, language)
        except ImportError:
            raise ImportError("xlrd is required to read .xls files")

    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=False)
    all_map_keys: set[str] = set(map_keys.keys())
    for aliases in map_keys.values():
        all_map_keys.update(a.lower() for a in aliases)

    tables: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        if not extract.all_sheets and ws.title != wb.sheetnames[0]:
            matched_rule = _match_sheet_rule(ws.title, extract.sheet_rules)
            if not matched_rule:
                continue
        else:
            matched_rule = _match_sheet_rule(ws.title, extract.sheet_rules)

        fill_merged = bool(matched_rule and matched_rule.merged_cells.fill)
        rows = _read_sheet_rows(ws, fill_merged=fill_merged)
        if not rows:
            continue

        header_idx = _detect_header_row(rows, all_map_keys)
        headers_raw = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
        headers_norm = normalize_headers(headers_raw, header_norm, language)

        data_rows: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            row_dict: dict[str, Any] = {}
            for col_idx, header in enumerate(headers_norm):
                if col_idx < len(row):
                    row_dict[header] = row[col_idx]
            data_rows.append(row_dict)

        tables.append({
            "sheet": ws.title,
            "headers_raw": headers_raw,
            "headers_norm": headers_norm,
            "rows": data_rows,
        })

    wb.close()
    return tables


def _read_xls(
    file_path: Path,
    extract: ExtractRule,
    header_norm: HeaderNormalization,
    map_keys: dict[str, list[str]],
    language: str,
) -> list[dict[str, Any]]:
    import xlrd  # type: ignore

    wb = xlrd.open_workbook(str(file_path))
    all_map_keys: set[str] = set(map_keys.keys())
    for aliases in map_keys.values():
        all_map_keys.update(a.lower() for a in aliases)

    tables: list[dict[str, Any]] = []
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        if not extract.all_sheets and sheet_idx != 0:
            matched_rule = _match_sheet_rule(ws.name, extract.sheet_rules)
            if not matched_rule:
                continue

        rows: list[list[Any]] = []
        for r in range(ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])

        if not rows:
            continue

        header_idx = _detect_header_row(rows, all_map_keys)
        headers_raw = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
        headers_norm = normalize_headers(headers_raw, header_norm, language)

        data_rows: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            row_dict: dict[str, Any] = {}
            for col_idx, header in enumerate(headers_norm):
                if col_idx < len(row):
                    row_dict[header] = row[col_idx]
            data_rows.append(row_dict)

        tables.append({
            "sheet": ws.name,
            "headers_raw": headers_raw,
            "headers_norm": headers_norm,
            "rows": data_rows,
        })

    return tables
