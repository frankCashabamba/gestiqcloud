"""Excel parser for products with automatic category detection."""

import re
from datetime import datetime
from typing import Any

import openpyxl


def _to_number(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s_norm = re.sub(r"[^0-9,.-]", "", s)
    if "," in s_norm and "." in s_norm:
        if s_norm.rfind(",") > s_norm.rfind("."):
            s_norm = s_norm.replace(".", "").replace(",", ".")
        else:
            s_norm = s_norm.replace(",", "")
    else:
        s_norm = s_norm.replace(",", ".")
    try:
        return float(s_norm)
    except (ValueError, TypeError):
        return None


def parse_products_excel(file_path: str, sheet_name: str = None) -> dict[str, Any]:
    """Parse products Excel with automatic category detection.

    Detects categories as rows with:
    - Product name filled
    - Quantity (cantidad) empty/None
    - Price (precio) empty/None

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name (default: first sheet or "REGISTRO")

    Returns:
        Dict with:
            - products: List of product dicts with categoria field
            - categories: List of detected category names
            - stats: Parsing statistics
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Try to find "REGISTRO" sheet, otherwise use first sheet
    if sheet_name:
        ws = wb[sheet_name]
    elif "REGISTRO" in wb.sheetnames:
        ws = wb["REGISTRO"]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {
            "products": [],
            "categories": [],
            "stats": {"total": 0, "categories": 0},
        }

    # Assume first row is header
    header = rows[0]

    # Find column indices (flexible mapping)
    col_map = {}
    for idx, col_name in enumerate(header):
        if not col_name:
            continue
        col_lower = str(col_name).lower().strip()

        # Producto/Nombre - más flexible
        if any(kw in col_lower for kw in ["producto", "nombre", "name", "item", "articulo"]):
            if "nombre" not in col_map:  # Evitar sobreescribir
                col_map["nombre"] = idx

        # Cantidad - detectar múltiples variantes (excluir "sobrante" y "venta diaria")
        elif any(kw in col_lower for kw in ["cantidad", "qty", "stock", "existencia", "unidades"]):
            # No mapear si es dato operacional
            if (
                "sobrante" not in col_lower
                and "venta" not in col_lower
                and "diaria" not in col_lower
            ):
                if "cantidad" not in col_map:
                    col_map["cantidad"] = idx

        # Precio - más flexible (no requiere "unit")
        elif any(
            kw in col_lower for kw in ["precio", "price", "pvp", "venta", "unitario", "valor"]
        ):
            if "precio" not in col_map:
                col_map["precio"] = idx

        # Costo
        elif any(
            kw in col_lower
            for kw in [
                "costo",
                "cost",
                "coste",
                "costo promedio",
                "costo unitario",
                "precio costo",
                "cost_price",
                "unit_cost",
            ]
        ):
            if "costo" not in col_map:
                col_map["costo"] = idx

        # SKU/Código
        elif any(
            kw in col_lower for kw in ["sku", "codigo", "código", "code", "referencia", "ref"]
        ):
            if "sku" not in col_map:
                col_map["sku"] = idx

    # Fallback to index-based mapping if columns not found
    if "nombre" not in col_map and len(header) >= 1:
        col_map["nombre"] = 0
    if "cantidad" not in col_map and len(header) >= 2:
        col_map["cantidad"] = 1
    if "precio" not in col_map and len(header) >= 3:
        col_map["precio"] = 2

    products = []
    categories = []
    current_category = "SIN_CATEGORIA"

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or len(row) == 0:
            continue

        nombre_idx = col_map.get("nombre", 0)
        cantidad_idx = col_map.get("cantidad", 1)
        precio_idx = col_map.get("precio", 2)
        costo_idx = col_map.get("costo")

        nombre = row[nombre_idx] if len(row) > nombre_idx else None
        cantidad = row[cantidad_idx] if len(row) > cantidad_idx else None
        precio = row[precio_idx] if len(row) > precio_idx else None
        costo = row[costo_idx] if costo_idx is not None and len(row) > costo_idx else None

        if not nombre:
            continue

        nombre_str = str(nombre).strip()

        # Detect category: has name but no quantity and no price
        # Also check for empty strings, not just None
        cantidad_empty = cantidad is None or (isinstance(cantidad, str) and not cantidad.strip())
        precio_empty = precio is None or (isinstance(precio, str) and not precio.strip())

        if cantidad_empty and precio_empty:
            current_category = nombre_str.upper()  # Categorías en mayúsculas
            if current_category not in categories:
                categories.append(current_category)
            continue

        # Parse product
        precio_val = _to_number(precio)
        if precio_val is None:
            precio_val = 0.0

        cantidad_val = _to_number(cantidad)
        if cantidad_val is None:
            cantidad_val = 0.0

        costo_val = _to_number(costo)

        # Extract SKU if available
        sku = None
        if "sku" in col_map:
            sku_idx = col_map["sku"]
            if len(row) > sku_idx:
                sku_val = row[sku_idx]
                if sku_val:
                    sku = str(sku_val).strip()

        product = {
            "nombre": nombre_str,
            "name": nombre_str,  # Alias for handler
            "producto": nombre_str,  # Alias for validator
            "precio": precio_val,
            "price": precio_val,
            "cantidad": cantidad_val,
            "quantity": cantidad_val,
            "stock": cantidad_val,
            "categoria": current_category,
            "category": current_category,
            "_row": row_idx,
            "_imported_at": datetime.utcnow().isoformat(),
        }
        if costo_val is not None:
            product.update(
                {
                    "cost_price": costo_val,
                    "cost": costo_val,
                    "costo": costo_val,
                    "unit_cost": costo_val,
                }
            )

        if sku:
            product["sku"] = sku

        products.append(product)

    stats = {
        "total": len(products),
        "categories": len(categories),
        "with_stock": sum(1 for p in products if p["cantidad"] > 0),
        "zero_stock": sum(1 for p in products if p["cantidad"] == 0),
    }

    return {
        "products": products,
        "categories": categories,
        "stats": stats,
    }


def normalize_product_row(
    raw_row: dict[str, Any], category_context: str | None = None
) -> dict[str, Any]:
    """Normalize a single product row for validation/promotion.

    Args:
        raw_row: Raw data dict
        category_context: Current category name from parser context

    Returns:
        Normalized dict with standardized field names
    """
    # Ensure all aliases are present
    nombre = (
        raw_row.get("nombre")
        or raw_row.get("name")
        or raw_row.get("producto")
        or raw_row.get("articulo")
        or ""
    )
    precio = _to_number(raw_row.get("precio") or raw_row.get("price"))
    if precio is None:
        precio = 0.0
    cantidad = _to_number(
        raw_row.get("cantidad") or raw_row.get("quantity") or raw_row.get("stock")
    )
    if cantidad is None:
        cantidad = 0.0
    costo = _to_number(
        raw_row.get("cost_price")
        or raw_row.get("cost")
        or raw_row.get("costo")
        or raw_row.get("costo_promedio")
        or raw_row.get("costo promedio")
        or raw_row.get("costo_unitario")
        or raw_row.get("costo unitario")
        or raw_row.get("unit_cost")
    )
    categoria = (
        raw_row.get("categoria") or raw_row.get("category") or category_context or "SIN_CATEGORIA"
    )

    normalized = {
        "name": str(nombre).strip(),
        "producto": str(nombre).strip(),
        "nombre": str(nombre).strip(),
        "price": float(precio) if precio is not None else 0.0,
        "precio": float(precio) if precio is not None else 0.0,
        "quantity": float(cantidad) if cantidad is not None else 0.0,
        "cantidad": float(cantidad) if cantidad is not None else 0.0,
        "stock": float(cantidad) if cantidad is not None else 0.0,
        "category": str(categoria).strip(),
        "categoria": str(categoria).strip(),
    }
    if costo is not None:
        normalized.update(
            {
                "cost_price": float(costo),
                "cost": float(costo),
                "costo": float(costo),
                "unit_cost": float(costo),
            }
        )

    # Preserve SKU if exists
    sku_val = (
        raw_row.get("sku") or raw_row.get("codigo") or raw_row.get("c?digo") or raw_row.get("code")
    )
    if sku_val not in (None, ""):
        normalized["sku"] = sku_val

    # Preserve metadata
    if "_row" in raw_row:
        normalized["_row"] = raw_row["_row"]
    if "_imported_at" in raw_row:
        normalized["_imported_at"] = raw_row["_imported_at"]

    return normalized
