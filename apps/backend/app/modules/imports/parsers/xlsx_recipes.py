"""Specialized parser for PAN KUSI recipe costing Excel."""

from datetime import datetime
from decimal import Decimal
from typing import Any

import openpyxl


def parse_xlsx_recipes(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """Parse PAN KUSI costing sheet into recipes + ingredients."""
    rows: list[dict[str, Any]] = []
    recipes: list[dict[str, Any]] = []
    materials: list[dict[str, Any]] = []
    errors: list[str] = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = [sheet_name] if sheet_name else wb.sheetnames
        for sname in sheets:
            if sname not in wb.sheetnames:
                errors.append(f"Sheet '{sname}' not found")
                continue
            ws = wb[sname]
            header = _parse_header(ws)
            if not header.get("name"):
                continue
            ingredients = _parse_ingredients(ws)
            for ing in ingredients:
                ing["recipe_name"] = header["name"]
            mats = _parse_materials(ws)
            for mat in mats:
                mat["recipe_name"] = header["name"]
            header["materials"] = mats
            header["ingredients_count"] = len(ingredients)
            header["materials_count"] = len(mats)
            recipes.append(header)
            rows.extend(ingredients)
            materials.extend(mats)
    except Exception as exc:  # pragma: no cover - reader errors
        errors.append(f"Error reading Excel: {exc}")

    return {
        "rows": rows,
        "recipes": recipes,
        "materials": materials,
        "headers": ["ingredient", "quantity", "unit", "real_yield", "amount"],
        "metadata": {
            "parser": "xlsx_recipes",
            "source_type": "xlsx",
            "sheets_parsed": len(recipes),
            "total_ingredients": len(rows),
            "total_materials": len(materials),
            "imported_at": datetime.utcnow().isoformat(),
        },
        "detected_type": "recipes",
        "errors": errors,
    }


def _to_decimal(val: Any) -> Decimal | None:
    if val in (None, ""):
        return None
    try:
        return Decimal(str(val).replace(",", "."))
    except Exception:
        return None


def _parse_header(ws) -> dict[str, Any]:
    name_cell = ws["G3"].value
    header = {
        "name": str(name_cell).strip() if name_cell not in (None, "") else None,
        "classification": ws["E4"].value,
        "recipe_type": ws["E5"].value,
        "origin": ws["E6"].value,
        "portions": _to_int(ws["E7"].value),
        "unit_cost": _to_decimal(ws["E8"].value),
        "service_temp": ws["E9"].value,
        "total_ingredients_cost": _find_total_ingredients_cost(ws),
    }
    return header


def _find_total_ingredients_cost(ws) -> Decimal | None:
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=15).value  # column O
        if isinstance(val, str) and val.strip().upper() == "COSTO TOTAL INGREDIENTES":
            return _to_decimal(ws.cell(row=row, column=19).value)  # column S
    return None


def _parse_ingredients(ws) -> list[dict[str, Any]]:
    start = None
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val.strip().lower() == "ingredientes":
            start = row + 1
            break
    if not start:
        return []

    data: list[dict[str, Any]] = []
    for row in range(start, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is None or str(name).strip() == "":
            break
        entry = {
            "ingredient": str(name).strip(),
            "quantity": _to_decimal(ws.cell(row=row, column=6).value),  # F
            "unit": _strip(ws.cell(row=row, column=7).value),  # G
            "real_yield": _to_decimal(ws.cell(row=row, column=15).value),  # O
            "investment": _to_decimal(ws.cell(row=row, column=17).value),  # Q
            "fraction_cost": _to_decimal(ws.cell(row=row, column=18).value),  # R
            "amount": _to_decimal(ws.cell(row=row, column=19).value),  # S
        }
        data.append(entry)
    return data


def _parse_materials(ws) -> list[dict[str, Any]]:
    title_row = None
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and "MATERIALES O INSUMOS ADICIONALES" in val.upper():
            title_row = row
            break
    if not title_row:
        return []
    start = title_row + 2  # skip title + header row
    data: list[dict[str, Any]] = []
    for row in range(start, ws.max_row + 1):
        desc = ws.cell(row=row, column=1).value
        if desc is None or str(desc).strip() == "":
            break
        entry = {
            "description": str(desc).strip(),
            "quantity": _to_decimal(ws.cell(row=row, column=7).value),  # G
            "purchase_unit": _strip(ws.cell(row=row, column=11).value),  # K
            "purchase_price": _to_decimal(ws.cell(row=row, column=17).value),  # Q
            "unit_cost": _to_decimal(ws.cell(row=row, column=18).value),  # R
            "amount": _to_decimal(ws.cell(row=row, column=19).value),  # S
        }
        data.append(entry)
    return data


def _strip(val: Any) -> str | None:
    return str(val).strip() if val not in (None, "") else None


def _to_int(val: Any) -> int | None:
    try:
        return int(val)
    except Exception:
        return None
