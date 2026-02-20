"""
Parser genérico para cualquier archivo Excel.
Auto-detecta estructura y normaliza datos sin importar el banco o formato.
Soporta múltiples hojas y celdas merged.
"""

from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.cell import MergedCell


def _get_merged_cell_value(ws, cell) -> Any:
    """Obtiene el valor de una celda, considerando si es parte de un merge."""
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def _expand_merged_headers(ws, header_row: tuple) -> list[Any]:
    """Expande headers de celdas merged a todas las columnas que cubren."""
    expanded = list(header_row)
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == merged_range.max_row:
            value = ws.cell(merged_range.min_row, merged_range.min_col).value
            for col in range(merged_range.min_col - 1, merged_range.max_col):
                if col < len(expanded):
                    expanded[col] = value
    return expanded


def _detect_header_row(rows: list[tuple], ws=None) -> int:
    """
    Detecta la fila de headers usando múltiples heurísticas:
    - Fila con más celdas llenas
    - Fila después de fila vacía
    - Fila con valores únicos tipo texto
    """
    if not rows:
        return 0

    scores = []
    for idx, row in enumerate(rows[:20]):
        score = 0
        filled = sum(1 for cell in row if cell is not None and str(cell).strip())
        score += filled * 2

        all_text = all(
            isinstance(cell, str) or cell is None
            for cell in row
            if cell is not None and str(cell).strip()
        )
        if all_text and filled > 2:
            score += 5

        if idx > 0:
            prev_row = rows[idx - 1]
            if all(cell is None or str(cell).strip() == "" for cell in prev_row):
                score += 3

        unique_values = {
            str(cell).strip().upper() for cell in row if cell is not None and str(cell).strip()
        }
        if len(unique_values) == filled and filled > 2:
            score += 4

        scores.append((idx, score, filled))

    if not scores:
        return 0

    best = max(scores, key=lambda x: (x[1], x[2]))
    return best[0]


def _normalize_headers(header: tuple | list) -> list[str | None]:
    """Normaliza los headers a strings."""
    normalized = []
    for cell in header:
        if cell is None or str(cell).strip() == "":
            normalized.append(None)
        else:
            normalized.append(str(cell).strip())
    return normalized


def _extract_rows(rows: list[tuple], header_row_idx: int, normalized_headers: list) -> list[dict]:
    """Extrae y normaliza las filas de datos."""
    data_rows = []
    for row_idx, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_dict = {}
        for idx, (header_name, cell_value) in enumerate(zip(normalized_headers, row, strict=False)):
            if header_name is None:
                header_name = f"_col_{idx}"

            if cell_value is None:
                row_dict[header_name] = None
            elif isinstance(cell_value, (int, float)):
                row_dict[header_name] = cell_value
            elif isinstance(cell_value, datetime):
                row_dict[header_name] = cell_value.date().isoformat()
            else:
                row_dict[header_name] = str(cell_value).strip()

        if all(v is None or v == "" for v in row_dict.values()):
            continue

        row_dict["_row"] = row_idx
        row_dict["_imported_at"] = datetime.utcnow().isoformat()
        data_rows.append(row_dict)

    return data_rows


def _detect_document_type(headers: list[str | None]) -> str:
    """Detecta el tipo de documento por keywords en headers."""
    headers_str = " ".join([str(h or "").upper() for h in headers])

    if any(kw in headers_str for kw in ["PRODUCTO", "NOMBRE", "PRECIO", "STOCK", "SKU", "CODIGO"]):
        return "products"
    if any(kw in headers_str for kw in ["FECHA", "IMPORTE", "SALDO", "BANCO", "IBAN", "CUENTA"]):
        return "bank"
    if any(
        kw in headers_str
        for kw in ["FACTURA", "INVOICE", "VENDOR", "SUPPLIER", "CLIENT", "IVA", "TAX"]
    ):
        return "invoices"
    return "generic"


def _parse_single_sheet(ws, file_path: str = "") -> dict[str, Any]:
    """Parsea una sola hoja del workbook."""
    rows = list(ws.iter_rows(values_only=True, max_row=10000))
    if not rows:
        return {
            "name": ws.title,
            "rows": [],
            "headers": [],
            "detected_type": "empty",
            "metadata": {"total_rows": 0},
        }

    header_row_idx = _detect_header_row(rows, ws)
    header = rows[header_row_idx]

    wb_for_merge = ws.parent
    if not getattr(wb_for_merge, "_read_only", True):
        ws_full = wb_for_merge[ws.title]
        header = _expand_merged_headers(ws_full, header)

    normalized_headers = _normalize_headers(header)
    data_rows = _extract_rows(rows, header_row_idx, normalized_headers)
    detected_type = _detect_document_type(normalized_headers)

    return {
        "name": ws.title,
        "rows": data_rows,
        "headers": [h for h in normalized_headers if h is not None],
        "detected_type": detected_type,
        "metadata": {
            "header_row": header_row_idx + 1,
            "total_rows": len(data_rows),
        },
    }


def _are_sheets_compatible(sheet1: dict, sheet2: dict, threshold: float = 0.7) -> bool:
    """Determina si dos hojas tienen estructura similar para combinarlas."""
    h1 = set(sheet1["headers"])
    h2 = set(sheet2["headers"])
    if not h1 or not h2:
        return False
    intersection = h1 & h2
    union = h1 | h2
    similarity = len(intersection) / len(union)
    return similarity >= threshold


def parse_excel_all_sheets(
    file_path: str,
    combine_similar: bool = True,
    similarity_threshold: float = 0.7,
) -> dict[str, Any]:
    """
    Parser que procesa todas las hojas de un archivo Excel.

    Args:
        file_path: Ruta al archivo Excel
        combine_similar: Si True, combina hojas con estructura similar
        similarity_threshold: Umbral de similitud para combinar hojas (0-1)

    Returns:
        Dict con datos de cada hoja, filas combinadas y metadata
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    wb_full = openpyxl.load_workbook(file_path, data_only=True, read_only=False)

    sheets_data = []
    for ws in wb.worksheets:
        ws_full = wb_full[ws.title]
        sheet_result = _parse_single_sheet_with_merged(ws, ws_full, file_path)
        if sheet_result["rows"]:
            sheets_data.append(sheet_result)

    wb.close()
    wb_full.close()

    combined_rows = []
    if combine_similar and sheets_data:
        groups = []
        used = set()

        for i, sheet in enumerate(sheets_data):
            if i in used:
                continue
            group = [sheet]
            used.add(i)

            for j, other_sheet in enumerate(sheets_data):
                if j not in used and _are_sheets_compatible(
                    sheet, other_sheet, similarity_threshold
                ):
                    group.append(other_sheet)
                    used.add(j)
            groups.append(group)

        for group in groups:
            for sheet in group:
                for row in sheet["rows"]:
                    row["_sheet"] = sheet["name"]
                    combined_rows.append(row)

    all_headers = set()
    all_types = set()
    total_rows = 0
    for sheet in sheets_data:
        all_headers.update(sheet["headers"])
        all_types.add(sheet["detected_type"])
        total_rows += len(sheet["rows"])

    primary_type = "generic"
    for t in ["products", "bank", "invoices"]:
        if t in all_types:
            primary_type = t
            break

    return {
        "sheets": sheets_data,
        "combined_rows": combined_rows,
        "metadata": {
            "file_path": file_path,
            "total_sheets": len(sheets_data),
            "total_rows": total_rows,
            "all_headers": list(all_headers),
            "detected_types": list(all_types),
            "primary_type": primary_type,
        },
    }


def _parse_single_sheet_with_merged(ws, ws_full, file_path: str) -> dict[str, Any]:
    """Parsea una hoja considerando celdas merged."""
    rows = list(ws.iter_rows(values_only=True, max_row=10000))
    if not rows:
        return {
            "name": ws.title,
            "rows": [],
            "headers": [],
            "detected_type": "empty",
            "metadata": {"total_rows": 0},
        }

    header_row_idx = _detect_header_row(rows, ws)
    header = rows[header_row_idx]

    header = _expand_merged_headers(ws_full, header)

    normalized_headers = _normalize_headers(header)
    data_rows = _extract_rows(rows, header_row_idx, normalized_headers)
    detected_type = _detect_document_type(normalized_headers)

    return {
        "name": ws.title,
        "rows": data_rows,
        "headers": [h for h in normalized_headers if h is not None],
        "detected_type": detected_type,
        "metadata": {
            "header_row": header_row_idx + 1,
            "total_rows": len(data_rows),
        },
    }


def parse_excel_generic(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """
    Parser universal para Excel de cualquier formato.

    Detecta automáticamente:
    - Headers (busca fila con más celdas llenas, considera merged cells)
    - Tipo de datos (productos, banco, facturas)
    - Columnas relevantes (flexible, múltiples alias)

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre de hoja (default: primera hoja o sheet con más datos)

    Returns:
        Dict con rows normalizadas y metadata
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        wb_full = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    except Exception:
        return _parse_excel_generic_pandas(file_path=file_path, sheet_name=sheet_name)

    if sheet_name:
        ws = wb[sheet_name]
        ws_full = wb_full[sheet_name]
    else:
        best_sheet = None
        max_rows = 0
        for sheet in wb.worksheets:
            rows_count = sum(1 for _ in sheet.iter_rows())
            if rows_count > max_rows:
                max_rows = rows_count
                best_sheet = sheet
        ws = best_sheet or wb.active
        ws_full = wb_full[ws.title]

    rows = list(ws.iter_rows(values_only=True, max_row=10000))
    if not rows:
        wb.close()
        wb_full.close()
        return {"rows": [], "headers": [], "metadata": {}, "detected_type": "empty"}

    header_row_idx = _detect_header_row(rows, ws)
    header = rows[header_row_idx]

    header = _expand_merged_headers(ws_full, header)

    normalized_headers = _normalize_headers(header)
    data_rows = _extract_rows(rows, header_row_idx, normalized_headers)
    detected_type = _detect_document_type(normalized_headers)

    metadata = {
        "sheet_name": ws.title,
        "header_row": header_row_idx + 1,
        "total_rows": len(data_rows),
        "detected_type": detected_type,
        "file_path": file_path,
        "has_merged_cells": bool(ws_full.merged_cells.ranges),
    }

    wb.close()
    wb_full.close()

    return {
        "rows": data_rows,
        "headers": [h for h in normalized_headers if h is not None],
        "metadata": metadata,
        "detected_type": detected_type,
    }


def _parse_excel_generic_pandas(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """Fallback parser for formats not supported by openpyxl (e.g., legacy .xls)."""
    try:
        import pandas as pd
    except Exception as e:
        return {
            "rows": [],
            "headers": [],
            "metadata": {"file_path": file_path, "error": f"pandas_not_available: {e}"},
            "detected_type": "generic",
            "errors": [f"pandas_not_available: {e}"],
        }

    engine = "xlrd" if file_path.lower().endswith(".xls") else None
    try:
        df = pd.read_excel(file_path, engine=engine, header=None, sheet_name=sheet_name or 0)
    except Exception as e:
        return {
            "rows": [],
            "headers": [],
            "metadata": {"file_path": file_path, "error": str(e)},
            "detected_type": "generic",
            "errors": [str(e)],
        }

    df = df.fillna("")
    rows = [tuple(row) for row in df.values.tolist()]
    if not rows:
        return {"rows": [], "headers": [], "metadata": {}, "detected_type": "empty"}

    header_row_idx = _detect_header_row(rows)
    header = rows[header_row_idx]
    normalized_headers = _normalize_headers(header)
    data_rows = _extract_rows(rows, header_row_idx, normalized_headers)
    detected_type = _detect_document_type(normalized_headers)

    return {
        "rows": data_rows,
        "headers": [h for h in normalized_headers if h is not None],
        "metadata": {
            "sheet_name": sheet_name or "sheet1",
            "header_row": header_row_idx + 1,
            "total_rows": len(data_rows),
            "detected_type": detected_type,
            "file_path": file_path,
            "has_merged_cells": False,
            "reader": "pandas",
        },
        "detected_type": detected_type,
    }
