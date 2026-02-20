"""
Robust Excel parser with unified header detection + row cleaning.
Handles problematic .xls/.xlsx files with:
- Garbage rows (instructions, summaries, blanks)
- Inconsistent headers
- Merged cells
- Formatting issues
"""

import logging
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("imports.robust_excel")


class RobustExcelParser:
    """
    Unified Excel parser for analyze + parse operations.
    
    Guarantees same header detection and row cleaning in both phases.
    """
    
    # Keywords indicating a row is junk/instructions
    JUNK_KEYWORDS = {
        "instruccion", "instrucciones", "instruction", "como", "apuntar",
        "rellenar", "llenar", "completar", "ejemplo", "example",
        "nota", "notas", "note", "notes", "aviso", "warning",
        "observacion", "observaciones", "observation",
        "por favor", "please", "favor", "requiere", "requires",
        "formato", "format", "manera", "modo", "modo de", "way",
        "guia", "guide", "ayuda", "help", "información",
        "descripción del", "field", "column", "columna"
    }
    
    # Minimum data density for a row to be considered real data
    MIN_DENSITY_FOR_HEADER = 0.5
    MIN_COLUMNS_FOR_HEADER = 3
    
    def __init__(self, min_header_columns: int = 3, max_header_scan_rows: int = 10):
        self.min_header_columns = min_header_columns
        self.max_header_scan_rows = max_header_scan_rows
    
    def analyze_file(self, file_path: str) -> dict[str, Any]:
        """
        Analyze Excel file: returns headers, sample rows, metadata.
        Same logic as parse_file but returns analysis instead of full data.
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            header_row_idx = self._detect_header_row(ws)
            headers = self._extract_headers(ws, header_row_idx)
            sample_rows = self._extract_sample_rows(ws, header_row_idx, headers, limit=5)
            
            wb.close()
            
            return {
                "headers": headers,
                "header_row": header_row_idx,
                "sample_rows": sample_rows,
                "row_count": ws.max_row - header_row_idx,
                "total_rows_including_header": ws.max_row,
                "success": True,
            }
        except Exception as e:
            logger.error(f"Error analyzing Excel: {e}")
            return {
                "headers": [],
                "header_row": 1,
                "sample_rows": [],
                "row_count": 0,
                "success": False,
                "error": str(e),
            }
    
    def parse_file(self, file_path: str) -> dict[str, Any]:
        """
        Parse entire Excel file: returns all data rows cleaned + validated.
        Uses same header detection as analyze_file.
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            header_row_idx = self._detect_header_row(ws)
            headers = self._extract_headers(ws, header_row_idx)
            
            rows = []
            errors = []
            
            # Extract all data rows (post-header)
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                try:
                    row_data = self._extract_row(ws, row_idx, headers)
                    
                    # Skip junk rows
                    if self._is_junk_row(row_data, headers):
                        continue
                    
                    # Skip empty rows
                    if not any(v for v in row_data.values()):
                        continue
                    
                    row_data["_row_number"] = row_idx
                    rows.append(row_data)
                except Exception as e:
                    logger.warning(f"Error extracting row {row_idx}: {e}")
                    errors.append({
                        "row": row_idx,
                        "error": str(e),
                    })
            
            wb.close()
            
            return {
                "headers": headers,
                "header_row": header_row_idx,
                "rows": rows,
                "row_count": len(rows),
                "errors": errors,
                "success": True,
            }
        except Exception as e:
            logger.error(f"Error parsing Excel: {e}")
            return {
                "headers": [],
                "rows": [],
                "row_count": 0,
                "errors": [{"error": str(e)}],
                "success": False,
            }
    
    def _detect_header_row(self, ws: Worksheet) -> int:
        """
        Unified header detection.
        
        Scans first N rows and finds row that:
        1. Has high text density (>= 50%)
        2. Has at least MIN_COLUMNS_FOR_HEADER columns with data
        3. Does NOT contain junk keywords
        4. Contains mostly text (not numbers)
        
        Returns: 1-based row index
        """
        for row_idx in range(1, min(self.max_header_scan_rows + 1, ws.max_row + 1)):
            row_values = list(self._extract_row_raw(ws, row_idx).values())
            
            # Skip completely empty rows
            non_empty = [v for v in row_values if v is not None and str(v).strip()]
            if len(non_empty) < self.min_header_columns:
                continue
            
            # Check for junk keywords in first cell
            first_cell_text = str(row_values[0] or "").lower().strip() if row_values else ""
            if self._contains_junk_keyword(first_cell_text):
                logger.debug(f"Row {row_idx} skipped (junk keyword in first cell)")
                continue
            
            # Count text cells (not pure numbers)
            text_count = sum(
                1 for v in row_values
                if v is not None
                and isinstance(v, str)
                and v.strip()
                and not v.strip().lstrip("-").isdigit()
            )
            
            # Check density: text cells >= MIN_DENSITY_FOR_HEADER of non-empty cells
            text_density = text_count / max(len(non_empty), 1)
            if text_density >= self.MIN_DENSITY_FOR_HEADER:
                logger.debug(
                    f"Detected header at row {row_idx} "
                    f"(text_count={text_count}, density={text_density:.2f})"
                )
                return row_idx
        
        logger.warning("No clear header detected, defaulting to row 1")
        return 1
    
    def _extract_headers(self, ws: Worksheet, header_row_idx: int) -> list[str]:
        """
        Extract and clean header names.
        
        Removes None, empty strings, and standardizes.
        """
        headers = []
        row_values = self._extract_row_raw(ws, header_row_idx)
        
        for col_idx, (_, value) in enumerate(row_values.items()):
            if value is None or (isinstance(value, str) and not value.strip()):
                # Check if there's data below in this column
                has_data = any(
                    ws.cell(r, col_idx + 1).value
                    for r in range(header_row_idx + 1, min(header_row_idx + 10, ws.max_row + 1))
                    if ws.cell(r, col_idx + 1).value
                )
                if has_data:
                    headers.append(f"Column_{col_idx + 1}")
                else:
                    break  # No more data in this column
            else:
                headers.append(str(value).strip())
        
        return headers
    
    def _extract_row(self, ws: Worksheet, row_idx: int, headers: list[str]) -> dict[str, Any]:
        """
        Extract single row as dict, mapped to headers.
        Handles missing columns gracefully.
        """
        row_dict = {}
        for col_idx, header in enumerate(headers, start=1):
            try:
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                # Normalize: strip whitespace from strings
                if isinstance(value, str):
                    value = value.strip()
                row_dict[header] = value
            except Exception:
                row_dict[header] = None
        
        return row_dict
    
    def _extract_row_raw(self, ws: Worksheet, row_idx: int) -> dict[int, Any]:
        """
        Extract raw row by column index (0-based keys).
        Used for header detection and analysis.
        
        Returns: {col_idx: value, ...}
        """
        row_dict = {}
        for col_idx in range(1, ws.max_column + 1):
            try:
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                if isinstance(value, str):
                    value = value.strip()
                row_dict[col_idx - 1] = value
            except Exception:
                row_dict[col_idx - 1] = None
        
        return row_dict
    
    def _extract_sample_rows(
        self,
        ws: Worksheet,
        header_row_idx: int,
        headers: list[str],
        limit: int = 5,
    ) -> list[dict[str, Any]]:
        """Extract sample data rows (clean junk)."""
        sample = []
        
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            if len(sample) >= limit:
                break
            
            try:
                row_data = self._extract_row(ws, row_idx, headers)
                
                if self._is_junk_row(row_data, headers):
                    continue
                
                if not any(v for v in row_data.values()):
                    continue
                
                sample.append(row_data)
            except Exception:
                continue
        
        return sample
    
    def _is_junk_row(self, row_data: dict[str, Any], headers: list[str]) -> bool:
        """
        Check if row is junk (instructions, summary, etc).
        
        Heuristics:
        - First column contains junk keywords
        - Row is completely empty
        - Row is all formulas/calculations
        """
        if not row_data:
            return True
        
        # Check first non-empty cell for junk keywords
        for col_idx, (header, value) in enumerate(row_data.items()):
            if value is not None and str(value).strip():
                text = str(value).lower().strip()
                if self._contains_junk_keyword(text):
                    logger.debug(f"Junk row detected: {header}={value}")
                    return True
                break
        
        # If all values are None, it's empty
        if not any(v for v in row_data.values()):
            return True
        
        return False
    
    def _contains_junk_keyword(self, text: str) -> bool:
        """Check if text contains junk keywords."""
        text_lower = text.lower()
        return any(kw in text_lower for kw in self.JUNK_KEYWORDS)


# Global instance
robust_parser = RobustExcelParser()
