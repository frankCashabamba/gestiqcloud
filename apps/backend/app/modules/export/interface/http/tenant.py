from __future__ import annotations

import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Request, Response
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.config.database import get_db
from app.core.access_guard import with_access_claims
from app.core.authz import require_scope
from app.db.rls import ensure_rls, tenant_id_from_request

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter(
    prefix="/export",
    tags=["Export"],
    dependencies=[
        Depends(with_access_claims),
        Depends(require_scope("tenant")),
        Depends(ensure_rls),
    ],
)


def _csv(rows, header):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    for r in rows:
        w.writerow([r.get(h) for h in header])
    return buf.getvalue().encode("utf-8")


@router.get("/products.csv", response_class=Response)
def export_products(request: Request, db: Session = Depends(get_db)):
    tenant_id = tenant_id_from_request(request)
    if tenant_id is None:
        raise HTTPException(status_code=403, detail="missing_tenant")
    rows = db.execute(
        text(
            "SELECT id, COALESCE(sku,'') AS sku, name, price, unit "
            "FROM products WHERE tenant_id=:tid ORDER BY id"
        ),
        {"tid": tenant_id},
    )
    items = [dict(r) for r in rows.mappings().all()]
    data = _csv(items, ["id", "sku", "name", "price", "unit"])
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=products.csv"},
    )


@router.get("/clients.csv", response_class=Response)
def export_clients(request: Request, db: Session = Depends(get_db)):
    tenant_id = tenant_id_from_request(request)
    if tenant_id is None:
        raise HTTPException(status_code=403, detail="missing_tenant")
    rows = db.execute(
        text(
            "SELECT id, nombre, COALESCE(email,'') AS email, "
            "COALESCE(telefono,'') AS telefono FROM clients "
            "WHERE tenant_id=:tid ORDER BY id"
        ),
        {"tid": tenant_id},
    )
    items = [dict(r) for r in rows.mappings().all()]
    data = _csv(items, ["id", "nombre", "email", "telefono"])
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clients.csv"},
    )


@router.get("/stock.csv", response_class=Response)
def export_stock(request: Request, db: Session = Depends(get_db)):
    tenant_id = tenant_id_from_request(request)
    if tenant_id is None:
        raise HTTPException(status_code=403, detail="missing_tenant")
    rows = db.execute(
        text(
            "SELECT w.code, p.sku, p.name, s.qty FROM stock_items s "
            "LEFT JOIN warehouses w ON s.warehouse_id = w.id "
            "LEFT JOIN products p ON s.product_id = p.id "
            "WHERE s.tenant_id=:tid ORDER BY w.code, p.sku"
        ),
        {"tid": tenant_id},
    )
    items = [dict(r) for r in rows.mappings().all()]
    data = _csv(items, ["code", "sku", "name", "qty"])
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=stock.csv"},
    )


@router.get("/stock.xlsx", response_class=Response)
def export_stock_xlsx(request: Request, db: Session = Depends(get_db)):
    """Export stock inventory as Excel"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="Excel export not available")

    tenant_id = tenant_id_from_request(request)
    if tenant_id is None:
        raise HTTPException(status_code=403, detail="missing_tenant")

    rows = db.execute(
        text(
            "SELECT w.code, p.sku, p.name, s.qty, p.price, (s.qty * p.price) as total "
            "FROM stock_items s "
            "LEFT JOIN warehouses w ON s.warehouse_id = w.id "
            "LEFT JOIN products p ON s.product_id = p.id "
            "WHERE s.tenant_id=:tid ORDER BY w.code, p.sku"
        ),
        {"tid": tenant_id},
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Existencias"

    # Headers
    headers = ["Almacén", "Código", "Producto", "Cantidad", "Precio Unit.", "Total Valor"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    for row in rows:
        ws.append(
            [
                row[0] or "",
                row[1] or "",
                row[2] or "",
                float(row[3] or 0),
                float(row[4] or 0),
                float(row[5] or 0),
            ]
        )

    # Adjust column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15

    # Format numbers
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = "#,##0.00"

    # Generate file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Informe_de_Existencias_Actuales_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
