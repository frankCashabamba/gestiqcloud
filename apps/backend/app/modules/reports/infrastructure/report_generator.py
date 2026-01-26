"""Report generation service"""

import csv
import io
import json
import logging
from abc import ABC, abstractmethod

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.modules.reports.domain.entities import (
    FinancialReport,
    InventoryReport,
    ReportData,
    ReportDefinition,
    ReportFormat,
    ReportType,
    SalesReport,
)

logger = logging.getLogger(__name__)


class BaseReportGenerator(ABC):
    """Abstract base for report generators"""

    @abstractmethod
    def generate(self, definition: ReportDefinition, db: Session) -> ReportData:
        """Generate report data"""
        pass


class SalesReportGenerator(BaseReportGenerator):
    """Sales report generator"""

    def generate(self, definition: ReportDefinition, db: Session) -> SalesReport:
        """Generate sales report"""
        try:
            query = """
                SELECT
                    DATE(so.created_at) as fecha,
                    COUNT(DISTINCT so.id) as num_pedidos,
                    SUM(soi.qty) as total_items,
                    SUM(soi.qty * soi.unit_price) as total_venta
                FROM sales_orders so
                LEFT JOIN sales_order_items soi ON so.id = soi.order_id
                WHERE so.tenant_id = :tenant_id
            """

            params = {"tenant_id": definition.tenant_id}

            # Add date filters
            if definition.date_from:
                query += " AND so.created_at >= :date_from"
                params["date_from"] = definition.date_from

            if definition.date_to:
                query += " AND so.created_at <= :date_to"
                params["date_to"] = definition.date_to

            query += " GROUP BY DATE(so.created_at) ORDER BY fecha DESC"

            result = db.execute(text(query), params).fetchall()

            # Calculate totals
            total_sales = sum(row[3] or 0 for row in result)
            total_items = sum(row[2] or 0 for row in result)
            order_count = len(result)
            avg_order_value = total_sales / order_count if order_count > 0 else 0

            columns = ["Fecha", "Pedidos", "Items", "Total"]
            rows = [[str(row[0]), row[1], row[2], f"${row[3]:.2f}"] for row in result]

            return SalesReport(
                columns=columns,
                rows=rows,
                total_sales=total_sales,
                total_items=total_items,
                order_count=order_count,
                average_order_value=avg_order_value,
                summary={
                    "total_sales": total_sales,
                    "total_items": total_items,
                    "order_count": order_count,
                    "average_order_value": avg_order_value,
                },
            )

        except Exception as e:
            logger.error(f"Failed to generate sales report: {e}")
            raise


class InventoryReportGenerator(BaseReportGenerator):
    """Inventory report generator"""

    def generate(self, definition: ReportDefinition, db: Session) -> InventoryReport:
        """Generate inventory report"""
        try:
            query = """
                SELECT
                    p.id,
                    p.nombre,
                    COALESCE(i.cantidad, 0) as stock,
                    COALESCE(i.precio_unitario, 0) as precio,
                    COALESCE(i.cantidad * i.precio_unitario, 0) as valor_total
                FROM products p
                LEFT JOIN inventario i ON p.id = i.product_id
                WHERE p.tenant_id = :tenant_id
                ORDER BY stock ASC
            """

            result = db.execute(text(query), {"tenant_id": definition.tenant_id}).fetchall()

            # Calculate statistics
            total_items = len(result)
            low_stock = sum(1 for row in result if row[2] < 10)  # Assuming low = < 10
            out_of_stock = sum(1 for row in result if row[2] == 0)
            total_value = sum(row[4] or 0 for row in result)

            columns = ["Producto", "Stock", "Precio Unit.", "Valor Total"]
            rows = [[row[1], row[2], f"${row[3]:.2f}", f"${row[4]:.2f}"] for row in result]

            return InventoryReport(
                columns=columns,
                rows=rows,
                total_items=total_items,
                low_stock_count=low_stock,
                out_of_stock_count=out_of_stock,
                total_value=total_value,
                summary={
                    "total_items": total_items,
                    "low_stock": low_stock,
                    "out_of_stock": out_of_stock,
                    "total_value": total_value,
                },
            )

        except Exception as e:
            logger.error(f"Failed to generate inventory report: {e}")
            raise


class FinancialReportGenerator(BaseReportGenerator):
    """Financial report generator"""

    def generate(self, definition: ReportDefinition, db: Session) -> FinancialReport:
        """Generate financial report"""
        try:
            # Revenue from invoices
            revenue_query = """
                SELECT COALESCE(SUM(total), 0)
                FROM invoices
                WHERE tenant_id = :tenant_id
                AND estado IN ('emitida', 'pagada')
            """

            # Expenses from purchase orders
            expense_query = """
                SELECT COALESCE(SUM(total), 0)
                FROM purchase_orders
                WHERE tenant_id = :tenant_id
                AND status != 'cancelled'
            """

            params = {"tenant_id": definition.tenant_id}

            total_revenue = db.execute(text(revenue_query), params).scalar() or 0
            total_expenses = db.execute(text(expense_query), params).scalar() or 0
            net_profit = total_revenue - total_expenses
            profit_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

            columns = ["Concepto", "Cantidad"]
            rows = [
                ["Ingresos", f"${total_revenue:,.2f}"],
                ["Gastos", f"${total_expenses:,.2f}"],
                ["Ganancia Neta", f"${net_profit:,.2f}"],
                ["Margen (%)", f"{profit_margin:.2f}%"],
            ]

            return FinancialReport(
                columns=columns,
                rows=rows,
                total_revenue=total_revenue,
                total_expenses=total_expenses,
                net_profit=net_profit,
                profit_margin=profit_margin,
                summary={
                    "total_revenue": total_revenue,
                    "total_expenses": total_expenses,
                    "net_profit": net_profit,
                    "profit_margin": profit_margin,
                },
            )

        except Exception as e:
            logger.error(f"Failed to generate financial report: {e}")
            raise


class ReportExporter:
    """Export reports to various formats"""

    @staticmethod
    def to_csv(data: ReportData) -> bytes:
        """Export to CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write columns
        writer.writerow(data.columns)

        # Write rows
        for row in data.rows:
            writer.writerow(row)

        # Write summary if present
        if data.summary:
            writer.writerow([])
            writer.writerow(["RESUMEN"])
            for key, value in data.summary.items():
                writer.writerow([key, value])

        return output.getvalue().encode("utf-8")

    @staticmethod
    def to_json(data: ReportData) -> bytes:
        """Export to JSON"""
        json_data = {
            "columns": data.columns,
            "rows": data.rows,
            "summary": data.summary,
            "metadata": data.metadata,
        }
        return json.dumps(json_data, indent=2, default=str).encode("utf-8")

    @staticmethod
    def to_excel(data: ReportData) -> bytes:
        """Export to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # Write header
            for col_idx, header in enumerate(data.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            # Write data
            for row_idx, row in enumerate(data.rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Write summary
            if data.summary:
                summary_row = len(data.rows) + 3
                ws.cell(row=summary_row, column=1, value="RESUMEN")
                for idx, (key, value) in enumerate(data.summary.items()):
                    ws.cell(row=summary_row + idx + 1, column=1, value=key)
                    ws.cell(row=summary_row + idx + 1, column=2, value=value)

            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise

    @staticmethod
    def to_pdf(data: ReportData, title: str = "Report") -> bytes:
        """Export to PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.pdfgen import canvas
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []

            # Title
            styles = getSampleStyleSheet()
            elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))

            # Create table
            table_data = [data.columns] + data.rows
            table = Table(table_data)

            # Style table
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            elements.append(table)

            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            return buffer.getvalue()

        except ImportError:
            logger.error("reportlab not installed. Install with: pip install reportlab")
            raise

    @staticmethod
    def to_html(data: ReportData, title: str = "Report") -> str:
        """Export to HTML"""
        html = f"""
        <html>
        <head>
            <title>{title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th {{ background-color: #4CAF50; color: white; padding: 12px; text-align: left; }}
                td {{ border: 1px solid #ddd; padding: 8px; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .summary {{ margin-top: 20px; }}
                .summary strong {{ color: #333; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <table>
                <tr>
                    {''.join(f'<th>{col}</th>' for col in data.columns)}
                </tr>
                {''.join(f'<tr>{"".join(f"<td>{val}</td>" for val in row)}</tr>' for row in data.rows)}
            </table>
        """

        if data.summary:
            html += '<div class="summary"><h3>Resumen</h3>'
            for key, value in data.summary.items():
                html += f"<p><strong>{key}:</strong> {value}</p>"
            html += "</div>"

        html += """
        </body>
        </html>
        """

        return html


class ReportService:
    """Main report service"""

    def __init__(self, db: Session):
        self.db = db
        self.generators = {
            ReportType.SALES_SUMMARY: SalesReportGenerator(),
            ReportType.INVENTORY_STATUS: InventoryReportGenerator(),
            ReportType.PROFIT_LOSS: FinancialReportGenerator(),
        }
        self.exporter = ReportExporter()

    def generate_report(self, definition: ReportDefinition, export_format: ReportFormat) -> bytes:
        """Generate and export report"""
        try:
            # Get generator
            generator = self.generators.get(definition.report_type)
            if not generator:
                raise ValueError(f"Unsupported report type: {definition.report_type}")

            # Generate report data
            data = generator.generate(definition, self.db)

            # Export to format
            if export_format == ReportFormat.CSV:
                return self.exporter.to_csv(data)
            elif export_format == ReportFormat.EXCEL:
                return self.exporter.to_excel(data)
            elif export_format == ReportFormat.JSON:
                return self.exporter.to_json(data)
            elif export_format == ReportFormat.PDF:
                return self.exporter.to_pdf(data, definition.name)
            elif export_format == ReportFormat.HTML:
                return self.exporter.to_html(data, definition.name).encode("utf-8")
            else:
                raise ValueError(f"Unsupported export format: {export_format}")

        except Exception as e:
            logger.error(f"Failed to generate report: {e}")
            raise
