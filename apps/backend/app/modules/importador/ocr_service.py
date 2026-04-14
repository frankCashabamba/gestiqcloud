"""Servicio OCR y extracción de texto para el Importador."""

from __future__ import annotations

import base64
import csv
import datetime
import hashlib
import io
import itertools
import json
import logging
import re
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable
from pathlib import Path
from threading import Lock
from typing import Any

import openpyxl
from PIL import Image, ImageFilter, ImageOps

try:
    import pandas as pd
except Exception:  # pragma: no cover - optional dependency
    pd = None

from .ocr_quality import estimate_text_quality as _estimate_text_quality
from .runtime_config import load_file_support_config, load_ocr_runtime_config
from .utils import json_safe as _json_safe

logger = logging.getLogger("importador.ocr")

try:
    from pillow_heif import register_heif_opener

    register_heif_opener()
except Exception:
    logger.debug("pillow-heif no disponible; HEIC/HEIF pueden no abrirse", exc_info=True)

_DEFAULT_FILE_SUPPORT = load_file_support_config(None)
SUPPORTED_EXTENSIONS = set(_DEFAULT_FILE_SUPPORT["accepted_extensions"])
IMAGE_EXTENSIONS = set(_DEFAULT_FILE_SUPPORT["image_extensions"])
OCR_EXTRACTION_CACHE_VERSION = "2026-04-14-1"
_EASYOCR_READERS: dict[tuple[tuple[str, ...], bool], Any] = {}
_EASYOCR_READER_LOCK = Lock()

# UBL 2.1 namespaces
_UBL_NS = {
    "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
    "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
}


def _ocr_cache_dir() -> Path:
    from app.config.settings import settings

    raw_dir = settings.uploads_path / "_importador_ocr_cache"
    raw_dir.mkdir(parents=True, exist_ok=True)
    return raw_dir


def _ocr_cache_path(file_bytes: bytes) -> Path:
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    return _ocr_cache_dir() / f"{file_hash}.json"


def _serialize_cached_extraction(extraction: dict[str, Any]) -> dict[str, Any]:
    payload = {
        "version": OCR_EXTRACTION_CACHE_VERSION,
        "text": str(extraction.get("text") or ""),
        "pages": int(extraction.get("pages") or 1),
        "structured_data": _json_safe(extraction.get("structured_data")),
        "format": str(extraction.get("format") or ""),
        "page_texts": _json_safe(extraction.get("page_texts")),
        "sheet_profiles": _json_safe(extraction.get("sheet_profiles")),
        "sheet_metadata": _json_safe(extraction.get("sheet_metadata")),
        "sheet_used": extraction.get("sheet_used"),
    }
    vision_image_bytes = extraction.get("vision_image_bytes")
    if isinstance(vision_image_bytes, (bytes, bytearray)):
        payload["vision_image_b64"] = base64.b64encode(bytes(vision_image_bytes)).decode("ascii")
    return payload


def _deserialize_cached_extraction(payload: dict[str, Any]) -> dict[str, Any]:
    extraction = {
        "text": str(payload.get("text") or ""),
        "pages": int(payload.get("pages") or 1),
        "structured_data": payload.get("structured_data"),
        "format": str(payload.get("format") or ""),
    }
    if payload.get("page_texts") is not None:
        extraction["page_texts"] = payload.get("page_texts")
    if payload.get("sheet_profiles") is not None:
        extraction["sheet_profiles"] = payload.get("sheet_profiles")
    if payload.get("sheet_metadata") is not None:
        extraction["sheet_metadata"] = payload.get("sheet_metadata")
    if payload.get("sheet_used") is not None:
        extraction["sheet_used"] = payload.get("sheet_used")
    vision_image_b64 = payload.get("vision_image_b64")
    if isinstance(vision_image_b64, str) and vision_image_b64:
        try:
            extraction["vision_image_bytes"] = base64.b64decode(vision_image_b64.encode("ascii"))
        except Exception:
            logger.warning("No se pudo decodificar vision_image_b64 de cache OCR", exc_info=True)
    return _rehydrate_virtual_sheet_context(extraction)


def _can_cache_extraction(extraction: dict[str, Any]) -> bool:
    fmt = str(extraction.get("format") or "").upper()
    if fmt.endswith("_ERROR"):
        return False
    text = str(extraction.get("text") or "").strip()
    structured_data = extraction.get("structured_data")
    page_texts = extraction.get("page_texts")
    sheet_profiles = extraction.get("sheet_profiles")
    vision_image_bytes = extraction.get("vision_image_bytes")
    return bool(
        text
        or structured_data
        or page_texts
        or sheet_profiles
        or isinstance(vision_image_bytes, (bytes, bytearray))
    )


def _load_cached_extraction(file_bytes: bytes) -> dict[str, Any] | None:
    cache_path = _ocr_cache_path(file_bytes)
    if not cache_path.exists():
        return None
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        logger.warning("No se pudo leer cache OCR %s", cache_path, exc_info=True)
        return None
    if payload.get("version") != OCR_EXTRACTION_CACHE_VERSION:
        return None
    return _deserialize_cached_extraction(payload)


def _store_cached_extraction(file_bytes: bytes, extraction: dict[str, Any]) -> None:
    if not _can_cache_extraction(extraction):
        return
    cache_path = _ocr_cache_path(file_bytes)
    payload = _serialize_cached_extraction(extraction)
    tmp_path = cache_path.with_suffix(".tmp")
    try:
        tmp_path.write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")
        tmp_path.replace(cache_path)
    except Exception:
        logger.warning("No se pudo guardar cache OCR %s", cache_path, exc_info=True)
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


def _image_to_jpeg_bytes(img: Image.Image, *, quality: int = 80) -> bytes:
    """Serialize a PIL image to JPEG bytes for optional vision fallback."""
    buffer = io.BytesIO()
    safe_img = img.convert("RGB") if img.mode != "RGB" else img
    safe_img.save(buffer, format="JPEG", quality=quality)
    return buffer.getvalue()


def _get_file_support_sets(db: Any | None = None) -> tuple[set[str], set[str], dict[str, str]]:
    config = load_file_support_config(db)
    return (
        set(config.get("accepted_extensions") or []),
        set(config.get("image_extensions") or []),
        dict(config.get("type_map") or {}),
    )


def detect_file_type(filename: str, db: Any | None = None) -> str:
    ext = Path(filename).suffix.lower()
    _, _, type_map = _get_file_support_sets(db)
    return type_map.get(ext, "UNKNOWN")


def _ocr_text_score(text: str) -> tuple[int, int]:
    raw = str(text or "").strip()
    if not raw:
        return (0, 0)
    words = [token for token in raw.replace("\n", " ").split(" ") if token.strip()]
    return (len(words), len(raw))


def _ocr_runtime_config() -> dict[str, Any]:
    return load_ocr_runtime_config(None)


def _is_weak_ocr_text(
    text: str,
    *,
    min_words: int | None = None,
    min_chars: int | None = None,
    min_quality: float | None = None,
) -> bool:
    config = _ocr_runtime_config()
    min_words = config["weak_text_min_words"] if min_words is None else min_words
    min_chars = config["weak_text_min_chars"] if min_chars is None else min_chars
    min_quality = (
        float(config.get("ocr_min_quality") or 0.45) if min_quality is None else float(min_quality)
    )
    quality = _estimate_text_quality(text, ocr_runtime=config)
    if float(quality.get("score") or 0.0) < min_quality:
        return True
    words, chars = _ocr_text_score(text)
    return words < min_words or chars < min_chars


def _copy_with_label(img: Image.Image, label: str) -> Image.Image:
    clone = img.copy()
    clone.info["ocr_variant"] = label
    return clone


def _channel_grayscale(img: Image.Image, channel_index: int) -> Image.Image:
    rgb = img.convert("RGB") if img.mode != "RGB" else img
    channel = rgb.split()[channel_index]
    return channel if channel.mode == "L" else channel.convert("L")


def _trim_document_edges(img: Image.Image) -> Image.Image:
    """Crop obvious outer margins/background without document-specific assumptions."""
    config = _ocr_runtime_config()
    gray = img.convert("L") if img.mode != "L" else img
    mask = gray.point(lambda px: 255 if px < int(config["trim_background_threshold"]) else 0)
    bbox = mask.getbbox()
    if not bbox:
        return img.copy()

    left, top, right, bottom = bbox
    width = max(1, img.width)
    height = max(1, img.height)
    crop_w = max(1, right - left)
    crop_h = max(1, bottom - top)

    min_crop_ratio = float(config["trim_min_crop_ratio"])
    if crop_w / width < min_crop_ratio or crop_h / height < min_crop_ratio:
        return img.copy()

    pad_ratio = float(config["trim_padding_ratio"])
    min_padding_px = int(config["trim_min_padding_px"])
    pad_x = max(min_padding_px, int(crop_w * pad_ratio))
    pad_y = max(min_padding_px, int(crop_h * pad_ratio))
    cropped = img.crop(
        (
            max(0, left - pad_x),
            max(0, top - pad_y),
            min(width, right + pad_x),
            min(height, bottom + pad_y),
        )
    )
    return cropped


def _order_quad_points(points) -> Any:
    import numpy as np

    pts = np.array(points, dtype="float32")
    ordered = np.zeros((4, 2), dtype="float32")
    sums = pts.sum(axis=1)
    diffs = np.diff(pts, axis=1)
    ordered[0] = pts[sums.argmin()]
    ordered[2] = pts[sums.argmax()]
    ordered[1] = pts[diffs.argmin()]
    ordered[3] = pts[diffs.argmax()]
    return ordered


def _rectify_document_perspective(img: Image.Image) -> tuple[Image.Image, bool]:
    """Detect a dominant document-like contour and rectify it generically."""
    try:
        import cv2
        import numpy as np
    except Exception:
        return img.copy(), False

    config = _ocr_runtime_config()
    rgb = img.convert("RGB")
    rgb_np = np.array(rgb)
    gray = cv2.cvtColor(rgb_np, cv2.COLOR_RGB2GRAY)
    blur_kernel_size = max(3, int(config["perspective_blur_kernel"]))
    if blur_kernel_size % 2 == 0:
        blur_kernel_size += 1
    blurred = cv2.GaussianBlur(gray, (blur_kernel_size, blur_kernel_size), 0)
    edges = cv2.Canny(
        blurred,
        int(config["perspective_canny_threshold1"]),
        int(config["perspective_canny_threshold2"]),
    )
    kernel_size = max(1, int(config["perspective_kernel_size"]))
    kernel = np.ones((kernel_size, kernel_size), dtype="uint8")
    edges = cv2.dilate(edges, kernel, iterations=1)
    edges = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, kernel, iterations=2)

    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return rgb.copy(), False

    height, width = gray.shape[:2]
    image_area = max(1, width * height)

    min_area_ratio = float(config["perspective_min_area_ratio"])
    min_output_ratio = float(config["perspective_min_output_ratio"])
    for contour in sorted(contours, key=cv2.contourArea, reverse=True):
        area = cv2.contourArea(contour)
        if area / image_area < min_area_ratio:
            continue

        perimeter = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.02 * perimeter, True)
        quad = None
        if len(approx) == 4:
            quad = approx.reshape(4, 2)
        else:
            rect = cv2.minAreaRect(contour)
            box = cv2.boxPoints(rect)
            if cv2.contourArea(box) / image_area < min_area_ratio:
                continue
            quad = box

        ordered = _order_quad_points(quad)
        (top_left, top_right, bottom_right, bottom_left) = ordered

        width_a = float(np.linalg.norm(bottom_right - bottom_left))
        width_b = float(np.linalg.norm(top_right - top_left))
        height_a = float(np.linalg.norm(top_right - bottom_right))
        height_b = float(np.linalg.norm(top_left - bottom_left))

        max_width = int(max(width_a, width_b))
        max_height = int(max(height_a, height_b))
        if max_width < width * min_output_ratio or max_height < height * min_output_ratio:
            continue

        destination = np.array(
            [
                [0, 0],
                [max_width - 1, 0],
                [max_width - 1, max_height - 1],
                [0, max_height - 1],
            ],
            dtype="float32",
        )
        transform = cv2.getPerspectiveTransform(ordered, destination)
        warped = cv2.warpPerspective(
            rgb_np,
            transform,
            (max_width, max_height),
            flags=cv2.INTER_CUBIC,
            borderMode=cv2.BORDER_REPLICATE,
        )
        warped_img = Image.fromarray(warped)
        if warped_img.width < 32 or warped_img.height < 32:
            continue
        if warped_img.size == rgb.size:
            return warped_img, True
        return warped_img, True

    return rgb.copy(), False


def _iter_small_rotations(
    img: Image.Image,
    *,
    label_prefix: str,
    angles: tuple[float, ...] | None = None,
) -> list[Image.Image]:
    """Generate mild deskew candidates without assuming a specific document layout."""
    if angles is None:
        config = _ocr_runtime_config()
        parsed_angles: list[float] = []
        for raw in config.get("small_rotation_angles") or []:
            try:
                parsed_angles.append(float(str(raw).strip()))
            except (TypeError, ValueError):
                continue
        angles = tuple(parsed_angles or (-4.0, -2.0, 2.0, 4.0))
    variants: list[Image.Image] = []
    for angle in angles:
        rotated = img.rotate(angle, expand=True, resample=Image.BICUBIC, fillcolor=255)
        variants.append(_copy_with_label(rotated, f"{label_prefix}_rot{angle:+.0f}"))
    return variants


def _iter_ocr_variants(img: Image.Image) -> list[Image.Image]:
    """Generate generic OCR-friendly variants without document-specific rules."""
    config = _ocr_runtime_config()
    base = img.convert("L") if img.mode != "L" else img.copy()
    variants = [_copy_with_label(base, "base")]
    variants.extend(_iter_small_rotations(base, label_prefix="base"))

    red_channel = _channel_grayscale(img, 0)
    variants.append(_copy_with_label(red_channel, "red_channel"))
    red_autocontrast = ImageOps.autocontrast(red_channel)
    variants.append(_copy_with_label(red_autocontrast, "red_autocontrast"))
    red_threshold = red_autocontrast.point(
        lambda px: 255 if px > int(config["threshold_low_value"]) else 0
    )
    variants.append(_copy_with_label(red_threshold, "red_threshold_low"))

    perspective_img, perspective_changed = _rectify_document_perspective(img)
    if perspective_changed:
        perspective_gray = perspective_img.convert("L")
        perspective_autocontrast = ImageOps.autocontrast(perspective_gray)
        variants.append(_copy_with_label(perspective_autocontrast, "perspective"))
        variants.extend(_iter_small_rotations(perspective_autocontrast, label_prefix="perspective"))
        perspective_threshold = perspective_autocontrast.point(
            lambda px: 255 if px > int(config["perspective_threshold_value"]) else 0
        )
        variants.append(_copy_with_label(perspective_threshold, "perspective_threshold"))

    autocontrast = ImageOps.autocontrast(base)
    variants.append(_copy_with_label(autocontrast, "autocontrast"))
    variants.extend(_iter_small_rotations(autocontrast, label_prefix="autocontrast"))

    sharpened = autocontrast.filter(
        ImageFilter.MedianFilter(size=int(config["median_filter_size"]))
    )
    sharpened = sharpened.filter(ImageFilter.SHARPEN)
    variants.append(_copy_with_label(sharpened, "median_sharpen"))

    threshold = autocontrast.point(lambda px: 255 if px > int(config["threshold_value"]) else 0)
    variants.append(_copy_with_label(threshold, "threshold"))

    strong_threshold = autocontrast.point(
        lambda px: 255 if px > int(config["threshold_low_value"]) else 0
    )
    variants.append(_copy_with_label(strong_threshold, "threshold_low"))

    trimmed = _trim_document_edges(autocontrast)
    if trimmed.size != autocontrast.size:
        variants.append(_copy_with_label(trimmed, "trimmed"))
        variants.extend(_iter_small_rotations(trimmed, label_prefix="trimmed"))
        variants.append(_copy_with_label(trimmed.rotate(90, expand=True), "trimmed_rot90"))
        variants.append(_copy_with_label(trimmed.rotate(270, expand=True), "trimmed_rot270"))

    return variants


def _variant_label(img: Image.Image) -> str:
    return str(img.info.get("ocr_variant") or "").strip()


def _partition_ocr_variants(
    variants: list[Image.Image],
    primary_variant_labels: list[str],
) -> tuple[list[Image.Image], list[Image.Image]]:
    variants_by_label = {_variant_label(variant): variant for variant in variants}
    primary: list[Image.Image] = []
    selected_labels: set[str] = set()

    for label in primary_variant_labels:
        variant = variants_by_label.get(label)
        if variant is None or label in selected_labels:
            continue
        primary.append(variant)
        selected_labels.add(label)

    rescue = [variant for variant in variants if _variant_label(variant) not in selected_labels]
    return primary, rescue


def _pick_easyocr_variant(
    variants: list[Image.Image],
    preferred_label: str,
) -> Image.Image | None:
    preferred = preferred_label.strip()
    if preferred:
        for variant in variants:
            if _variant_label(variant) == preferred:
                return variant
    return variants[0] if variants else None


def _tesseract_configs(psm_modes: list[str] | tuple[str, ...] | None = None) -> tuple[str, ...]:
    config = _ocr_runtime_config()
    raw_modes = psm_modes if psm_modes is not None else config["primary_psm_modes"]
    parsed: list[str] = []
    for value in raw_modes:
        token = str(value or "").strip()
        if not token:
            continue
        parsed.append(token if token.startswith("--") else f"--psm {token}")
    return tuple(parsed or ("--psm 6",))


def _run_tesseract(
    img: Image.Image,
    *,
    psm_modes: list[str] | tuple[str, ...] | None = None,
) -> str:
    import pytesseract

    runtime_config = _ocr_runtime_config()
    languages = [str(lang).strip() for lang in runtime_config.get("tesseract_languages") or []]
    lang = "+".join(lang for lang in languages if lang) or "spa+eng"
    candidates: list[str] = []
    for config in _tesseract_configs(psm_modes):
        text = str(pytesseract.image_to_string(img, lang=lang, config=config) or "").strip()
        if not text:
            continue
        candidates.append(text)
        if not _is_weak_ocr_text(text):
            return text
    if not candidates:
        return ""
    return _best_text_candidate(candidates, ocr_runtime=_ocr_runtime_config())


def _run_easyocr(img: Image.Image) -> str:
    import easyocr
    import numpy as np

    config = _ocr_runtime_config()
    languages = tuple(
        str(lang).strip() for lang in config["easyocr_languages"] if str(lang).strip()
    )
    gpu_enabled = bool(config.get("easyocr_gpu", False))
    reader_key = (languages or ("es", "en"), gpu_enabled)
    reader = _EASYOCR_READERS.get(reader_key)
    if reader is None:
        with _EASYOCR_READER_LOCK:
            reader = _EASYOCR_READERS.get(reader_key)
            if reader is None:
                reader = easyocr.Reader(list(reader_key[0]), gpu=reader_key[1])
                _EASYOCR_READERS[reader_key] = reader
    results = reader.readtext(np.array(img))
    return "\n".join([r[1] for r in results]).strip()


def _pdf_render_dpi_candidates(base_dpi: int) -> tuple[int, ...]:
    """Return a small set of render DPIs for weak PDF pages."""
    boosted = max(int(round(base_dpi * 1.4)), base_dpi + 120)
    boosted = min(boosted, max(base_dpi, 450))
    candidates = []
    for dpi in (base_dpi, boosted):
        dpi = max(72, int(dpi))
        if dpi not in candidates:
            candidates.append(dpi)
    return tuple(candidates)


def _best_text_candidate(candidates: list[str], *, ocr_runtime: dict[str, Any]) -> str:
    """Pick the OCR candidate with the best text-quality score."""
    best_text = ""
    best_key = (-1.0, -1.0, -1.0)
    for candidate in candidates:
        text = str(candidate or "").strip()
        if not text:
            continue
        quality = _estimate_text_quality(text, ocr_runtime=ocr_runtime)
        key = (
            float(quality.get("score") or 0.0),
            float(quality.get("words") or 0.0),
            float(quality.get("chars") or 0.0),
        )
        if key > best_key:
            best_key = key
            best_text = text
    return best_text


async def extract_text_from_file(
    file_bytes: bytes,
    filename: str,
    *,
    bypass_cache: bool = False,
) -> dict[str, Any]:
    """Extrae texto de cualquier archivo soportado.
    Returns: {"text": str, "pages": int, "structured_data": list[dict] | None, "format": str}
    """
    ext = Path(filename).suffix.lower()
    if bypass_cache:
        logger.info("OCR cache bypassed for %s", Path(filename).name)
    else:
        cached = _load_cached_extraction(file_bytes)
        if cached is not None:
            logger.info("OCR cache hit for %s", Path(filename).name)
            result = dict(cached)
            result["_cache_hit"] = True
            result["_cache_bypassed"] = False
            return result

    if ext == ".pdf":
        extraction = await _extract_pdf(file_bytes)
    elif ext in IMAGE_EXTENSIONS:
        extraction = await _extract_image(file_bytes)
    elif ext in (".xlsx", ".xls"):
        try:
            extraction = _extract_excel(file_bytes, ext=ext)
        except Exception as exc:
            logger.warning("Excel parse failed (%s): %s", ext, exc)
            extraction = {
                "text": "",
                "pages": 1,
                "structured_data": None,
                "format": "EXCEL_ERROR",
                "error": str(exc),
            }
    elif ext == ".csv":
        extraction = _extract_csv(file_bytes)
    elif ext == ".json":
        extraction = _extract_json(file_bytes)
    elif ext == ".xml":
        try:
            extraction = _extract_xml(file_bytes)
        except Exception as exc:
            logger.warning("XML parse failed: %s", exc)
            preview = file_bytes[:4000].decode("utf-8", errors="ignore")
            extraction = {
                "text": preview,
                "pages": 1,
                "structured_data": None,
                "format": "XML_PARSE_ERROR",
                "error": str(exc),
            }
    elif ext == ".txt":
        extraction = _extract_txt(file_bytes)
    elif ext == ".zip":
        extraction = _extract_zip_summary(file_bytes, filename)
    else:
        raise ValueError(f"Formato no soportado: {ext}")
    extraction = _rehydrate_virtual_sheet_context(extraction)
    extraction["_cache_hit"] = False
    extraction["_cache_bypassed"] = bool(bypass_cache)
    if not bypass_cache:
        _store_cached_extraction(file_bytes, extraction)
    return extraction


def iter_zip_entries(
    file_bytes: bytes,
    max_files: int = 20,
    max_size_bytes: int = 8 * 1024 * 1024,
    db: Any | None = None,
) -> Iterable[tuple[str, bytes]]:
    """Itera ficheros válidos dentro de un ZIP.

    - Ignora directorios y ficheros vacíos.
    - Limita número de entradas y tamaño por archivo para evitar OOM.
    - Solo devuelve extensiones soportadas.
    """
    supported_extensions, _, _ = _get_file_support_sets(db)
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        count = 0
        for info in zf.infolist():
            if info.is_dir():
                continue
            if info.file_size <= 0 or info.file_size > max_size_bytes:
                logger.warning(
                    "Zip entry %s skipped (size %s bytes)", info.filename, info.file_size
                )
                continue
            ext = Path(info.filename).suffix.lower()
            if ext not in supported_extensions:
                logger.warning("Zip entry %s skipped (ext %s no soportada)", info.filename, ext)
                continue
            with zf.open(info) as fp:
                yield info.filename, fp.read()
                count += 1
                if count >= max_files:
                    logger.warning("Zip truncado a %s ficheros", max_files)
                    break


def _extract_zip_summary(file_bytes: bytes, zip_name: str) -> dict[str, Any]:
    """Devuelve un resumen de contenido de un ZIP (no reemplaza el fan-out por archivo).

    Se usa solo para mostrar preview rápida si se sube el ZIP como un único documento.
    El fan-out real lo maneja router/run antes de llamar a este extractor.
    """
    summaries = []
    for inner_name, inner_bytes in iter_zip_entries(file_bytes):
        summaries.append({"filename": inner_name, "size": len(inner_bytes)})
    text = "\n".join(f"{s['filename']} ({s['size']} bytes)" for s in summaries)
    return {"text": text, "pages": 1, "structured_data": summaries, "format": "ZIP"}


async def _extract_pdf(file_bytes: bytes) -> dict[str, Any]:
    """PDF: intenta texto embebido con PyMuPDF, si no hay, usa OCR."""
    config = _ocr_runtime_config()
    ocr_runtime = load_ocr_runtime_config(None)
    pdf_render_dpi = int(config["pdf_render_dpi"])
    try:
        import fitz
    except ImportError:
        raise RuntimeError("PyMuPDF (fitz) no disponible")

    doc = fitz.open(stream=file_bytes, filetype="pdf")
    text_parts = []
    page_texts: list[str] = []
    used_ocr = False
    vision_image_bytes: bytes | None = None
    _page_quality_scores: list[float] = []

    try:
        selected_page_texts: list[str] = []
        for page in doc:
            page_text = page.get_text("text")
            page_texts.append(page_text)
            page_text_clean = str(page_text or "").strip()
            if page_text_clean:
                page_quality = _estimate_text_quality(page_text_clean, ocr_runtime=ocr_runtime)
                _page_quality_scores.append(float(page_quality.get("score") or 0.0))
                is_sufficient = (
                    page_quality["score"] >= float(ocr_runtime.get("ocr_min_quality") or 0.45)
                    and page_quality["words"]
                    >= int(ocr_runtime.get("ocr_min_words_for_vision") or 18)
                )
                if is_sufficient:
                    text_parts.append(page_text_clean)
                    selected_page_texts.append(page_text_clean)
                    continue

            used_ocr = True
            ocr_candidates: list[str] = [page_text_clean] if page_text_clean else []
            best_page_text: str = page_text_clean
            for dpi in _pdf_render_dpi_candidates(pdf_render_dpi):
                pix = page.get_pixmap(dpi=dpi)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                if vision_image_bytes is None:
                    vision_image_bytes = _image_to_jpeg_bytes(img)
                candidate = _ocr_image(img)
                if candidate:
                    ocr_candidates.append(candidate)
                best_page_text = _best_text_candidate(ocr_candidates, ocr_runtime=ocr_runtime)
                if candidate and not _is_weak_ocr_text(candidate):
                    break
            if best_page_text:
                text_parts.append(best_page_text)
                selected_page_texts.append(best_page_text)
            else:
                selected_page_texts.append(page_text_clean)

        pages = len(doc)
    finally:
        doc.close()

    _ocr_quality_score: float | None = (
        sum(_page_quality_scores) / len(_page_quality_scores) if _page_quality_scores else None
    )

    if any(str(text or "").strip() for text in selected_page_texts):
        return {
            "text": "\n".join(selected_page_texts),
            "pages": pages,
            "structured_data": None,
            "format": "PDF_OCR" if used_ocr else "PDF",
            "page_texts": selected_page_texts,
            "vision_image_bytes": vision_image_bytes,
            "ocr_quality_score": _ocr_quality_score,
        }

    # Otherwise, convert pages to images and OCR
    # Strategy 1: Use PyMuPDF native rendering (no Poppler needed)
    doc2 = fitz.open(stream=file_bytes, filetype="pdf")
    vision_image_bytes = None
    try:
        ocr_texts = []
        for page in doc2:
            page_candidates: list[str] = []
            for dpi in _pdf_render_dpi_candidates(pdf_render_dpi):
                pix = page.get_pixmap(dpi=dpi)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                if vision_image_bytes is None:
                    vision_image_bytes = _image_to_jpeg_bytes(img)
                candidate = _ocr_image(img)
                if candidate:
                    page_candidates.append(candidate)
                    if not _is_weak_ocr_text(candidate):
                        break
            ocr_texts.append(_best_text_candidate(page_candidates, ocr_runtime=ocr_runtime))
        combined = "\n\n".join(t for t in ocr_texts if t)
        if combined.strip():
            return {
                "text": combined,
                "pages": len(doc2),
                "structured_data": None,
                "format": "PDF_OCR",
                "vision_image_bytes": vision_image_bytes,
                "page_texts": ocr_texts,
                "ocr_quality_score": 0.0,
            }
    except Exception as exc:
        logger.warning("PyMuPDF OCR fallback failed: %s", exc)
    finally:
        doc2.close()

    # Strategy 2: pdf2image (requires Poppler)
    try:
        from pdf2image import convert_from_bytes

        images = convert_from_bytes(file_bytes, dpi=pdf_render_dpi)
        ocr_texts = []
        vision_image_bytes = _image_to_jpeg_bytes(images[0]) if images else None
        for img in images:
            ocr_texts.append(_ocr_image(img))
        return {
            "text": "\n\n".join(ocr_texts),
            "pages": len(images),
            "structured_data": None,
            "format": "PDF_OCR",
            "vision_image_bytes": vision_image_bytes,
            "page_texts": ocr_texts,
            "ocr_quality_score": 0.0,
        }
    except Exception as exc:
        logger.warning("pdf2image OCR fallback failed: %s", exc)
        return {
            "text": "\n".join(text_parts) if text_parts else "",
            "pages": pages,
            "structured_data": None,
            "format": "PDF",
        }


async def _extract_image(file_bytes: bytes) -> dict[str, Any]:
    """Image: OCR with Tesseract with preprocessing for better results."""
    from PIL import ImageEnhance

    config = _ocr_runtime_config()
    img = Image.open(io.BytesIO(file_bytes))

    # Resize if too small — Tesseract works best at ~300 DPI equivalent
    min_width = int(config["min_width"])
    if img.width < min_width:
        scale = min_width / img.width
        img = img.resize((int(img.width * scale), int(img.height * scale)), Image.LANCZOS)

    # Keep color information when available; some handwritten notes OCR better on channels.
    img = ImageEnhance.Contrast(img).enhance(float(config["image_contrast"]))
    img = ImageEnhance.Sharpness(img).enhance(float(config["image_sharpness"]))
    img = img.filter(ImageFilter.SHARPEN)

    text = _ocr_image(img)
    vision_image_bytes = _image_to_jpeg_bytes(img)
    return {
        "text": text,
        "pages": 1,
        "structured_data": None,
        "format": "IMAGE_OCR",
        "vision_image_bytes": vision_image_bytes,
    }


def _ocr_image(img: Image.Image) -> str:
    """Run Tesseract OCR on a PIL Image."""
    candidates: list[str] = []
    config = _ocr_runtime_config()
    variants = _iter_ocr_variants(img)
    primary_variants, rescue_variants = _partition_ocr_variants(
        variants,
        list(config["primary_variant_labels"]),
    )
    try:
        for variant in itertools.chain(primary_variants, rescue_variants):
            psm_modes = (
                config["primary_psm_modes"]
                if variant in primary_variants
                else config["rescue_psm_modes"]
            )
            cleaned = _run_tesseract(variant, psm_modes=psm_modes)
            if cleaned:
                candidates.append(cleaned)
            if cleaned and not _is_weak_ocr_text(cleaned):
                logger.info(
                    "Tesseract OCR accepted variant=%s words=%s chars=%s",
                    _variant_label(variant) or "unknown",
                    *_ocr_text_score(cleaned),
                )
                return cleaned
        logger.info("Tesseract OCR weak output detected across variants; skipping non-native fallback")
    except Exception as exc:
        logger.warning("Tesseract OCR failed: %s", exc)

    if not candidates:
        return ""

    return _best_text_candidate(candidates, ocr_runtime=config)


def _normalize_header(h: Any, idx: int) -> str:
    return str(h or f"col_{idx}").strip().lower().replace(" ", "_")


def _detect_col_type(values: list[Any]) -> str:
    num = sum(isinstance(v, (int, float)) for v in values if v is not None)
    dates = sum(isinstance(v, (datetime.date, datetime.datetime)) for v in values if v is not None)
    strings = sum(isinstance(v, str) for v in values if v is not None)
    counts = {"number": num, "date": dates, "string": strings}
    return max(counts, key=counts.get)


def _iter_xls_rows(file_bytes: bytes):
    try:
        import xlrd
    except ImportError:
        raise RuntimeError("Falta dependencia xlrd para archivos .xls (pip install xlrd)")
    book = xlrd.open_workbook(file_contents=file_bytes)
    for sheet in book.sheets():
        yield sheet.name, (sheet.nrows, sheet.ncols), (
            [sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)
        )


def _iter_xlsx_rows(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            yield ws.title, (ws.max_row or 0, ws.max_column or 0), ws.iter_rows(values_only=True)
    finally:
        wb.close()


def _score_header_row(row: tuple | list) -> float:
    """Score a row as a potential header row. Higher = more likely to be a header.

    Good headers: many non-empty strings, no numbers, short label-like text (ALL-CAPS / Title).
    Bad headers: lots of numbers, single text value, very long strings (prose, not labels).
    Minimum 2 unique text values required — single-value rows (titles, totals) are excluded.
    """
    if not row:
        return 0.0
    values = [v for v in row if v is not None]
    if not values:
        return 0.0
    text_vals = [str(v).strip() for v in values if isinstance(v, str) and str(v).strip()]
    numeric_count = sum(
        1 for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)
    )
    if not text_vals:
        return 0.0
    unique_text = len({t.upper() for t in text_vals})
    # Require at least 2 unique text values — single-value rows (titles/totals) cannot be headers
    if unique_text < 2:
        return 0.0
    # Prefer all-caps or title-case labels (typical for column headers)
    caps_count = sum(1 for t in text_vals if t.isupper() or t.istitle())
    # Prefer short labels (column names are usually < 40 chars); long prose is likely data
    avg_len = sum(len(t) for t in text_vals) / len(text_vals)
    length_bonus = 1.0 if avg_len <= 25 else 0.0
    # Prefer rows where most cells are non-null (dense rows = likely headers)
    fill_ratio = len(values) / max(len(row), 1)
    # Penalize rows with numbers (data rows have lots of numbers; header rows rarely do)
    score = (
        (unique_text * 2.0)
        + (caps_count * 0.5)
        + (length_bonus * 1.0)
        + (fill_ratio * 1.5)
        - (numeric_count * 3.0)
    )
    return score


def _find_header_row(initial_rows: list[tuple | list]) -> tuple[int, list]:
    """Scan rows and return (best_index, best_row) most likely to be the header.

    Prefers rows closer to the start when scores are equal, so a title in row 0
    does not displace a proper header in row 1.
    """
    if not initial_rows:
        return 0, []
    scores = [_score_header_row(r) for r in initial_rows]
    best_score = max(scores)
    if best_score <= 0.0:
        # No valid header found — return first non-empty row
        for i, r in enumerate(initial_rows):
            if any(v is not None for v in r):
                return i, list(r)
        return 0, list(initial_rows[0])
    # Among rows with the best score, prefer the earliest (smallest index)
    best_idx = next(i for i, s in enumerate(scores) if s == best_score)
    return best_idx, list(initial_rows[best_idx])


def _extract_kv_pairs(rows_before_header: list[tuple | list]) -> dict[str, Any]:
    """Extrae pares clave/valor en la sección superior de la hoja (antes del header).

    Heurística simple: toma cada celda de texto como posible clave y asigna
    el primer valor no vacío hacia la derecha en la misma fila.
    """
    kv: dict[str, Any] = {}
    for row in rows_before_header:
        if not row:
            continue
        # Requiere al menos dos valores no vacíos para considerar la fila útil
        non_null = [v for v in row if v not in (None, "")]
        if len(non_null) < 2:
            continue
        idx = 0
        row_list = list(row)
        while idx < len(row_list):
            label = row_list[idx]
            if not (isinstance(label, str) and label.strip() and len(label.strip()) <= 60):
                idx += 1
                continue
            label_clean = label.strip()
            value = None
            next_idx = idx + 1
            while next_idx < len(row_list):
                v = row_list[next_idx]
                if v is None or (isinstance(v, str) and not v.strip()):
                    next_idx += 1
                    continue
                value = v
                break
            if value is not None:
                key_norm = _normalize_header(label_clean, idx)
                if key_norm not in kv:
                    kv[key_norm] = value
                idx = next_idx + 1
            else:
                idx += 1
    return kv


def _extract_excel(file_bytes: bytes, ext: str = ".xlsx") -> dict[str, Any]:
    """Stream Excel safely, build fingerprint and a small preview (no OOM).

    No hay límite de tamaño de archivo aquí: openpyxl en read_only ignora
    imágenes embebidas y solo lee celdas. La protección real son los límites
    de filas (MAX_SCAN_ROWS_PER_SHEET) que acotan el uso de memoria
    independientemente del peso del fichero.
    """
    config = _ocr_runtime_config()
    max_header_scan = int(config["excel_max_header_scan_rows"])
    max_preview_rows_per_sheet = int(config["excel_max_preview_rows_per_sheet"])
    max_scan_rows_per_sheet = max_preview_rows_per_sheet * int(config["excel_scan_rows_multiplier"])
    max_text_chars = int(config["excel_max_text_chars"])

    row_iters = _iter_xls_rows(file_bytes) if ext == ".xls" else _iter_xlsx_rows(file_bytes)

    all_preview_rows: list[dict[str, Any]] = []
    text_lines: list[str] = []
    sheet_profiles: dict[str, Any] = {}
    sheet_metadata: dict[str, dict[str, Any]] = {}
    sheet_used = None
    best_score = -1

    for sheet_name, (nrows, ncols), rows_iter in row_iters:
        rows_iter = iter(rows_iter)

        # Scan first N rows to find the most likely header row
        initial_rows: list[tuple | list] = []
        for _ in range(max_header_scan):
            try:
                initial_rows.append(next(rows_iter))
            except StopIteration:
                break

        if not initial_rows:
            continue

        header_idx, first_row = _find_header_row(initial_rows)
        # Capturar metadatos antes del header (títulos, pares clave/valor)
        kv_pairs = _extract_kv_pairs(initial_rows[:header_idx])

        # Rows AFTER the header row (from the scanned batch) become the first data rows
        data_prefix = [initial_rows[i] for i in range(header_idx + 1, len(initial_rows))]
        rows_iter = itertools.chain(data_prefix, rows_iter)

        # Build headers from the detected header row, using the max column count
        # seen in ALL scanned rows (so data columns not represented in headers get col_N names)
        max_cols = max((len(r) for r in initial_rows), default=len(first_row))
        if len(first_row) < max_cols:
            first_row = list(first_row) + [None] * (max_cols - len(first_row))

        headers = [_normalize_header(h, i) for i, h in enumerate(first_row)]
        header_display = [str(h or f"col_{i}") for i, h in enumerate(first_row)]
        sample_values_by_col: dict[str, list[Any]] = {h: [] for h in headers}

        preview_rows_sheet: list[dict[str, Any]] = []
        total_rows_counted = 0

        for row in rows_iter:
            total_rows_counted += 1
            if all(v is None for v in row):
                continue
            # Si la fila tiene más columnas que el header, crear col_N para las extras
            row_list = list(row)
            if len(row_list) > len(headers):
                for extra_i in range(len(headers), len(row_list)):
                    headers.append(f"col_{extra_i}")
                    header_display.append(f"col_{extra_i}")
                    sample_values_by_col[f"col_{extra_i}"] = []
            row_dict = {headers[j]: row_list[j] for j in range(min(len(headers), len(row_list)))}
            # Normaliza fechas numéricas en .xls si parecen fechas Excel
            if ext == ".xls":
                for k, v in list(row_dict.items()):
                    if isinstance(v, (int, float)):
                        try:
                            import xlrd

                            dt = xlrd.xldate_as_datetime(v, 0)
                            row_dict[k] = dt
                        except Exception:
                            pass
            row_dict["_sheet"] = sheet_name

            if len(preview_rows_sheet) < max_preview_rows_per_sheet:
                preview_rows_sheet.append(row_dict)
                for h in headers:
                    sample_values_by_col[h].append(row_dict.get(h))

            if total_rows_counted >= max_scan_rows_per_sheet:
                break

        text_lines.append(f"[{sheet_name}] " + "\t".join(header_display))
        for row in preview_rows_sheet[:30]:
            text_lines.append(
                f"[{sheet_name}] " + "\t".join(str(row.get(h, "") or "") for h in headers)
            )

        col_profiles = {}
        for h, vals in sample_values_by_col.items():
            if not vals:
                continue
            col_profiles[h] = {"type": _detect_col_type(vals)}

        score = len(preview_rows_sheet)
        headers_lower = [h.lower() for h in headers]
        if any("fecha" in h or "date" in h for h in headers_lower):
            score += 50
        if any("monto" in h or "total" in h or "importe" in h for h in headers_lower):
            score += 20
        if score > best_score:
            best_score = score
            sheet_used = sheet_name

        all_preview_rows.extend(preview_rows_sheet)
        sheet_profiles[sheet_name] = {
            "rows_previewed": len(preview_rows_sheet),
            "rows_counted": total_rows_counted,
            "headers": header_display,
            "headers_norm": headers,
            "column_profiles": col_profiles,
            "kv_pairs": kv_pairs,
        }
        sheet_metadata[sheet_name] = kv_pairs

    text = "\n".join(text_lines)
    if len(text) > max_text_chars:
        text = text[:max_text_chars]

    return {
        "text": text,
        "pages": 1,
        "structured_data": all_preview_rows,
        "format": "EXCEL",
        "sheet_used": sheet_used,
        "sheet_profiles": sheet_profiles,
        "sheet_metadata": sheet_metadata,
    }


def _build_virtual_sheet_context(
    structured_data: Any,
    *,
    sheet_name: str,
    metadata: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], dict[str, Any], str | None]:
    rows: list[dict[str, Any]] = []
    if isinstance(structured_data, dict):
        rows = [{str(key): value for key, value in structured_data.items()}]
    elif isinstance(structured_data, list):
        rows = [
            {str(key): value for key, value in row.items()}
            for row in structured_data
            if isinstance(row, dict)
        ]

    if not rows:
        return {}, {}, None

    headers_display: list[str] = []
    seen_headers: set[str] = set()
    for row in rows:
        for raw_key in row.keys():
            key = str(raw_key).strip()
            if not key or key.startswith("_") or key in seen_headers:
                continue
            seen_headers.add(key)
            headers_display.append(key)

    if not headers_display:
        return {}, {}, None

    headers_norm = [_normalize_header(header, index) for index, header in enumerate(headers_display)]
    sample_values_by_col: dict[str, list[Any]] = {header: [] for header in headers_norm}
    preview_rows = min(len(rows), 50)
    for row in rows[:preview_rows]:
        for raw_header, norm_header in zip(headers_display, headers_norm):
            value = row.get(raw_header)
            if value is not None and value != "":
                sample_values_by_col[norm_header].append(value)

    column_profiles = {
        header: {"type": _detect_col_type(values) if values else "string"}
        for header, values in sample_values_by_col.items()
    }
    kv_pairs = {
        str(key): value
        for key, value in (metadata or {}).items()
        if value not in (None, "", [], {}) and not str(key).startswith("_")
    }

    return (
        {
            sheet_name: {
                "rows_previewed": preview_rows,
                "rows_counted": len(rows),
                "headers": headers_display,
                "headers_norm": headers_norm,
                "column_profiles": column_profiles,
                "kv_pairs": kv_pairs,
            }
        },
        {sheet_name: kv_pairs},
        sheet_name,
    )


def _rehydrate_virtual_sheet_context(extraction: dict[str, Any]) -> dict[str, Any]:
    fmt = str(extraction.get("format") or "").upper()
    if fmt not in {"CSV", "JSON", "XML", "XML_UBL", "XML_FACTURAE"}:
        return extraction

    if extraction.get("sheet_profiles"):
        return extraction

    structured_data = extraction.get("structured_data")
    if not structured_data:
        return extraction

    sheet_name = "XML" if fmt.startswith("XML") else fmt
    metadata: dict[str, Any] | None = None
    raw_metadata = extraction.get("sheet_metadata")
    if isinstance(raw_metadata, dict):
        for value in raw_metadata.values():
            if isinstance(value, dict) and value:
                metadata = value
                break
    if metadata is None and isinstance(structured_data, list) and structured_data:
        first_row = structured_data[0]
        if isinstance(first_row, dict):
            metadata = {
                str(key): value
                for key, value in first_row.items()
                if not str(key).startswith("_")
            }
    elif metadata is None and isinstance(structured_data, dict):
        metadata = {
            str(key): value
            for key, value in structured_data.items()
            if not str(key).startswith("_")
        }

    sheet_profiles, sheet_metadata, sheet_used = _build_virtual_sheet_context(
        structured_data,
        sheet_name=sheet_name,
        metadata=metadata,
    )
    if sheet_profiles:
        extraction["sheet_profiles"] = sheet_profiles
    if sheet_metadata and not extraction.get("sheet_metadata"):
        extraction["sheet_metadata"] = sheet_metadata
    if sheet_used and not extraction.get("sheet_used"):
        extraction["sheet_used"] = sheet_used
    return extraction


_CSV_SUMMARY_TITLE_KEYS = {
    "resumen",
    "summary",
    "totales",
    "totals",
    "total",
}
_CSV_TOTAL_KEY_ALIASES = {
    "total_amount",
    "total_sales",
    "sales_total",
    "total_ventas",
    "ventas_total",
    "grand_total",
    "importe_total",
    "monto_total",
    "total",
    "amount",
}


def _csv_cell_has_value(value: Any) -> bool:
    if value is None:
        return False
    if pd is not None:
        try:
            if pd.isna(value):
                return False
        except Exception:
            pass
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _csv_key_norm(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    normalized = unicodedata.normalize("NFD", raw)
    normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    return normalized.strip("_")


def _csv_parse_summary_and_promote(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    main_rows: list[dict[str, Any]] = []
    summary_meta: dict[str, Any] = {}
    summary_mode = False

    for row in rows:
        clean_row = {
            str(key): value
            for key, value in row.items()
            if _csv_cell_has_value(value)
        }
        if not clean_row:
            continue

        values = list(clean_row.values())
        if not values:
            continue

        if not summary_mode:
            if len(values) == 1 and _csv_key_norm(values[0]) in _CSV_SUMMARY_TITLE_KEYS:
                summary_mode = True
                continue
            main_rows.append(clean_row)
            continue

        if len(values) < 2:
            continue

        key = _csv_key_norm(values[0])
        if not key:
            continue
        summary_meta.setdefault(key, values[1])
        if key in _CSV_TOTAL_KEY_ALIASES:
            summary_meta.setdefault("total_amount", values[1])

    if main_rows and "issue_date" not in summary_meta:
        for row in main_rows:
            for key, value in row.items():
                key_norm = _csv_key_norm(key)
                if "fecha" not in key_norm and not key_norm.endswith("_date") and key_norm != "date":
                    continue
                if _csv_cell_has_value(value):
                    summary_meta["issue_date"] = value
                    break
            if "issue_date" in summary_meta:
                break

    return main_rows, summary_meta


def _extract_csv(file_bytes: bytes) -> dict[str, Any]:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            text_content = file_bytes.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        text_content = file_bytes.decode("utf-8", errors="replace")

    if pd is not None:
        try:
            frame = pd.read_csv(io.BytesIO(file_bytes))
            frame = frame.dropna(how="all")
            data = frame.where(pd.notna(frame), None).to_dict(orient="records")
            data = [
                {str(key): value for key, value in row.items() if value is not None}
                for row in data
                if any(value is not None for value in row.values())
            ]
            data, summary_meta = _csv_parse_summary_and_promote(data)
            sheet_profiles, sheet_metadata, sheet_used = _build_virtual_sheet_context(
                data,
                sheet_name="CSV",
                metadata=summary_meta or None,
            )
            return {
                "text": text_content[:10000],
                "pages": 1,
                "structured_data": data,
                "format": "CSV",
                "sheet_profiles": sheet_profiles,
                "sheet_metadata": sheet_metadata,
                "sheet_used": sheet_used,
            }
        except Exception as exc:
            logger.debug("pandas CSV parse failed; falling back to stdlib parser: %s", exc)

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text_content[:2048])
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)
    data = [row for row in reader if any(v and v.strip() for v in row.values())]
    data, summary_meta = _csv_parse_summary_and_promote(data)
    sheet_profiles, sheet_metadata, sheet_used = _build_virtual_sheet_context(
        data,
        sheet_name="CSV",
        metadata=summary_meta or None,
    )
    return {
        "text": text_content[:10000],
        "pages": 1,
        "structured_data": data,
        "format": "CSV",
        "sheet_profiles": sheet_profiles,
        "sheet_metadata": sheet_metadata,
        "sheet_used": sheet_used,
    }


def _extract_json(file_bytes: bytes) -> dict[str, Any]:
    text_content = file_bytes.decode("utf-8", errors="replace")
    try:
        payload = json.loads(text_content)
    except json.JSONDecodeError as exc:
        return {
            "text": text_content[:10000],
            "pages": 1,
            "structured_data": None,
            "format": "JSON_PARSE_ERROR",
            "error": str(exc),
        }

    if isinstance(payload, list):
        structured_data = [item for item in payload if isinstance(item, dict)]
    else:
        structured_data = payload if isinstance(payload, dict) else {"value": payload}
    sheet_profiles, sheet_metadata, sheet_used = _build_virtual_sheet_context(
        structured_data,
        sheet_name="JSON",
    )

    return {
        "text": json.dumps(payload, ensure_ascii=False)[:10000],
        "pages": 1,
        "structured_data": structured_data,
        "format": "JSON",
        "sheet_profiles": sheet_profiles,
        "sheet_metadata": sheet_metadata,
        "sheet_used": sheet_used,
    }


def _extract_xml(file_bytes: bytes) -> dict[str, Any]:
    """XML extraction for UBL/Facturae with graceful fallback on malformed XML."""
    raw_xml_text = file_bytes.decode("utf-8", errors="ignore")
    try:
        root = ET.fromstring(file_bytes)
    except Exception as exc:
        root = None
        for closing_tag in ("</Facturae>", "</Invoice>", "</CreditNote>", "</DebitNote>"):
            close_at = raw_xml_text.find(closing_tag)
            if close_at < 0:
                continue
            candidate = raw_xml_text[: close_at + len(closing_tag)]
            try:
                root = ET.fromstring(candidate)
                logger.info(
                    "XML parse recovered by truncating trailing content closing_tag=%s",
                    closing_tag,
                )
                break
            except Exception:
                continue
        if root is None:
            # Malformed XML: degrade to text preview and mark parse error, but do NOT raise
            preview = file_bytes[:4000].decode("utf-8", errors="ignore")
            return {
                "text": preview,
                "pages": 1,
                "structured_data": None,
                "format": "XML_PARSE_ERROR",
                "error": str(exc),
            }
    tag = root.tag.split("}")[-1] if "}" in root.tag else root.tag
    root_ns = root.tag.split("}")[0].strip("{") if "}" in root.tag else ""

    if root_ns == "http://www.facturae.gob.es/formato" or tag.lower() == "facturae":
        ns = {"fe": root_ns or "http://www.facturae.gob.es/formato"}

        def find_text(*paths):
            for path in paths:
                el = root.find(path, ns)
                if el is not None and el.text and el.text.strip():
                    return el.text.strip()
            return None

        def find_text_in(element: ET.Element, *paths: str):
            for path in paths:
                el = element.find(path, ns)
                if el is not None and el.text and el.text.strip():
                    return el.text.strip()
            return None

        invoice_lines = root.findall(".//fe:InvoiceLine", ns)
        line_items: list[dict[str, Any]] = []
        tax_total = 0.0
        for line in invoice_lines:
            item = {
                "descripcion": find_text_in(line, ".//fe:ItemDescription"),
                "cantidad": find_text_in(line, ".//fe:Quantity"),
                "precio_unitario": find_text_in(line, ".//fe:UnitPriceWithoutTax"),
                "total": find_text_in(line, ".//fe:TotalAmountWithoutTax"),
            }
            tax_amount_text = find_text_in(line, ".//fe:TaxesOutputs//fe:TaxAmount/fe:TotalAmount")
            tax_rate_text = find_text_in(line, ".//fe:TaxesOutputs//fe:TaxRate")
            if tax_rate_text:
                item["impuesto_pct"] = tax_rate_text
            if tax_amount_text:
                item["impuesto"] = tax_amount_text
                try:
                    tax_total += float(str(tax_amount_text).replace(",", "."))
                except (TypeError, ValueError):
                    pass
            if any(value not in (None, "", [], {}) for value in item.values()):
                line_items.append(item)

        header = {
            "documento": " ".join(
                part
                for part in [
                    find_text(".//fe:InvoiceNumber"),
                    find_text(".//fe:InvoiceSeriesCode"),
                ]
                if part
            ) or find_text(".//fe:InvoiceNumber"),
            "fecha": find_text(".//fe:IssueDate"),
            "tipo_documento": "FACTURA",
            "proveedor": find_text(".//fe:SellerParty//fe:CorporateName"),
            "comprador": find_text(".//fe:BuyerParty//fe:CorporateName"),
            "subtotal": find_text(".//fe:InvoiceTotals//fe:TotalGrossAmount"),
            "monto": find_text(".//fe:InvoiceTotals//fe:InvoiceTotal")
            or find_text(".//fe:InvoiceTotals//fe:TotalGrossAmount"),
        }
        if tax_total:
            header["impuesto"] = f"{tax_total:.2f}"

        full_text = ET.tostring(root, encoding="unicode", method="text")
        structured_rows = line_items or [{k: v for k, v in header.items() if v not in (None, "")}]
        sheet_profiles, sheet_metadata, sheet_used = _build_virtual_sheet_context(
            structured_rows,
            sheet_name="XML",
            metadata=header,
        )
        return {
            "text": full_text[:10000] if full_text else str(header),
            "pages": 1,
            "structured_data": structured_rows,
            "format": "XML_FACTURAE",
            "sheet_profiles": sheet_profiles,
            "sheet_metadata": sheet_metadata,
            "sheet_used": sheet_used,
        }

    is_credit_note = tag.lower() in ("creditnote", "debitnote")
    ns = _UBL_NS

    def find_text(*paths):
        for p in paths:
            el = root.find(p, ns)
            if el is not None and el.text:
                return el.text.strip()
        return None

    header = {
        "documento": find_text("cbc:ID", ".//cbc:ID"),
        "fecha": find_text("cbc:IssueDate", ".//cbc:IssueDate"),
        "moneda": find_text("cbc:DocumentCurrencyCode"),
        "tipo_documento": "NOTA_CREDITO" if is_credit_note else "FACTURA",
    }

    supplier = root.find(".//cac:AccountingSupplierParty//cac:Party", ns)
    if supplier is not None:
        tax_scheme = supplier.find(".//cac:PartyTaxScheme", ns)
        if tax_scheme is not None:
            el = tax_scheme.find("cbc:CompanyID", ns)
            if el is not None and el.text:
                header["ruc"] = el.text.strip()
        name_el = supplier.find(".//cac:PartyName/cbc:Name", ns)
        if name_el is not None and name_el.text:
            header["proveedor"] = name_el.text.strip()

    monetary = root.find(".//cac:LegalMonetaryTotal", ns)
    if monetary is not None:
        st = monetary.find("cbc:TaxExclusiveAmount", ns)
        if st is not None and st.text:
            header["subtotal"] = st.text.strip()
        pa = monetary.find("cbc:PayableAmount", ns)
        if pa is not None and pa.text:
            header["monto"] = pa.text.strip()

    tax_total = root.find(".//cac:TaxTotal", ns)
    if tax_total is not None:
        ta = tax_total.find("cbc:TaxAmount", ns)
        if ta is not None and ta.text:
            header["igv"] = ta.text.strip()

    if is_credit_note:
        for field in ("monto", "subtotal", "igv"):
            if header.get(field):
                try:
                    val = float(header[field])
                    if val > 0:
                        header[field] = str(-val)
                except (ValueError, TypeError):
                    pass

    full_text = ET.tostring(root, encoding="unicode", method="text")
    sheet_profiles, sheet_metadata, sheet_used = _build_virtual_sheet_context(
        [header],
        sheet_name="XML",
        metadata=header,
    )
    return {
        "text": full_text[:10000] if full_text else str(header),
        "pages": 1,
        "structured_data": [header],
        "format": "XML_UBL",
        "sheet_profiles": sheet_profiles,
        "sheet_metadata": sheet_metadata,
        "sheet_used": sheet_used,
    }


def _extract_txt(file_bytes: bytes) -> dict[str, Any]:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            text_content = file_bytes.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        text_content = file_bytes.decode("utf-8", errors="replace")

    return {"text": text_content, "pages": 1, "structured_data": None, "format": "TXT"}
