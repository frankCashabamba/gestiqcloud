"""Parser for PAN KUSI costing workbook as products (price list).

This workbook is a recipe costing template with one recipe per sheet. For the
"products" import flow, users often want to import the suggested selling prices
as products, not the recipe/ingredients structure.
"""

from __future__ import annotations

import re
import unicodedata
from datetime import datetime
from typing import Any

import openpyxl


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


def _to_number(val: Any) -> float | None:
    if val in (None, ""):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s_norm = re.sub(r"[^0-9,.-]", "", s)
    if "," in s_norm and "." in s_norm:
        if s_norm.rfind(",") > s_norm.rfind("."):
            s_norm = s_norm.replace(".", "").replace(",", ".")
        else:
            s_norm = s_norm.replace(",", "")
    else:
        s_norm = s_norm.replace(",", ".")
    try:
        return float(s_norm)
    except Exception:
        return None


def _find_label_value(ws, label: str) -> Any:
    """Find a cell containing label and return the value cell to the right.

    The PAN KUSI template usually stores the numeric value 2-3 columns to the right
    of the label. We try a small window on the same row.
    """
    want = _norm(label)
    max_row = min(getattr(ws, "max_row", 200) or 200, 200)
    max_col = min(getattr(ws, "max_column", 50) or 50, 50)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str) or not v.strip():
                continue
            if want in _norm(v):
                # Prefer value at +3 columns (matches the template), then scan right.
                for dc in (3, 1, 2, 4, 5, 6):
                    vv = ws.cell(row=r, column=c + dc).value
                    if vv not in (None, ""):
                        return vv
                return None
    return None


def _get_recipe_name(ws) -> str | None:
    # Template: "Nombre de la Receta" in A3, actual value in G3.
    v = ws["G3"].value
    if v not in (None, ""):
        return str(v).strip() or None
    # Fallback: search top-left for the label.
    for r in range(1, 15):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c).value
            if isinstance(cell, str) and "NOMBRE DE LA RECETA" in _norm(cell):
                vv = ws.cell(row=r, column=7).value
                if vv not in (None, ""):
                    return str(vv).strip() or None
    return None


def _parse_ingredients(ws) -> list[dict[str, Any]]:
    start = None
    max_row = ws.max_row or 0
    for row in range(1, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and _norm(val) == "INGREDIENTES":
            start = row + 1
            break
    if not start:
        return []

    out: list[dict[str, Any]] = []
    for row in range(start, max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is None or str(name).strip() == "":
            break
        out.append(
            {
                "ingredient": str(name).strip(),
                "quantity": _to_number(ws.cell(row=row, column=6).value),  # F
                "unit": (
                    str(ws.cell(row=row, column=7).value).strip()
                    if ws.cell(row=row, column=7).value not in (None, "")
                    else None
                ),  # G
                "amount": _to_number(ws.cell(row=row, column=19).value),  # S (IMPORTE)
            }
        )
    return out


def _parse_materials(ws) -> list[dict[str, Any]]:
    title_row = None
    max_row = ws.max_row or 0
    for row in range(1, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and "MATERIALES O INSUMOS ADICIONALES" in _norm(val):
            title_row = row
            break
    if not title_row:
        return []

    start = title_row + 2  # skip title + header row
    out: list[dict[str, Any]] = []
    for row in range(start, max_row + 1):
        desc = ws.cell(row=row, column=1).value
        if desc is None or str(desc).strip() == "":
            break
        out.append(
            {
                "description": str(desc).strip(),
                "quantity": _to_number(ws.cell(row=row, column=7).value),  # G
                "purchase_unit": (
                    str(ws.cell(row=row, column=11).value).strip()
                    if ws.cell(row=row, column=11).value not in (None, "")
                    else None
                ),  # K
                "amount": _to_number(ws.cell(row=row, column=19).value),  # S (IMPORTE)
            }
        )
    return out


def parse_xlsx_costing_products(file_path: str) -> dict[str, Any]:
    """Parse costing workbook into a product list (one product per sheet)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    products: list[dict[str, Any]] = []
    errors: list[str] = []

    try:
        for ws in wb.worksheets:
            name = _get_recipe_name(ws)
            if not name:
                continue

            ingredients = _parse_ingredients(ws)
            materials = _parse_materials(ws)

            classification = ws["E4"].value
            portions = _to_number(ws["E7"].value)

            cost_total = _to_number(_find_label_value(ws, "COSTO DE PROD. FINAL"))
            price_total = _to_number(_find_label_value(ws, "PRECIO DE VENTA TOTAL"))
            price_unit = _to_number(_find_label_value(ws, "PRECIO DE VENTA UNITARIO"))
            # Fallback: compute unit price from total/portions.
            if price_unit is None and price_total is not None and portions:
                try:
                    price_unit = float(price_total) / float(portions)
                except Exception:
                    price_unit = None

            # Cost per unit: prefer total/portions; fallback to "Costo Unitario Ingredientes" (E8).
            cost_unit = None
            if cost_total is not None and portions:
                try:
                    cost_unit = float(cost_total) / float(portions)
                except Exception:
                    cost_unit = None
            if cost_unit is None:
                cost_unit = _to_number(ws["E8"].value)

            product: dict[str, Any] = {
                "doc_type": "product",
                "name": str(name).strip(),
                "nombre": str(name).strip(),
                "producto": str(name).strip(),
                "sku": str(name).strip(),  # sanitized later in import pipeline
                "price": float(price_unit or 0.0),
                "precio": float(price_unit or 0.0),
                "category": (
                    str(classification).strip() if classification not in (None, "") else None
                ),
                "categoria": (
                    str(classification).strip() if classification not in (None, "") else None
                ),
                "ingredients": ingredients,
                "materials": materials,
                "source": "xlsx_costing",
                "_metadata": {
                    "parser": "xlsx_costing_products",
                    "sheet": ws.title,
                    "portions": portions,
                    "price_total": price_total,
                    "cost_total": cost_total,
                    "imported_at": datetime.utcnow().isoformat(),
                },
            }
            if cost_unit is not None:
                product.update(
                    {
                        "cost_price": float(cost_unit),
                        "cost": float(cost_unit),
                        "costo": float(cost_unit),
                        "unit_cost": float(cost_unit),
                    }
                )

            products.append({k: v for k, v in product.items() if v is not None})
    except Exception as exc:  # pragma: no cover
        errors.append(str(exc))
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "products": products,
        "rows_parsed": len(products),
        "errors": errors,
        "source_type": "xlsx",
        "parser": "xlsx_costing_products",
        "detected_type": "products",
    }
