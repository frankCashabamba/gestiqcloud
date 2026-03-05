"""Excel parser for expenses/receipts."""

from datetime import datetime
from typing import Any

import openpyxl


def parse_xlsx_expenses(file_path: str, sheet_name: str = None) -> dict[str, Any]:
    """Parse Excel file with expense/receipt data.

    Detects columns: date, description, category, amount, vendor,
    payment_method, reference, tax, currency

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name (default: first sheet or "GASTOS"/"EXPENSES")

    Returns:
        Dict with 'expenses' list and metadata
    """
    expenses = []
    errors = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Try to find standard sheet names
        if sheet_name:
            ws = wb[sheet_name]
        elif "GASTOS" in wb.sheetnames:
            ws = wb["GASTOS"]
        elif "EXPENSES" in wb.sheetnames:
            ws = wb["EXPENSES"]
        elif "RECIBOS" in wb.sheetnames:
            ws = wb["RECIBOS"]
        elif "RECEIPTS" in wb.sheetnames:
            ws = wb["RECEIPTS"]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {
                "expenses": [],
                "rows_processed": 0,
                "rows_parsed": 0,
                "errors": [],
                "source_type": "xlsx",
                "parser": "xlsx_expenses",
            }

        # Parse header
        header = rows[0]
        col_map = _map_columns(header)

        # Parse expense rows
        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or all(cell is None for cell in row):
                continue

            try:
                expense = _parse_expense_row(row, row_idx, col_map)
                if expense:
                    expenses.append(expense)
            except Exception as e:
                errors.append(f"Error parsing row {row_idx}: {str(e)}")

    except Exception as e:
        errors.append(f"Error reading Excel: {str(e)}")

    return {
        "expenses": expenses,
        "rows_processed": len(rows) - 1,  # Exclude header
        "rows_parsed": len(expenses),
        "errors": errors,
        "source_type": "xlsx",
        "parser": "xlsx_expenses",
    }


def _map_columns(header: tuple) -> dict[str, int]:
    """Map column names to indices.

    Args:
        header: Header row tuple

    Returns:
        Dict mapping field name to column index
    """
    col_map = {}

    for idx, col_name in enumerate(header):
        if not col_name:
            continue

        col_lower = str(col_name).lower().strip()

        # Date
        if any(kw in col_lower for kw in ["date", "fecha", "transaction_date"]):
            if "date" not in col_map:
                col_map["date"] = idx

        # Description
        elif any(
            kw in col_lower for kw in ["description", "descripcion", "desc", "concepto", "detail"]
        ):
            if "description" not in col_map:
                col_map["description"] = idx

        # Category
        elif any(kw in col_lower for kw in ["category", "categoria", "tipo", "type"]):
            if "category" not in col_map:
                col_map["category"] = idx

        # Amount
        elif any(kw in col_lower for kw in ["amount", "monto", "valor", "quantity", "total"]):
            if "amount" not in col_map:
                col_map["amount"] = idx

        # Vendor
        elif any(kw in col_lower for kw in ["vendor", "supplier", "from", "source"]):
            if "vendor" not in col_map:
                col_map["vendor"] = idx

        # Payment method
        elif any(kw in col_lower for kw in ["payment", "forma_pago", "payment_method", "metodo"]):
            if "payment_method" not in col_map:
                col_map["payment_method"] = idx

        # Reference
        elif any(kw in col_lower for kw in ["reference", "referencia", "ref", "numero"]):
            if "reference" not in col_map:
                col_map["reference"] = idx

        # Tax
        elif any(kw in col_lower for kw in ["tax", "iva", "impuesto", "tax_amount"]):
            if "tax" not in col_map:
                col_map["tax"] = idx

        # Currency
        elif any(kw in col_lower for kw in ["currency", "moneda", "curreny"]):
            if "currency" not in col_map:
                col_map["currency"] = idx

    return col_map


def _parse_expense_row(row: tuple, row_idx: int, col_map: dict[str, int]) -> dict[str, Any] | None:
    """Extract expense data from row.

    Args:
        row: Data row tuple
        row_idx: Row index
        col_map: Column mapping

    Returns:
        Dict with expense data or None
    """
    # Extract fields
    date_str = row[col_map["date"]] if "date" in col_map and col_map["date"] < len(row) else None
    description = (
        row[col_map["description"]]
        if "description" in col_map and col_map["description"] < len(row)
        else None
    )
    category = (
        row[col_map["category"]]
        if "category" in col_map and col_map["category"] < len(row)
        else None
    )
    amount = (
        row[col_map["amount"]] if "amount" in col_map and col_map["amount"] < len(row) else None
    )
    vendor = (
        row[col_map["vendor"]] if "vendor" in col_map and col_map["vendor"] < len(row) else None
    )
    payment_method = (
        row[col_map["payment_method"]]
        if "payment_method" in col_map and col_map["payment_method"] < len(row)
        else None
    )
    reference = (
        row[col_map["reference"]]
        if "reference" in col_map and col_map["reference"] < len(row)
        else None
    )
    tax = row[col_map["tax"]] if "tax" in col_map and col_map["tax"] < len(row) else None
    currency = (
        row[col_map["currency"]]
        if "currency" in col_map and col_map["currency"] < len(row)
        else None
    )

    # Require at least description and amount
    if not description or not amount:
        return None

    expense = {
        "doc_type": "expense",
        "description": str(description).strip(),
        "amount": _to_float(amount),
        "currency": str(currency).strip() if currency else "USD",
        "source": "xlsx",
        "_metadata": {
            "parser": "xlsx_expenses",
            "row_index": row_idx,
            "imported_at": datetime.utcnow().isoformat(),
        },
    }

    if date_str:
        expense["date"] = _parse_date(date_str)

    if category:
        expense["category"] = str(category).strip()

    if vendor:
        expense["vendor"] = str(vendor).strip()

    if payment_method:
        expense["payment_method"] = str(payment_method).strip()

    if reference:
        expense["reference"] = str(reference).strip()

    if tax:
        expense["tax"] = _to_float(tax)

    return _clean_dict(expense)


def _to_float(val) -> float | None:
    """Convert to float or None."""
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def _parse_date(val) -> str | None:
    """Parse date to ISO format string."""
    if not val:
        return None

    # If already string date-like
    if isinstance(val, str):
        return val

    # If datetime object
    if hasattr(val, "isoformat"):
        return val.isoformat()

    return str(val)


def _clean_dict(d: dict) -> dict:
    """Remove keys with None or empty string values."""
    if not isinstance(d, dict):
        return d
    return {
        k: _clean_dict(v) if isinstance(v, dict) else v
        for k, v in d.items()
        if v is not None and v != "" and (not isinstance(v, dict) or any(_clean_dict(v).values()))
    }
