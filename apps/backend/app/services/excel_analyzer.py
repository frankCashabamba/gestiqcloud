"""
Excel Column Analyzer - Detecta automáticamente columnas y sugiere mapeos
"""

from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def detect_header_row(ws: Worksheet, max_rows: int = 10) -> int:
    """
    Detecta qué fila contiene los encabezados
    Busca la primera fila con mayoría de celdas tipo texto
    """
    # Palabras que indican que NO es un header (instrucciones, notas, etc)
    exclude_keywords = [
        "formato",
        "instrucciones",
        "como",
        "apuntar",
        "rellenar",
        "llenar",
        "ejemplo",
        "nota",
        "aviso",
        "observaciones",
        "completar",
        "guia",
        "compras",
        "ventas",
        "manera",
        "forma",
        "modo",
        "instruccion",
    ]

    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_values = [cell.value for cell in ws[row_idx]]

        # Filtrar None y vacíos
        non_empty = [v for v in row_values if v]
        if len(non_empty) == 0:
            continue

        # Si la primera celda contiene palabras de exclusión, skip
        first_cell = str(row_values[0] or "").lower().strip()
        if any(kw in first_cell for kw in exclude_keywords):
            continue

        # Contar celdas con texto (no None, no números puros)
        text_count = sum(
            1
            for v in row_values
            if v and isinstance(v, str) and v.strip() and not v.strip().isdigit()
        )

        # Si 60% o más son texto Y no son instrucciones, es probable que sea header
        if text_count >= len(row_values) * 0.5 and text_count >= 3:
            return row_idx

    return 1  # Default a primera fila


def extract_headers(ws: Worksheet, header_row: int) -> list[str]:
    """Extrae headers limpiando valores None"""
    headers = []
    for idx, cell in enumerate(ws[header_row]):
        value = cell.value
        if value and str(value).strip():
            headers.append(str(value).strip())
        else:
            # Si hay datos en columnas posteriores, usar placeholder
            # Sino, terminar extracción
            has_data_below = any(
                ws.cell(r, idx + 1).value
                for r in range(header_row + 1, min(header_row + 6, ws.max_row + 1))
            )
            if has_data_below:
                headers.append(f"Columna_{idx + 1}")
            else:
                break  # No hay más columnas con datos
    return headers


def get_sample_rows(
    ws: Worksheet, header_row: int, headers: list[str], limit: int = 5
) -> list[dict[str, Any]]:
    """
    Extrae filas de muestra con sus valores
    """
    sample = []
    start_row = header_row + 1
    end_row = min(start_row + limit, ws.max_row + 1)

    for row_idx in range(start_row, end_row):
        row_data = {}
        for col_idx, header in enumerate(headers):
            cell_value = ws.cell(row_idx, col_idx + 1).value
            row_data[header] = cell_value

        # Solo incluir filas con al menos un valor
        if any(v for v in row_data.values()):
            sample.append(row_data)

    return sample


def suggest_column_mapping(headers: list[str]) -> dict[str, str]:
    """
    Sugerencia básica de mapeo por palabras clave
    Retorna: {"columna_excel": "campo_destino"}
    """
    mapping = {}

    # Keywords por campo destino
    keywords_map = {
        "name": [
            "producto",
            "nombre",
            "descripcion",
            "description",
            "item",
            "articulo",
            "artículo",
            "product",
            "name",
            "denominacion",
        ],
        "precio": [
            "precio",
            "price",
            "pvp",
            "venta",
            "unitario",
            "valor",
            "importe",
            "amount",
            "cost",
            "costo",
        ],
        "cantidad": [
            "cantidad",
            "stock",
            "qty",
            "quantity",
            "existencia",
            "unidades",
            "units",
            "disponible",
        ],
        "categoria": [
            "categoria",
            "categoría",
            "category",
            "grupo",
            "familia",
            "type",
            "tipo",
            "class",
            "clase",
        ],
        "codigo": [
            "codigo",
            "código",
            "code",
            "sku",
            "referencia",
            "ref",
            "reference",
            "barcode",
            "ean",
        ],
        "costo": ["costo", "cost", "compra", "purchase", "adquisicion"],
        "supplier": ["proveedor", "supplier", "vendor", "fabricante"],
        "unidad": ["unidad", "unit", "medida", "uom", "u/m"],
    }

    for header in headers:
        header_lower = header.lower().strip()

        # Buscar match con keywords
        for target_field, keywords in keywords_map.items():
            if any(kw in header_lower for kw in keywords):
                mapping[header] = target_field
                break

        # Si no match, marcar como "ignore" si parece irrelevante
        if header not in mapping:
            irrelevant_keywords = [
                "formato",
                "instrucciones",
                "notas",
                "observaciones",
                "comentario",
                "nota",
                "aviso",
                "sobrante",  # Dato operacional
                "venta",  # Dato operacional (no confundir con precio venta)
                "total",  # Cálculo derivado
                "diario",  # Datos temporales/operacionales
                "unnamed",  # Columnas sin nombre
            ]
            if any(kw in header_lower for kw in irrelevant_keywords):
                mapping[header] = "ignore"

    return mapping


def analyze_excel_file(file_path: str) -> dict[str, Any]:
    """
    Análisis completo de archivo Excel

    Returns:
        {
            "headers": [...],
            "header_row": 1,
            "sample_data": [{...}, ...],
            "suggested_mapping": {"col": "field", ...},
            "total_rows": 100,
            "total_columns": 5
        }
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Detectar fila de headers
    header_row = detect_header_row(ws)

    # Extraer headers
    headers = extract_headers(ws, header_row)

    # Muestra de datos
    sample_data = get_sample_rows(ws, header_row, headers, limit=5)

    # Sugerencias de mapeo
    suggested_mapping = suggest_column_mapping(headers)

    # Estadísticas
    total_rows = ws.max_row - header_row
    total_columns = len(headers)

    wb.close()

    return {
        "headers": headers,
        "header_row": header_row,
        "sample_data": sample_data,
        "suggested_mapping": suggested_mapping,
        "total_rows": total_rows,
        "total_columns": total_columns,
        "analysis_metadata": {
            "detected_header_row": header_row,
            "columns_with_suggestions": len(
                [m for m in suggested_mapping.values() if m != "ignore"]
            ),
            "ignored_columns": len([m for m in suggested_mapping.values() if m == "ignore"]),
        },
    }


def analyze_excel_stream(file_stream) -> dict[str, Any]:
    """
    Analiza un archivo Excel desde un stream (UploadFile)
    """
    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    header_row = detect_header_row(ws)
    headers = extract_headers(ws, header_row)
    sample_data = get_sample_rows(ws, header_row, headers, limit=5)
    suggested_mapping = suggest_column_mapping(headers)

    total_rows = ws.max_row - header_row
    total_columns = len(headers)

    wb.close()

    return {
        "headers": headers,
        "header_row": header_row,
        "sample_data": sample_data,
        "suggested_mapping": suggested_mapping,
        "total_rows": total_rows,
        "total_columns": total_columns,
    }
