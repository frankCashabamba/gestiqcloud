import openpyxl

wb = openpyxl.load_workbook("Stock-02-11-2025.xlsx", data_only=True)

print("=== TODAS LAS HOJAS ===")
for i, sheet_name in enumerate(wb.sheetnames):
    print(f"\n{i+1}. HOJA: '{sheet_name}'")
    ws = wb[sheet_name]

    # Mostrar primeras 2 filas
    print(f"   Fila 1: {[c.value for c in ws[1]][:7]}")
    print(f"   Fila 2: {[c.value for c in ws[2]][:7]}")

    # Buscar headers relevantes
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_values = [c.value for c in ws[row_idx]]
        row_str = str(row_values).upper()
        if "PRODUCTO" in row_str and ("PRECIO" in row_str or "CANTIDAD" in row_str):
            print(f"   ✅ HEADERS DETECTADOS EN FILA {row_idx}: {row_values[:7]}")
            print("   Datos:")
            for data_row in range(row_idx + 1, min(row_idx + 6, ws.max_row + 1)):
                print(f"      {[c.value for c in ws[data_row]][:7]}")
            break
