import openpyxl
import sys

files = [
    '67 Y 68 CATALOGO.xlsx',
    'Stock-02-11-2025.xlsx'
]

for filepath in files:
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        print(f'\n📁 {filepath}')
        print('='*80)
        
        for row_idx in range(1, min(8, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(15, ws.max_column + 1)):
                val = ws.cell(row_idx, col_idx).value
                row_data.append(str(val)[:30] if val else '')
            print(f'Fila {row_idx}: {" | ".join(row_data)}')
        
        print(f'\nTotal filas: {ws.max_row}, Total columnas: {ws.max_column}')
        wb.close()
    except Exception as e:
        print(f'❌ Error en {filepath}: {e}')
